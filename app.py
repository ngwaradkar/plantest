import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import matplotlib.pyplot as plt
import importlib
import data_loader as dl
importlib.reload(dl)
import allocation_engine as ae
importlib.reload(ae)
import datetime
from streamlit_paste_button import paste_image_button

# IST (India Standard Time) Timezone offset = UTC + 5:30
IST_OFFSET = datetime.timedelta(hours=5, minutes=30)
IST_TZ = datetime.timezone(IST_OFFSET)

def get_ist_now():
    """Returns current datetime in Indian Standard Time (IST)."""
    return datetime.datetime.now(datetime.timezone.utc).astimezone(IST_TZ)

def format_ist_now(fmt="%d-%m-%Y %I:%M %p"):
    """Returns formatted current IST time string."""
    return get_ist_now().strftime(fmt)

def format_ist_nearest_15min():
    """Returns formatted IST time string rounded to the nearest 15-minute interval (e.g. 03.15 PM, 09.45 AM)."""
    dt = get_ist_now()
    rem = dt.minute % 15
    if rem >= 8:
        add_minutes = 15 - rem
    else:
        add_minutes = -rem
    rounded_dt = dt + datetime.timedelta(minutes=add_minutes)
    return rounded_dt.strftime("%I.%M %p")

def format_ist_mtime(filepath, fmt="%d-%m-%Y %I:%M %p"):
    """Returns formatted file modification time converted to IST."""
    try:
        mtime = os.path.getmtime(filepath)
        dt = datetime.datetime.fromtimestamp(mtime, tz=datetime.timezone.utc).astimezone(IST_TZ)
        return dt.strftime(fmt)
    except Exception:
        return None

# Initialize session state for theme preference from database
db_theme = '☀️ White Theme'
try:
    db_theme = dl.load_metadata('theme', '☀️ White Theme')
except Exception:
    pass

if 'theme' not in st.session_state:
    st.session_state.theme = db_theme

# Set page config
st.set_page_config(
    page_title="TCF1 & TCF2 VIN generation PPC Dashboard",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Render theme selector at top right
col_title_card, col_theme_select = st.columns([5.5, 1])
with col_theme_select:
    st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
    theme_option = st.selectbox(
        "🎨 Interface Theme",
        options=["☀️ White Theme", "🌙 Dark Theme"],
        index=0 if st.session_state.theme == '☀️ White Theme' else 1,
        key="theme_selection_key"
    )
    if theme_option != st.session_state.theme:
        st.session_state.theme = theme_option
        try:
            dl.save_metadata('theme', theme_option)
        except Exception:
            pass
        st.rerun()

is_dark = st.session_state.theme == "🌙 Dark Theme"

# Inject style blocks dynamically using CSS variables
if is_dark:
    theme_vars = """
    :root {
        --bg-primary: #0E1117;
        --bg-secondary: #1F2937;
        --card-bg: #161B22;
        --text-primary: #FAFAFA;
        --text-secondary: #D1D5DB;
        --border-color: #30363D;
        --accent-color: #4A9EFF;
        --accent-hover: #7BB4FF;
        --success-color: #10B981;
        --warning-color: #F59E0B;
        --danger-color: #EF4444;
        --hover-tint: rgba(74, 158, 255, 0.08);
        --card-ready-bg: #064E3B;
        --card-ready-text: #D1FAE5;
        --card-blocked-bg: #7F1D1D;
        --card-blocked-text: #FEE2E2;
        
        /* Map Streamlit native properties to match */
        --primary-color: var(--accent-color) !important;
        --background-color: var(--bg-primary) !important;
        --secondary-background-color: var(--bg-secondary) !important;
        --text-color: var(--text-primary) !important;
    }
    """
else:
    theme_vars = """
    :root {
        --bg-primary: #F9FAFB;
        --bg-secondary: #FFFFFF;
        --card-bg: #FFFFFF;
        --text-primary: #111827;
        --text-secondary: #374151;
        --border-color: #E5E7EB;
        --accent-color: #1D4ED8;
        --accent-hover: #1E3A8A;
        --success-color: #16A34A;
        --warning-color: #F59E0B;
        --danger-color: #DC2626;
        --hover-tint: rgba(29, 78, 216, 0.05);
        --card-ready-bg: #F0FAF4;
        --card-ready-text: #166534;
        --card-blocked-bg: #FFF5F5;
        --card-blocked-text: #B91C1C;
        
        /* Map Streamlit native properties to match */
        --primary-color: var(--accent-color) !important;
        --background-color: var(--bg-primary) !important;
        --secondary-background-color: var(--bg-secondary) !important;
        --text-color: var(--text-primary) !important;
    }
    """

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    {theme_vars}
    
    html, body, [class*="css"] {{
        font-family: 'Inter', sans-serif !important;
    }}
    
    /* Global App Background */
    .stApp {{
        background-color: var(--bg-primary) !important;
        color: var(--text-primary) !important;
    }}
    
    /* Completely hide sidebar and collapse button */
    [data-testid="stSidebar"], section[data-testid="stSidebar"] {{
        display: none !important;
        width: 0px !important;
    }}
    [data-testid="collapsedControl"] {{
        display: none !important;
    }}
    .stApp [data-testid="stHeader"] {{
        left: 0px !important;
        background-color: transparent !important;
    }}
    .main .block-container {{
        padding: 2.5rem 3rem !important;
        max-width: 100% !important;
    }}
    
    /* Premium Metric Cards with Hover lift animation */
    div[data-testid="stMetric"] {{
        background-color: var(--card-bg) !important;
        border: 1px solid var(--border-color) !important;
        padding: 1rem 1.25rem !important;
        border-radius: 12px !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03) !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        min-width: 0 !important;
        width: 100% !important;
    }}
    div[data-testid="stMetric"]:hover {{
        transform: translateY(-4px) !important;
        box-shadow: 0 12px 24px rgba(0, 0, 0, 0.08) !important;
        border-color: var(--accent-hover) !important;
    }}
    
    /* Colorful left border accents for each metric column */
    div[data-testid="column"]:nth-of-type(1) div[data-testid="stMetric"] {{
        border-left: 5px solid var(--accent-color) !important;
    }}
    div[data-testid="column"]:nth-of-type(2) div[data-testid="stMetric"] {{
        border-left: 5px solid #06B6D4 !important; /* Cyan */
    }}
    div[data-testid="column"]:nth-of-type(3) div[data-testid="stMetric"] {{
        border-left: 5px solid var(--success-color) !important; /* Success Green */
    }}
    div[data-testid="column"]:nth-of-type(4) div[data-testid="stMetric"] {{
        border-left: 5px solid var(--danger-color) !important; /* Danger Red */
    }}
    
    div[data-testid="stMetric"] label {{
        color: var(--text-secondary) !important;
        font-weight: 600 !important;
        font-size: 0.82rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.06em !important;
    }}
    
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {{
        color: var(--text-primary) !important;
        font-weight: 750 !important;
        font-size: 1.15rem !important;
        line-height: 1.3 !important;
        letter-spacing: -0.015em !important;
        word-break: normal !important;
        white-space: normal !important;
        overflow: visible !important;
    }}
    
    /* Segmented control for tabs */
    div[data-baseweb="tab-list"] {{
        background-color: var(--bg-secondary) !important;
        padding: 0.3rem !important;
        border-radius: 12px !important;
        gap: 6px !important;
        margin-bottom: 2rem !important;
        border: 1px solid var(--border-color) !important;
    }}
    button[data-baseweb="tab"] {{
        background-color: transparent !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.6rem 1.4rem !important;
        color: var(--text-secondary) !important;
        font-weight: 600 !important;
        font-size: 0.92rem !important;
        transition: all 0.2s ease !important;
    }}
    button[data-baseweb="tab"]:hover {{
        color: var(--text-primary) !important;
        background-color: var(--hover-tint) !important;
    }}
    button[data-baseweb="tab"][aria-selected="true"] {{
        background-color: var(--card-bg) !important;
        color: var(--accent-color) !important;
        border: 1px solid var(--border-color) !important;
        border-bottom: 3px solid var(--accent-color) !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.04) !important;
    }}
    div[data-baseweb="tab-border"] {{
        display: none !important;
    }}
    
    /* Expander styling */
    details[data-testid="stExpander"] {{
        background-color: var(--card-bg) !important;
        border: 1px solid var(--border-color) !important;
        border-radius: 12px !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.01) !important;
        margin-bottom: 1.5rem !important;
    }}
    details[data-testid="stExpander"] summary {{
        font-weight: 600 !important;
        color: var(--text-primary) !important;
    }}
    
    /* Form inputs and select boxes */
    div[data-baseweb="select"], div[data-baseweb="input"], input, textarea {{
        background-color: var(--card-bg) !important;
        color: var(--text-primary) !important;
        border-color: var(--border-color) !important;
        border-radius: 10px !important;
    }}
    div[data-baseweb="select"]:hover, div[data-baseweb="input"]:hover {{
        border-color: var(--accent-color) !important;
    }}
    
    /* Multiselect Tag Pill Fix */
    span[data-baseweb="tag"] {{
        background-color: #2563eb !important;
        background: #2563eb !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        padding-left: 8px !important;
        padding-right: 8px !important;
    }}
    span[data-baseweb="tag"] div, span[data-baseweb="tag"] span, span[data-baseweb="tag"] a {{
        background-color: transparent !important;
        background: transparent !important;
        color: #ffffff !important;
        border-radius: 0px !important;
    }}
    span[data-baseweb="tag"] svg {{
        fill: #ffffff !important;
        color: #ffffff !important;
    }}
    
    /* Button premium styling with micro-interaction hover/active states */
    button[kind="primary"] {{
        background-color: var(--accent-color) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.6rem 1.6rem !important;
        box-shadow: 0 4px 12px rgba(37, 99, 235, 0.15) !important;
        transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }}
    button[kind="primary"]:hover {{
        background-color: var(--accent-hover) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 16px rgba(37, 99, 235, 0.25) !important;
    }}
    button[kind="primary"]:active {{
        transform: translateY(0) !important;
    }}
    
    button[kind="secondary"] {{
        background-color: var(--card-bg) !important;
        color: var(--text-primary) !important;
        border: 1px solid var(--border-color) !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.6rem 1.6rem !important;
        transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }}
    button[kind="secondary"]:hover {{
        background-color: var(--hover-tint) !important;
        border-color: var(--accent-color) !important;
        transform: translateY(-2px) !important;
    }}
    button[kind="secondary"]:active {{
        transform: translateY(0) !important;
    }}
    
    /* Table / Dataframe premium look with soft shadow card mapping */
    div[data-testid="stDataFrame"] {{
        border-radius: 12px !important;
        border: 1px solid var(--border-color) !important;
        background-color: var(--card-bg) !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.02) !important;
        overflow: hidden !important;
    }}
    
    /* st.form container override */
    div[data-testid="stForm"] {{
        border: 1px solid var(--border-color) !important;
        border-radius: 12px !important;
        background-color: var(--card-bg) !important;
        padding: 1.5rem !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.02) !important;
    }}
    
    /* Headings */
    h1, h2, h3, h4, h5, h6 {{
        font-family: 'Inter', sans-serif !important;
        color: var(--text-primary) !important;
        letter-spacing: -0.02em !important;
        line-height: 1.3 !important;
    }}
    h1 {{ font-size: 2.2rem !important; font-weight: 850 !important; }}
    h2 {{ font-size: 1.8rem !important; font-weight: 750 !important; }}
    h3 {{ font-size: 1.45rem !important; font-weight: 700 !important; }}
    h4 {{ font-size: 1.2rem !important; font-weight: 650 !important; }}
    
    /* Horizontal rules */
    hr {{
        border-color: var(--border-color) !important;
    }}
    
    /* Custom file uploader border */
    section[data-testid="stFileUploader"] {{
        border: 2px dashed var(--border-color) !important;
        border-radius: 12px !important;
        background-color: var(--bg-secondary) !important;
    }}
</style>
""", unsafe_allow_html=True)

# Render the Title Card
with col_title_card:
    st.markdown("""
    <div style="background: linear-gradient(135deg, var(--accent-color) 0%, #06B6D4 100%); padding: 1.8rem 2.2rem; border-radius: 16px; margin-bottom: 2rem; color: white; box-shadow: 0 10px 25px rgba(0, 0, 0, 0.05);">
        <h1 style="color: white !important; font-weight: 850; margin: 0; font-size: 2.2rem; letter-spacing: -0.03em; font-family: 'Inter', sans-serif;">TCF1 & TCF2 VIN Generation PPC Dashboard</h1>
        <p style="color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0; font-size: 1.05rem; font-weight: 400; font-family: 'Inter', sans-serif;">Painted Body Storage (PBS) Buffer Allocation & Multi-Stage Material Availability Summary</p>
    </div>
    """, unsafe_allow_html=True)

# ----------------- SESSION STATE & INITIALIZATION -----------------
import datetime

# Pre-populated Engine Stocks (Defaulting to 0 as requested)
# Pre-populated Engine Stocks (ICE engines only; EV models managed separately)
engine_default_data = [
    {"TCF Line": "TCF1", "Engine Part No": "54850000PTP001", "Model": "Punch MT SA", "TA Code": "3302", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54850000PTP002", "Model": "Punch AMT SA", "TA Code": "3404", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54970000PTP002", "Model": "Punch TC MCE", "TA Code": "7349", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54970000PTP003", "Model": "Punch MCE MT", "TA Code": "3641", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54970000PTP004", "Model": "Punch MCE AMT", "TA Code": "3406", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54970000PTP005", "Model": "Punch MCE CNG MT", "TA Code": "3627", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF1", "Engine Part No": "54970000PTP031", "Model": "Punch MCE CNG AMT", "TA Code": "3403", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF2", "Engine Part No": "572900000118", "Model": "Harrier / Safari Diesel AT", "TA Code": "—", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF2", "Engine Part No": "572900000120", "Model": "Harrier / Safari Diesel MT", "TA Code": "—", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF2", "Engine Part No": "54780000PTP001", "Model": "Harrier / Safari Petrol TGDI MT", "TA Code": "—", "Clearance After 6:30AM": 0},
    {"TCF Line": "TCF2", "Engine Part No": "54780000PTP002", "Model": "Harrier / Safari Petrol TGDI AT", "TA Code": "—", "Clearance After 6:30AM": 0}
]

# Pre-populated Punch EV (Nova) Material Stocks (Defaulting to 0)
nova_default_data = [
    {"Model": "Nova", "Material": "Battery", "Clearance Qty": 0},
    {"Model": "Nova", "Material": "Combo", "Clearance Qty": 0},
    {"Model": "Nova", "Material": "Tube Frame(Craddle)", "Clearance Qty": 0},
    {"Model": "Nova", "Material": "Subframe", "Clearance Qty": 0},
    {"Model": "Nova", "Material": "RTB", "Clearance Qty": 0}
]

# Load last reset time from database metadata
db_last_reset = None
try:
    db_last_reset = dl.load_metadata('last_reset_time')
except Exception:
    pass
now = get_ist_now()

if 'last_reset_time' not in st.session_state:
    if db_last_reset:
        st.session_state.last_reset_time = datetime.datetime.fromisoformat(db_last_reset)
    else:
        st.session_state.last_reset_time = now
        try:
            dl.save_metadata('last_reset_time', now.isoformat())
        except Exception:
            pass

# Load engine stocks from DB or defaults (filter out legacy EV entries from engine_df)
if 'engine_df' not in st.session_state:
    db_engine_df = None
    try:
        db_engine_df = dl.load_engine_stocks_from_db()
    except Exception:
        pass
    if db_engine_df is not None:
        db_engine_df = db_engine_df[~db_engine_df['Engine Part No'].astype(str).str.strip().isin(['546816111212', '547380400103'])]
        st.session_state.engine_df = db_engine_df
    else:
        st.session_state.engine_df = pd.DataFrame(engine_default_data)
else:
    st.session_state.engine_df = st.session_state.engine_df[~st.session_state.engine_df['Engine Part No'].astype(str).str.strip().isin(['546816111212', '547380400103'])]

# Load Punch EV Nova materials from DB or defaults
if 'nova_materials_df' not in st.session_state:
    db_nova_df = None
    try:
        db_nova_df = dl.load_nova_stocks_from_db()
    except Exception:
        pass
    if db_nova_df is not None and not db_nova_df.empty:
        st.session_state.nova_materials_df = db_nova_df
    else:
        st.session_state.nova_materials_df = pd.DataFrame(nova_default_data)

# Load Model-Wise Shortages from DB or defaults
if 'model_shortages_df' not in st.session_state:
    db_ms_df = None
    try:
        db_ms_df = dl.load_model_shortages_from_db()
    except Exception:
        pass
    if db_ms_df is not None and not db_ms_df.empty:
        if 'Trims' not in db_ms_df.columns:
            db_ms_df['Trims'] = 'All Trims'
        st.session_state.model_shortages_df = db_ms_df
    else:
        st.session_state.model_shortages_df = pd.DataFrame(columns=['Model', 'Trims', 'Part Name', 'Clearance Qty'])

# Load Telegram bot credentials from DB metadata or pre-populated defaults
DEFAULT_TELEGRAM_TOKEN = "8817304754:AAGT6lfz17PE2BgSAMd10h6HIrHUFfU8pGk"
DEFAULT_TELEGRAM_CHAT_ID = "680536291"

if 'telegram_token' not in st.session_state:
    try:
        saved_tok = dl.load_metadata('telegram_token')
        st.session_state.telegram_token = saved_tok if saved_tok else DEFAULT_TELEGRAM_TOKEN
    except Exception:
        st.session_state.telegram_token = DEFAULT_TELEGRAM_TOKEN

if 'telegram_chat_id' not in st.session_state:
    try:
        saved_cid = dl.load_metadata('telegram_chat_id')
        st.session_state.telegram_chat_id = saved_cid if saved_cid else DEFAULT_TELEGRAM_CHAT_ID
    except Exception:
        st.session_state.telegram_chat_id = DEFAULT_TELEGRAM_CHAT_ID

# Auto reset at 6:30 AM is completely disabled. Data reset is only performed when manually triggered by planner via Control Center.

def get_demand_qty_for_model_trims(model_name, trims_str, tcf1_drops, tcf2_drops):
    """Calculates demand (VIN generation / DPT plan) for a given model and trim filter."""
    vgl_df = None
    if model_name in ['PUNCH', 'PUNCH.EV', 'ALTROZ']:
        vgl_df = tcf1_drops
    elif model_name in ['HARRIER / SAFARI', 'HARRIER', 'SAFARI', 'HARRIER.EV']:
        vgl_df = tcf2_drops

    if vgl_df is None or vgl_df.empty:
        dfs = [df for df in [tcf1_drops, tcf2_drops] if df is not None and not df.empty]
        vgl_df = pd.concat(dfs, ignore_index=True) if dfs else None

    if vgl_df is None or vgl_df.empty:
        return 0

    sub = vgl_df.copy()
    model_col = 'ProductFamily' if 'ProductFamily' in sub.columns else ('Model_Family' if 'Model_Family' in sub.columns else ('Model' if 'Model' in sub.columns else None))
    
    if model_col:
        m_str = str(model_name).upper()
        if 'HARRIER / SAFARI' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['HARRIER', 'SAFARI', 'GRAVITAS', 'Q5'])]
        elif 'HARRIER.EV' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['HARRIER.EV', 'ETURNA'])]
        elif 'PUNCH.EV' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['PUNCH.EV', 'NOVA'])]
        elif 'PUNCH' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['PUNCH', 'HORNBILL'])]
        else:
            sub = sub[sub[model_col].astype(str).str.upper().str.contains(m_str, na=False)]

    if trims_str and "All Trims" not in str(trims_str) and not sub.empty:
        trims_list = [t.strip().upper() for t in str(trims_str).split(',') if t.strip()]
        desc_col = 'SALES DESC' if 'SALES DESC' in sub.columns else ('SALES DESCRIPTION' if 'SALES DESCRIPTION' in sub.columns else ('MODEL' if 'MODEL' in sub.columns else None))
        if desc_col and trims_list:
            patterns = [re.escape(t) for t in trims_list]
            regex_pat = '|'.join(patterns)
            sub = sub[sub[desc_col].astype(str).str.upper().str.contains(regex_pat, na=False)]

    if 'VIN_Count' in sub.columns and pd.to_numeric(sub['VIN_Count'], errors='coerce').fillna(0).sum() > 0:
        return int(pd.to_numeric(sub['VIN_Count'], errors='coerce').fillna(0).sum())
    
    for col in ['TCF/-VIN', 'TCF2-VIN', 'TOTAL']:
        if col in sub.columns and pd.to_numeric(sub[col], errors='coerce').fillna(0).sum() > 0:
            return int(pd.to_numeric(sub[col], errors='coerce').fillna(0).sum())

    for col in ['TCF/-Plan', 'TCF2-Plan', 'Plan_Count']:
        if col in sub.columns and pd.to_numeric(sub[col], errors='coerce').fillna(0).sum() > 0:
            return int(pd.to_numeric(sub[col], errors='coerce').fillna(0).sum())

    return len(sub)

def get_backflushed_vin_count_for_model_trims(model_name, trims_str, tcf1_drops, tcf2_drops):
    """Calculates backflushed VIN drops generated today for a given model and trim filter."""
    vgl_df = None
    if model_name in ['PUNCH', 'PUNCH.EV', 'ALTROZ']:
        vgl_df = tcf1_drops
    elif model_name in ['HARRIER / SAFARI', 'HARRIER', 'SAFARI', 'HARRIER.EV']:
        vgl_df = tcf2_drops

    if vgl_df is None or vgl_df.empty:
        dfs = [df for df in [tcf1_drops, tcf2_drops] if df is not None and not df.empty]
        vgl_df = pd.concat(dfs, ignore_index=True) if dfs else None

    if vgl_df is None or vgl_df.empty:
        return 0

    sub = vgl_df.copy()
    model_col = 'ProductFamily' if 'ProductFamily' in sub.columns else ('Model_Family' if 'Model_Family' in sub.columns else ('Model' if 'Model' in sub.columns else None))
    
    if model_col:
        m_str = str(model_name).upper()
        if 'HARRIER / SAFARI' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['HARRIER', 'SAFARI', 'GRAVITAS', 'Q5'])]
        elif 'HARRIER.EV' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['HARRIER.EV', 'ETURNA'])]
        elif 'PUNCH.EV' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['PUNCH.EV', 'NOVA'])]
        elif 'PUNCH' in m_str:
            sub = sub[sub[model_col].astype(str).str.upper().isin(['PUNCH', 'HORNBILL'])]
        else:
            sub = sub[sub[model_col].astype(str).str.upper().str.contains(m_str, na=False)]

    if trims_str and "All Trims" not in str(trims_str) and not sub.empty:
        trims_list = [t.strip().upper() for t in str(trims_str).split(',') if t.strip()]
        desc_col = 'SALES DESC' if 'SALES DESC' in sub.columns else ('SALES DESCRIPTION' if 'SALES DESCRIPTION' in sub.columns else ('MODEL' if 'MODEL' in sub.columns else None))
        if desc_col and trims_list:
            patterns = [re.escape(t) for t in trims_list]
            regex_pat = '|'.join(patterns)
            sub = sub[sub[desc_col].astype(str).str.upper().str.contains(regex_pat, na=False)]

    for col in ['VIN_Count', 'TCF/-VIN', 'TCF2-VIN', 'TOTAL']:
        if col in sub.columns and pd.to_numeric(sub[col], errors='coerce').fillna(0).sum() > 0:
            return int(pd.to_numeric(sub[col], errors='coerce').fillna(0).sum())

    return 0

def evaluate_all_clearance_shortage_alerts(model_shortages_df, nova_materials_df, engine_df, tcf1_drops, tcf2_drops):
    """Evaluates all starting clearance stocks against VIN/DPT demand and returns active shortage alerts."""
    alerts = []
    
    # 1. Model-Wise Shortages
    if model_shortages_df is not None and not model_shortages_df.empty:
        for idx, row in model_shortages_df.iterrows():
            m_name = str(row['Model']).strip()
            m_trims = str(row.get('Trims', 'All Trims')).strip()
            p_name = str(row['Part Name']).strip()
            c_qty = int(row['Clearance Qty'])
            
            vins_today = get_backflushed_vin_count_for_model_trims(m_name, m_trims, tcf1_drops, tcf2_drops)
            true_stock = max(0, c_qty - vins_today)
            d_qty = get_demand_qty_for_model_trims(m_name, m_trims, tcf1_drops, tcf2_drops)
            
            if true_stock <= 0 or c_qty < d_qty:
                deficit = max(d_qty - c_qty, vins_today - c_qty, 0)
                alerts.append({
                    'Category': 'Model-Wise Shortage',
                    'Model': m_name,
                    'Trims': m_trims,
                    'Part Name': p_name,
                    'Clearance Qty': c_qty,
                    'VINs Today': vins_today,
                    'True Stock': true_stock,
                    'Demand Qty': d_qty,
                    'Shortage Qty': deficit,
                    'Message': f"{m_name} [{m_trims}] - {p_name}: Clearance {c_qty} - Produced VINs {vins_today} = {true_stock} Available Stock (Deficit: -{deficit} units)"
                })

    # 2. Punch EV (Nova) Materials
    if nova_materials_df is not None and not nova_materials_df.empty:
        nova_vin_qty = get_demand_qty_for_model_trims("PUNCH.EV", "All Trims", tcf1_drops, tcf2_drops)
        for idx, row in nova_materials_df.iterrows():
            m_name = str(row['Material']).strip().replace('Craddle', 'Cradle')
            c_qty = int(row['Clearance Qty'])
            if c_qty < nova_vin_qty:
                deficit = nova_vin_qty - c_qty
                alerts.append({
                    'Category': 'Punch EV Material',
                    'Model': 'PUNCH.EV',
                    'Trims': 'All Trims',
                    'Part Name': m_name,
                    'Clearance Qty': c_qty,
                    'Demand Qty': nova_vin_qty,
                    'Shortage Qty': deficit,
                    'Message': f"PUNCH.EV - {m_name}: Clearance Stock {c_qty} vs VIN Demand {nova_vin_qty} (Shortage: -{deficit} units)"
                })
                
    return alerts

# Load active directory selection from database metadata
db_data_source = 'Root Directory (Production)'
try:
    db_data_source = dl.load_metadata('data_source_dir', 'Root Directory (Production)')
except Exception:
    pass

if 'data_source_dir' not in st.session_state:
    st.session_state.data_source_dir = db_data_source

# Auto-detect file mappings in workspace
workspace_dir = os.path.dirname(os.path.abspath(__file__))

if st.session_state.data_source_dir == 'TEST Directory (Sample Data)':
    active_dir = os.path.join(workspace_dir, "TEST")
else:
    active_dir = workspace_dir

# Scan active folder for default files dynamically
detected_files = dl.detect_and_classify_files(active_dir)

all_categories = [
    'BOM', 
    'FLOAT_REPORT', 
    'FLOAT_PAINT_SUMMARY',
    'SHOP_WISE_REPORT',
    'TCF1_VGL', 'TCF2_VGL',
    'TCF1_ALTROZ_COCKPIT_STOCK', 'TCF1_NOVA_COCKPIT_STOCK', 'TCF2_COCKPIT_STOCK',
    'TCF1_WIRING_STOCK', 'TCF2_WIRING_STOCK'
]

loaded_data = {}

# ----------------- CONTROL CENTER (UPLOADS & ENGINE STOCKS) -----------------
# We put the control panel in a clean main-body expander.
# The expander starts collapsed if the core files are already auto-loaded.
db_bom_exists = dl.load_bom_from_db() is not None
default_bom_path = detected_files.get('BOM')
default_float_path = detected_files.get('FLOAT_REPORT') or detected_files.get('FLOAT_PAINT_SUMMARY')
default_core_available = (db_bom_exists or default_bom_path is not None) and default_float_path is not None

config_expander = st.expander(
    "⚙️ Control Panel: File Uploads & Engine Starting Stocks (Click to Expand/Collapse)",
    expanded=not default_core_available and not st.session_state.get('run_report', False)
)

with config_expander:
    col_upload, col_engine = st.columns([1.15, 1.2])
    
    with col_upload:
        with st.container(border=True):
            st.markdown("#### 📤 Upload Raw Plant Files")
            
            uploaded_files = st.file_uploader(
                "Upload plant reports to replace existing ones",
                accept_multiple_files=True,
                help="Upload raw spreadsheets (Float, Wiring, Cockpit, or VGL). They will automatically replace older files on disk."
            )
            
            # Process uploads immediately, saving to session state buffers and optionally to disk
            uploaded_ids = [f"{f.name}_{f.size}" for f in uploaded_files] if uploaded_files else []
            last_processed_ids = st.session_state.get("last_processed_upload_ids", [])
            
            if not uploaded_files:
                st.session_state.last_processed_upload_ids = []
                
            if uploaded_files and uploaded_ids != last_processed_ids:
                uploaded_mappings = dl.classify_files(uploaded_files)
                replaced_any = False
                upload_ts = format_ist_now("%d-%m-%Y %I:%M %p")
                for category, uploaded_file in uploaded_mappings.items():
                    mem_buf = io.BytesIO(uploaded_file.getvalue())
                    mem_buf.name = uploaded_file.name
                    st.session_state[f"buffer_{category}"] = mem_buf
                    st.session_state[f"upload_time_{category}"] = upload_ts
                    try:
                        dl.save_metadata(f"upload_time_{category}", upload_ts)
                    except Exception:
                        pass
                    
                    if category == 'BOM':
                        try:
                            parsed_bom = dl.load_bom(uploaded_file)
                            try:
                                dl.save_bom_to_db(parsed_bom)
                                dl.save_metadata(f"uploaded_{category}", uploaded_file.name)
                            except Exception:
                                pass
                            st.toast("💾 Master BOM replaced!", icon="💾")
                            replaced_any = True
                        except Exception as e:
                            st.error(f"Failed to parse uploaded BOM: {e}")
                    else:
                        # Try to save file to disk (succeeds locally, fails safely in read-only cloud)
                        try:
                            old_path = detected_files.get(category)
                            if old_path and os.path.exists(old_path):
                                os.remove(old_path)
                            
                            new_path = os.path.join(active_dir, uploaded_file.name)
                            with open(new_path, "wb") as f:
                                f.write(uploaded_file.getbuffer())
                            dl.save_metadata(f"uploaded_{category}", uploaded_file.name)
                        except Exception:
                            pass
                        
                        st.toast(f"✅ Loaded {category.replace('_',' ')}: {uploaded_file.name}", icon="✅")
                        replaced_any = True
                
                # Record that we processed these files
                st.session_state.last_processed_upload_ids = uploaded_ids
                if replaced_any:
                    detected_files = dl.detect_and_classify_files(active_dir)
                    st.session_state.run_report = False
                    st.rerun()
            
            st.markdown("##### Loaded Files Status")
            
            # Check if database has BOM
            db_bom_df = None
            try:
                db_bom_df = dl.load_bom_from_db()
            except Exception:
                pass
            
            status_items = []
            seq_num = 1
            
            # Populate loaded_data
            for category in all_categories:
                detected_path = detected_files.get(category)
                is_default_exists = detected_path is not None and os.path.exists(detected_path)
                
                # Check session state buffer first
                in_mem_buffer = st.session_state.get(f"buffer_{category}")
                
                if category == 'BOM':
                    if in_mem_buffer is not None:
                        loaded_data[category] = in_mem_buffer
                    elif db_bom_df is not None:
                        loaded_data[category] = 'DATABASE'
                    elif is_default_exists:
                        loaded_data[category] = detected_path
                    continue
                    
                display_name = category.replace('_',' ').replace('VGL', 'VIN Generation')
                if category in ['TCF1_VGL', 'TCF2_VGL']:
                    display_name = f"DPT {display_name}"

                upload_time = st.session_state.get(f"upload_time_{category}")
                if not upload_time:
                    try:
                        upload_time = dl.load_metadata(f"upload_time_{category}")
                    except Exception:
                        upload_time = None

                if not upload_time and is_default_exists:
                    upload_time = format_ist_mtime(detected_path)

                time_label = f"({upload_time})" if upload_time else "(No upload time)"

                if in_mem_buffer is not None:
                    loaded_data[category] = in_mem_buffer
                    status_icon = "🟢 Uploaded"
                    source_label = time_label
                elif is_default_exists:
                    loaded_data[category] = detected_path
                    status_icon = "🟢 Uploaded"
                    source_label = time_label
                else:
                    status_icon = "🔴 Missing"
                    source_label = "(Pending for upload)"
                    
                status_items.append((seq_num, display_name, status_icon, source_label))
                seq_num += 1

            is_dark_theme = st.session_state.get('theme', '☀️ White Theme') == '🌙 Dark Theme'
            sub_text_color = "#94A3B8" if is_dark_theme else "#64748B"
            label_color = "#FAFAFA" if is_dark_theme else "#1E293B"

            status_html = "<div style='font-family: \"Inter\", sans-serif; font-size: 13px; line-height: 1.8; margin-top: 8px;'>"
            for num, name, icon, src in status_items:
                status_html += f'<div style="display: flex; align-items: center; margin-bottom: 4px;"><span style="font-weight: 700; width: 255px; color: {label_color}; display: inline-block; flex-shrink: 0;">{num}. {name}:</span><span style="margin-right: 12px; flex-shrink: 0;">{icon}</span><span style="color: {sub_text_color}; font-size: 12px; word-break: break-all;">{src}</span></div>'
            status_html += "</div>"
            st.markdown(status_html, unsafe_allow_html=True)

        st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("#### ⚡ Punch EV (Nova) Component Starting Stocks")
            st.markdown("<small style='color:#8896AB'>Enter clearance counts for 5 critical Punch EV materials (Defaults: 0)</small>", unsafe_allow_html=True)
            
            nova_vals = {}
            if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None:
                for idx, r_nova in st.session_state.nova_materials_df.iterrows():
                    nova_vals[str(r_nova['Material']).strip()] = int(r_nova['Clearance Qty'])

            materials_list = ["Battery", "Combo", "Tube Frame(Craddle)", "Subframe", "RTB"]

            nova_cols = st.columns(5)
            new_nova_input_vals = {}
            for idx, mat in enumerate(materials_list):
                current_val = nova_vals.get(mat, 0)
                with nova_cols[idx]:
                    st.markdown(f"<div style='font-weight: 700; font-size: 12px; color: {label_color}; margin-bottom: 3px; word-break: break-word;'>{mat}</div>", unsafe_allow_html=True)
                    new_nova_input_vals[mat] = st.number_input(
                        label=mat,
                        min_value=0,
                        value=int(current_val),
                        step=1,
                        key=f"nova_input_{idx}",
                        label_visibility="collapsed"
                    )

            st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)
            if st.button("💾 Save Punch EV Clearances", type="primary", use_container_width=True, key="save_nova_btn"):
                updated_nova_list = []
                for mat in materials_list:
                    updated_nova_list.append({"Model": "Nova", "Material": mat, "Clearance Qty": new_nova_input_vals[mat]})
                st.session_state.nova_materials_df = pd.DataFrame(updated_nova_list)
                try:
                    dl.save_nova_stocks_to_db(st.session_state.nova_materials_df)
                except Exception:
                    pass
                st.toast("💾 Punch EV starting clearance stocks saved successfully!", icon="💾")
                st.session_state.run_report = True
                st.rerun()

        st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("#### 📦 Model Wise Shortage")
            st.markdown("<small style='color:#8896AB'>Configure shortages by model, trim, part name, and clearance quantity.</small>", unsafe_allow_html=True)
            
            # Trim Options Mapping per Model
            trim_options_map = {
                "HARRIER / SAFARI": ["ACC", "ADV", "PUR", "SMT", "FRL", "FRLR", "TGDI", "CNG"],
                "HARRIER.EV": ["ADV", "EMP", "FEA", "FEA+"],
                "PUNCH": ["SMART", "PURE", "PURE +", "PURE + S", "ADVT", "ADVT S", "ACCOMP", "ACCOMP + S", "CREATIVE / LUX", "CNG", "TGDI"],
                "PUNCH.EV": ["SMT", "SMT+", "ADV", "EMP", "EMP+ S"]
            }

            def pause_calculation_on_edit():
                st.session_state.run_report = False

            col_m, col_t = st.columns([1.2, 1.8])
            with col_m:
                selected_ms_model = st.selectbox(
                    "Select Model",
                    options=list(trim_options_map.keys()),
                    key="ms_model_select",
                    on_change=pause_calculation_on_edit
                )
            with col_t:
                available_trims = trim_options_map.get(selected_ms_model, [])
                selected_trims = st.multiselect(
                    "Select Trim(s)",
                    options=available_trims,
                    default=[],
                    placeholder="Select trim(s)... (leave empty for All Trims)",
                    key=f"ms_trim_select_{selected_ms_model}",
                    on_change=pause_calculation_on_edit
                )

            col_p, col_q = st.columns([2.0, 1.0])
            with col_p:
                input_ms_part_name = st.text_input(
                    "Part Name",
                    placeholder="Enter part name manually...",
                    key="ms_part_input",
                    on_change=pause_calculation_on_edit
                )
            with col_q:
                input_ms_clearance_qty = st.number_input(
                    "Clearance Qty",
                    min_value=0,
                    value=0,
                    step=1,
                    key="ms_qty_input",
                    on_change=pause_calculation_on_edit
                )
            
            st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)
            if st.button("➕ Add Shortage Item", type="primary", use_container_width=True, key="add_ms_btn"):
                if not input_ms_part_name.strip():
                    st.error("Please enter a Part Name before adding.")
                else:
                    trims_str = ", ".join(selected_trims) if selected_trims else "All Trims"
                    new_ms_entry = {
                        "Model": selected_ms_model,
                        "Trims": trims_str,
                        "Part Name": input_ms_part_name.strip(),
                        "Clearance Qty": int(input_ms_clearance_qty)
                    }
                    if 'model_shortages_df' not in st.session_state or st.session_state.model_shortages_df is None or st.session_state.model_shortages_df.empty:
                        st.session_state.model_shortages_df = pd.DataFrame([new_ms_entry])
                    else:
                        df_ms = st.session_state.model_shortages_df.copy()
                        mask = (df_ms['Model'] == selected_ms_model) & (df_ms['Trims'] == trims_str) & (df_ms['Part Name'].str.lower() == input_ms_part_name.strip().lower())
                        if mask.any():
                            df_ms.loc[mask, 'Clearance Qty'] = int(input_ms_clearance_qty)
                        else:
                            df_ms = pd.concat([df_ms, pd.DataFrame([new_ms_entry])], ignore_index=True)
                        st.session_state.model_shortages_df = df_ms
                    
                    try:
                        dl.save_model_shortages_to_db(st.session_state.model_shortages_df)
                    except Exception:
                        pass
                    st.toast(f"✅ Added {input_ms_part_name.strip()} for {selected_ms_model} [{trims_str}] (Qty: {input_ms_clearance_qty})", icon="✅")
                    st.rerun()

            # Display existing model-wise shortages table if present
            if 'model_shortages_df' in st.session_state and st.session_state.model_shortages_df is not None and not st.session_state.model_shortages_df.empty:
                st.markdown("<div style='height: 6px;'></div>", unsafe_allow_html=True)
                st.markdown("##### Current Model-Wise Shortages")
                
                ms_disp_df = st.session_state.model_shortages_df.copy()
                
                for idx_ms, row_ms in ms_disp_df.iterrows():
                    c1, c2, c3, c4, c5, c6 = st.columns([1.1, 1.1, 1.4, 0.7, 1.5, 0.4])
                    with c1:
                        st.markdown(f"**{row_ms['Model']}**")
                    with c2:
                        st.markdown(f"`{row_ms.get('Trims', 'All Trims')}`")
                    with c3:
                        st.markdown(f"{row_ms['Part Name']}")
                    with c4:
                        st.markdown(f"Clearance: `{row_ms['Clearance Qty']}`")
                    with c5:
                        c_qty = int(row_ms['Clearance Qty'])
                        tcf1_tmp = locals().get('tcf1_drops', None)
                        tcf2_tmp = locals().get('tcf2_drops', None)
                        vins_today = get_backflushed_vin_count_for_model_trims(row_ms['Model'], row_ms.get('Trims', 'All Trims'), tcf1_tmp, tcf2_tmp)
                        true_buf = max(0, c_qty - vins_today)
                        if true_buf <= 0:
                            st.markdown(f"<span style='background:#FEE2E2; color:#DC2626; border: 1px solid #FCA5A5; padding:3px 8px; border-radius:6px; font-weight:700; font-size:11px;'>🚨 SHORTAGE: 0 Buffer (VIN Gen: {vins_today})</span>", unsafe_allow_html=True)
                        else:
                            st.markdown(f"<span style='background:#DCFCE7; color:#166534; border: 1px solid #86EFAC; padding:3px 8px; border-radius:6px; font-weight:600; font-size:11px;'>🟢 OK: {true_buf} Buffer (VIN Gen: {vins_today})</span>", unsafe_allow_html=True)
                    with c6:
                        if st.button("🗑️", key=f"del_ms_{idx_ms}", help="Delete item"):
                            st.session_state.model_shortages_df = st.session_state.model_shortages_df.drop(idx_ms).reset_index(drop=True)
                            try:
                                dl.save_model_shortages_to_db(st.session_state.model_shortages_df)
                            except Exception:
                                pass
                            st.toast("Item removed", icon="🗑️")
                            st.rerun()

                if st.button("🗑️ Clear All Model Shortages", key="clear_all_ms_btn", use_container_width=True):
                    st.session_state.model_shortages_df = pd.DataFrame(columns=['Model', 'Trims', 'Part Name', 'Clearance Qty'])
                    try:
                        dl.save_model_shortages_to_db(st.session_state.model_shortages_df)
                    except Exception:
                        pass
                    st.toast("All model shortages cleared", icon="🗑️")
                    st.rerun()
            
    with col_engine:
        with st.container(border=True):
            st.markdown("#### ⚙️ ICE Engine Starting Stocks")
            st.markdown("<small style='color:#8896AB'>Enter clearance counts for ICE engines (Defaults: 0)</small>", unsafe_allow_html=True)
            
            eng_vals = {}
            if 'engine_df' in st.session_state and st.session_state.engine_df is not None:
                for idx, r_eng in st.session_state.engine_df.iterrows():
                    p_no = str(r_eng['Engine Part No']).strip()
                    eng_vals[p_no] = int(r_eng.get('Clearance After 6:30AM', 0))

            tcf1_engine_list = [
                ("54850000PTP001", "Punch MT SA"),
                ("54850000PTP002", "Punch AMT SA"),
                ("54970000PTP002", "Punch TC MCE"),
                ("54970000PTP003", "Punch MCE MT"),
                ("54970000PTP004", "Punch MCE AMT"),
                ("54970000PTP005", "Punch MCE CNG MT"),
                ("54970000PTP031", "Punch MCE CNG AMT")
            ]

            tcf2_engine_list = [
                ("572900000118", "Harrier / Safari Diesel AT"),
                ("572900000120", "Harrier / Safari Diesel MT"),
                ("54780000PTP001", "Harrier / Safari TGDI MT"),
                ("54780000PTP002", "Harrier / Safari TGDI AT")
            ]

            new_eng_input_vals = {}

            st.markdown("<div style='font-weight: 700; font-size: 13px; color: #3B82F6; margin-top: 6px; margin-bottom: 6px;'>TCF1 Engines (Punch ICE)</div>", unsafe_allow_html=True)
            tcf1_cols = st.columns(4)
            for idx, (p_no, model_name) in enumerate(tcf1_engine_list):
                c_idx = idx % 4
                curr_val = eng_vals.get(p_no, 0)
                with tcf1_cols[c_idx]:
                    st.markdown(f"<div style='font-weight: 700; font-size: 12px; color: {label_color}; margin-bottom: 3px; word-break: break-word;'>{model_name}</div>", unsafe_allow_html=True)
                    new_eng_input_vals[p_no] = st.number_input(
                        label=model_name,
                        min_value=0,
                        value=int(curr_val),
                        step=1,
                        key=f"eng_input_{p_no}",
                        label_visibility="collapsed"
                    )

            st.markdown("<div style='font-weight: 700; font-size: 13px; color: #3B82F6; margin-top: 10px; margin-bottom: 6px;'>TCF2 Engines (Harrier / Safari ICE)</div>", unsafe_allow_html=True)
            tcf2_cols = st.columns(4)
            for idx, (p_no, model_name) in enumerate(tcf2_engine_list):
                curr_val = eng_vals.get(p_no, 0)
                with tcf2_cols[idx]:
                    st.markdown(f"<div style='font-weight: 700; font-size: 12px; color: {label_color}; margin-bottom: 3px; word-break: break-word;'>{model_name}</div>", unsafe_allow_html=True)
                    new_eng_input_vals[p_no] = st.number_input(
                        label=model_name,
                        min_value=0,
                        value=int(curr_val),
                        step=1,
                        key=f"eng_input_{p_no}",
                        label_visibility="collapsed"
                    )

            st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)
            if st.button("💾 Save Engine Starting Stocks", type="primary", use_container_width=True, key="save_engine_btn"):
                for idx, r_eng in st.session_state.engine_df.iterrows():
                    p_no = str(r_eng['Engine Part No']).strip()
                    if p_no in new_eng_input_vals:
                        st.session_state.engine_df.at[idx, 'Clearance After 6:30AM'] = new_eng_input_vals[p_no]
                try:
                    dl.save_engine_stocks_to_db(st.session_state.engine_df)
                except Exception:
                    pass
                st.toast("💾 ICE Engine starting clearance stocks saved successfully!", icon="💾")
                st.session_state.run_report = True
                st.rerun()

        st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
        with st.container(border=True):
            reset_card_bg = "rgba(225, 29, 72, 0.05)" if is_dark_theme else "#FFF1F2"
            reset_border = "rgba(225, 29, 72, 0.25)" if is_dark_theme else "#FECDD3"
            reset_text_color = "#FDA4AF" if is_dark_theme else "#9F1239"
            reset_subtext = "#94A3B8" if is_dark_theme else "#64748B"

            st.markdown(f"""
            <div style="background-color: {reset_card_bg}; border: 1px solid {reset_border}; border-radius: 10px; padding: 12px 16px; margin-bottom: 8px;">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px;">
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <span style="font-size: 15px;">🔄</span>
                        <span style="font-weight: 700; font-size: 13px; color: {reset_text_color}; font-family: 'Inter', sans-serif;">Shift Start Stock Reset</span>
                    </div>
                    <span style="font-size: 11px; background: rgba(225, 29, 72, 0.12); color: {reset_text_color}; padding: 2px 8px; border-radius: 12px; font-weight: 600;">Manual Trigger</span>
                </div>
                <div style="font-size: 11px; color: {reset_subtext}; margin-bottom: 8px; line-height: 1.4;">
                    Resets all starting clearance quantities to 0 and clears daily uploaded report cache.
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("🔄 Reset Clearances to 0 (Shift Start)", type="primary", use_container_width=True, key="reset_clearances_btn"):
                st.session_state.engine_df = pd.DataFrame(engine_default_data)
                st.session_state.nova_materials_df = pd.DataFrame(nova_default_data)
                st.session_state.model_shortages_df = pd.DataFrame(columns=['Model', 'Part Name', 'Clearance Qty'])
                try:
                    dl.save_engine_stocks_to_db(st.session_state.engine_df)
                    dl.save_nova_stocks_to_db(st.session_state.nova_materials_df)
                    dl.save_model_shortages_to_db(st.session_state.model_shortages_df)
                except Exception:
                    pass
                now_time = get_ist_now()
                st.session_state.last_reset_time = now_time
                try:
                    dl.save_metadata('last_reset_time', now_time.isoformat())
                except Exception:
                    pass
                # Clear uploaded file registry in metadata and session state buffers (except BOM which is constant)
                for cat in all_categories:
                    if cat == 'BOM':
                        continue  # Master BOM details remain constant all the time
                    try:
                        dl.save_metadata(f"uploaded_{cat}", "")
                        dl.save_metadata(f"upload_time_{cat}", "")
                    except Exception:
                        pass
                    if f"buffer_{cat}" in st.session_state:
                        del st.session_state[f"buffer_{cat}"]
                    if f"upload_time_{cat}" in st.session_state:
                        del st.session_state[f"upload_time_{cat}"]
                st.session_state.run_report = False
                st.toast("🔄 Engine clearances and daily report upload registries reset (Master BOM preserved)!", icon="🔄")
                st.rerun()

core_available = 'BOM' in loaded_data and ('FLOAT_REPORT' in loaded_data or 'FLOAT_PAINT_SUMMARY' in loaded_data)

if not core_available:
    st.warning("⚠️ Please ensure the **Master BOM** and **PPC Float Report** (or **Paint Float Summary Report**) are available in the database/workspace or uploaded above to run the dashboard.")
    st.stop()

# ----------------- GENERATE REPORT CONTROL -----------------
if 'run_report' not in st.session_state:
    st.session_state.run_report = False

st.markdown("---")
col_gen1, col_gen2 = st.columns([1.5, 3.5])
with col_gen1:
    btn_label = "🚀 Generate Report" if not st.session_state.run_report else "🔄 Update Dashboard Data"
    if st.button(btn_label, type="primary", use_container_width=True, key="btn_generate_report_control"):
        st.session_state.run_report = True
        st.rerun()

with col_gen2:
    if not st.session_state.run_report:
        st.info("💡 Files or engine data have been updated. Click **'🚀 Generate Report'** on the left to run calculations.")
    else:
        st.caption("✅ Report is generated. Uploading new files or updating engine clearances will pause auto-runs until you click **'Update Dashboard Data'**.")

if not st.session_state.run_report:
    st.stop()

# ----------------- PARSING & CALCULATING DATA -----------------
try:
    with st.spinner("⏳ Parsing plant reports and running allocation calculations..."):
        # 1. Load BOM
        if loaded_data['BOM'] == 'DATABASE':
            bom_df = dl.load_bom_from_db()
        else:
            bom_df = dl.load_bom(loaded_data['BOM'])
            try:
                dl.save_bom_to_db(bom_df)
            except Exception:
                pass
        
        # 2. Load Float Reports
        float_df = dl.load_float_report(loaded_data['FLOAT_REPORT']) if 'FLOAT_REPORT' in loaded_data else None
        paint_summary_dict = dl.load_paint_summary_report(loaded_data['FLOAT_PAINT_SUMMARY']) if 'FLOAT_PAINT_SUMMARY' in loaded_data else None
        paint_summary_vc_dict = dl.load_paint_summary_by_vc(loaded_data['FLOAT_PAINT_SUMMARY']) if 'FLOAT_PAINT_SUMMARY' in loaded_data else None
        
        # 3. Load VGL drops & Shop-Wise Report
        tcf1_drops = dl.load_vgl(loaded_data['TCF1_VGL']) if 'TCF1_VGL' in loaded_data else None
        tcf2_drops = dl.load_vgl(loaded_data['TCF2_VGL']) if 'TCF2_VGL' in loaded_data else None
        
        shop_totals = None
        shop_vehicles_df = None
        shop_ta_df = None
        if 'SHOP_WISE_REPORT' in loaded_data:
            shop_totals, shop_vehicles_df, shop_ta_df, shop_debug_info = dl.load_shop_wise_report(
                loaded_data['SHOP_WISE_REPORT'], return_debug=True
            )
            if not shop_debug_info.get('success'):
                with st.expander("⚠️ Shop Wise Report failed to load — click for details", expanded=True):
                    st.error(shop_debug_info.get('reason', 'Unknown error while parsing the Shop Wise Report.'))
                    attempts = shop_debug_info.get('attempts') or []
                    if attempts:
                        st.markdown("**Engines/paths tried and why each failed:**")
                        for a in attempts:
                            st.markdown(f"- `{a['stage']}`: {a['error']}")
        
        # 4. Parse stocks
        # Wiring Stock
        tcf1_wiring_start = None
        tcf1_wiring_vc_map = {}
        if 'TCF1_WIRING_STOCK' in loaded_data:
            tcf1_wiring_start, tcf1_wiring_vc_map = dl.load_stock_grouped(
                loaded_data['TCF1_WIRING_STOCK'],
                sheet_name='Coverage file 6.30 AM New',
                vc_col_idx=2, part_col_idx=3, qty_col_idx=9
            )
            
        tcf2_wiring_start = None
        tcf2_wiring_vc_map = {}
        if 'TCF2_WIRING_STOCK' in loaded_data:
            tcf2_wiring_start, tcf2_wiring_vc_map = dl.load_stock_grouped(
                loaded_data['TCF2_WIRING_STOCK'],
                sheet_name='coverage file 6.30 pm',
                vc_col_idx=1, part_col_idx=2, qty_col_idx=9
            )
            
        # Cockpit Stock
        # Combine Altroz and Nova cockpits for TCF1
        tcf1_cockpit_start = None
        tcf1_cockpit_vc_map = {}
        
        if 'TCF1_ALTROZ_COCKPIT_STOCK' in loaded_data:
            if tcf1_cockpit_start is None:
                tcf1_cockpit_start = {}
            altroz_start, altroz_map = dl.load_stock_grouped(
                loaded_data['TCF1_ALTROZ_COCKPIT_STOCK'],
                sheet_name='Fresh VIN PPC',
                vc_col_idx=4, part_col_idx=3, qty_col_idx=12, skip_rows=3
            )
            tcf1_cockpit_start.update(altroz_start)
            tcf1_cockpit_vc_map.update(altroz_map)
            
        if 'TCF1_NOVA_COCKPIT_STOCK' in loaded_data:
            if tcf1_cockpit_start is None:
                tcf1_cockpit_start = {}
            nova_start, nova_map = dl.load_stock_grouped(
                loaded_data['TCF1_NOVA_COCKPIT_STOCK'],
                sheet_name='Fresh VIN PPC',
                vc_col_idx=1, part_col_idx=0, qty_col_idx=9, skip_rows=3
            )
            tcf1_cockpit_start.update(nova_start)
            tcf1_cockpit_vc_map.update(nova_map)
            
        tcf2_cockpit_start = None
        tcf2_cockpit_vc_map = {}
        if 'TCF2_COCKPIT_STOCK' in loaded_data:
            tcf2_cockpit_start, tcf2_cockpit_vc_map = dl.load_stock_grouped(
                loaded_data['TCF2_COCKPIT_STOCK'],
                sheet_name='Fresh vin PPC',
                vc_col_idx=3, part_col_idx=1, qty_col_idx=9, skip_rows=3
            )
            
        # Load Engine Stock from data_editor
        engine_stocks_tcf1 = {}
        engine_stocks_tcf2 = {}
        for idx, row in st.session_state.engine_df.iterrows():
            part = str(row['Engine Part No']).strip()
            qty = int(row['Clearance After 6:30AM'])
            if row['TCF Line'] == 'TCF1':
                engine_stocks_tcf1[part] = qty
            else:
                engine_stocks_tcf2[part] = qty

        # Add Punch EV Nova starting clearance (min across 5 materials)
        if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None and not st.session_state.nova_materials_df.empty:
            engine_stocks_tcf1['546816111212'] = int(st.session_state.nova_materials_df['Clearance Qty'].min())
        else:
            engine_stocks_tcf1['546816111212'] = 182

        # Add Harrier EV starting clearance (default 160)
        engine_stocks_tcf2['547380400103'] = 160

        # ----------------- BACKFLUSH LOGIC (calculate true stock) -----------------
        # TCF1 Backflush
        true_engine_tcf1, eng_cons_tcf1, eng_warn_tcf1 = ae.calculate_true_stock(engine_stocks_tcf1, tcf1_drops, bom_df, 'Engine')
        true_cockpit_tcf1, ck_cons_tcf1, ck_warn_tcf1 = ae.calculate_true_stock(tcf1_cockpit_start, tcf1_drops, bom_df, 'Cockpit')
        true_wiring_tcf1, wh_cons_tcf1, wh_warn_tcf1 = ae.calculate_true_stock(tcf1_wiring_start, tcf1_drops, bom_df, 'Front Wiring')
        
        # TCF2 Backflush
        true_engine_tcf2, eng_cons_tcf2, eng_warn_tcf2 = ae.calculate_true_stock(engine_stocks_tcf2, tcf2_drops, bom_df, 'Engine')
        true_cockpit_tcf2, ck_cons_tcf2, ck_warn_tcf2 = ae.calculate_true_stock(tcf2_cockpit_start, tcf2_drops, bom_df, 'Cockpit')
        true_wiring_tcf2, wh_cons_tcf2, wh_warn_tcf2 = ae.calculate_true_stock(tcf2_wiring_start, tcf2_drops, bom_df, 'Front Wiring')
        
        # ----------------- PBS QUEUE ALLOCATION (FOR TCF1 & TCF2 LINE TABS) -----------------
        # Cabs physically in PBS buffer (PBS LIFT not null)
        pbs_all = float_df[float_df['PBS LIFT'].notna()].copy()
        
        is_hold_pbs = pbs_all['HOLD BY'].notna() & (pbs_all['HOLD BY'].astype(str).str.strip() != '') & (pbs_all['HOLD BY'].astype(str).str.upper() != 'NONE')
        pbs_on_hold = pbs_all[is_hold_pbs].copy()
        pbs_active = pbs_all[~is_hold_pbs].copy()
        
        tcf1_queue = pbs_active[pbs_active['SHOP'] == 'TCF1'].copy()
        tcf2_queue = pbs_active[pbs_active['SHOP'] == 'TCF2'].copy()
        
        tcf1_queue.sort_values(by='PBS LIFT', ascending=True, inplace=True)
        tcf2_queue.sort_values(by='PBS LIFT', ascending=True, inplace=True)
        
        # Build Punch EV (Nova) true material stock dict (Starting Clearance minus TCF1 Backflushed Drops)
        true_nova_dict = {}
        nova_backflushed = eng_cons_tcf1.get('546816111212', 0)
        if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None and not st.session_state.nova_materials_df.empty:
            for idx, r_nova in st.session_state.nova_materials_df.iterrows():
                mat_name = str(r_nova['Material']).strip()
                start_qty = int(r_nova['Clearance Qty'])
                true_nova_dict[mat_name] = start_qty - nova_backflushed

        # Build Model-Wise Shortage stock list (Starting Clearance minus Backflushed VIN Drops today)
        model_shortages_list = []
        if 'model_shortages_df' in st.session_state and st.session_state.model_shortages_df is not None and not st.session_state.model_shortages_df.empty:
            for idx, r_ms in st.session_state.model_shortages_df.iterrows():
                m_mod = str(r_ms['Model']).strip()
                m_trm = str(r_ms.get('Trims', 'All Trims')).strip()
                m_part = str(r_ms['Part Name']).strip()
                start_qty = int(r_ms['Clearance Qty'])
                
                vins_today = get_backflushed_vin_count_for_model_trims(m_mod, m_trm, tcf1_drops, tcf2_drops)
                true_stock = max(0, start_qty - vins_today)
                
                model_shortages_list.append({
                    'Model': m_mod,
                    'Trims': m_trm,
                    'Part Name': m_part,
                    'Stock': true_stock,
                    'Clearance Qty': start_qty,
                    'Backflushed VINs': vins_today
                })

        # Run allocation engine for PBS cabs only (for TCF1 & TCF2 tabs)
        tcf1_alloc, tcf1_final_stocks = ae.run_allocation(
            tcf1_queue, bom_df, true_engine_tcf1, true_cockpit_tcf1, true_wiring_tcf1,
            true_nova=true_nova_dict, model_shortages=model_shortages_list
        )
        tcf2_alloc, tcf2_final_stocks = ae.run_allocation(
            tcf2_queue, bom_df, true_engine_tcf2, true_cockpit_tcf2, true_wiring_tcf2,
            model_shortages=model_shortages_list
        )
        
        tcf1_alloc_df = pd.DataFrame(tcf1_alloc)
        tcf2_alloc_df = pd.DataFrame(tcf2_alloc)
        
        # ----------------- TOTAL FLOAT QUEUE ALLOCATION (FOR TOTAL FLOAT SEARCH & EXCEL EXPORT) -----------------
        is_hold_float = float_df['HOLD BY'].notna() & (float_df['HOLD BY'].astype(str).str.strip() != '') & (float_df['HOLD BY'].astype(str).str.upper() != 'NONE')
        float_active = float_df[~is_hold_float].copy()
        
        tcf1_total_queue = float_active[float_active['SHOP'] == 'TCF1'].copy()
        tcf2_total_queue = float_active[float_active['SHOP'] == 'TCF2'].copy()
        
        stage_sort_cols = [c for c in ['PBS LIFT', 'TOPCOAT', 'SEALANT', 'PTCED', 'BIW LIFTING'] if c in float_df.columns]
        if stage_sort_cols:
            tcf1_total_queue.sort_values(by=stage_sort_cols, ascending=[True]*len(stage_sort_cols), na_position='last', inplace=True)
            tcf2_total_queue.sort_values(by=stage_sort_cols, ascending=[True]*len(stage_sort_cols), na_position='last', inplace=True)

        tcf1_total_alloc, _ = ae.run_allocation(
            tcf1_total_queue, bom_df, true_engine_tcf1, true_cockpit_tcf1, true_wiring_tcf1,
            true_nova=true_nova_dict, model_shortages=model_shortages_list
        )
        tcf2_total_alloc, _ = ae.run_allocation(
            tcf2_total_queue, bom_df, true_engine_tcf2, true_cockpit_tcf2, true_wiring_tcf2,
            model_shortages=model_shortages_list
        )
        
        tcf1_total_alloc_df = pd.DataFrame(tcf1_total_alloc)
        tcf2_total_alloc_df = pd.DataFrame(tcf2_total_alloc)
        
        # Load All models.xlsx catalog for VC -> Trim mapping
        all_models_path = os.path.join(active_dir, "All models.xlsx")
        if not os.path.exists(all_models_path):
            all_models_path = os.path.join(workspace_dir, "TEST", "All models.xlsx")
        vc_to_desc, vc_to_trim = dl.load_all_models_catalog(all_models_path)

        def get_row_trim(row):
            s_desc = row.get('SALES DESCRIPTION', row.get('SALES DESC', None))
            m_name = row.get('Model', row.get('PRODUCT', ''))
            if s_desc and pd.notna(s_desc) and str(s_desc).strip() not in ['', 'nan']:
                return dl.extract_trim_from_sales_desc(s_desc, m_name)
            v_code = str(row.get('VEHICLE CODE', row.get('VC', ''))).strip()
            if v_code in vc_to_trim:
                return vc_to_trim[v_code]
            if v_code[:9] in vc_to_trim:
                return vc_to_trim[v_code[:9]]
            return "—"

        if not tcf1_alloc_df.empty:
            tcf1_alloc_df['Trim'] = tcf1_alloc_df.apply(get_row_trim, axis=1)
        else:
            tcf1_alloc_df['Trim'] = pd.Series(dtype='object')

        if not tcf2_alloc_df.empty:
            tcf2_alloc_df['Trim'] = tcf2_alloc_df.apply(get_row_trim, axis=1)
        else:
            tcf2_alloc_df['Trim'] = pd.Series(dtype='object')
        
        # Add Model column by mapping Engine Part No to Model & Line
        if 'engine_df' in st.session_state and not st.session_state.engine_df.empty:
            engine_to_model = dict(zip(st.session_state.engine_df['Engine Part No'].astype(str).str.strip(), st.session_state.engine_df['Model']))
            engine_to_line = dict(zip(st.session_state.engine_df['Engine Part No'].astype(str).str.strip(), st.session_state.engine_df['TCF Line']))
        else:
            engine_to_model = {item['Engine Part No']: item['Model'] for item in engine_default_data}
            engine_to_line = {item['Engine Part No']: item['TCF Line'] for item in engine_default_data}
            
        # Add explicit EV Model mappings
        engine_to_model['546816111212'] = 'Punch EV (Nova)'
        engine_to_line['546816111212'] = 'TCF1'
        engine_to_model['547380400103'] = 'Harrier EV'
        engine_to_line['547380400103'] = 'TCF2'

        # Helper function to map model name dynamically with EV & Tayrona overrides
        def map_row_model(df):
            if df is None or df.empty:
                return pd.Series(dtype='object')
            
            # Default mapping from Engine Part
            if 'Engine_Part' in df.columns:
                models_s = df['Engine_Part'].astype(str).str.strip().map(engine_to_model)
            else:
                models_s = pd.Series(dtype='object', index=df.index)
            
            models_s = pd.Series(models_s, index=df.index)
            is_missing = models_s.isna() | (models_s == '—') | (models_s == '')
            
            vc_col = 'VEHICLE CODE' if 'VEHICLE CODE' in df.columns else ('VC' if 'VC' in df.columns else None)
            if vc_col:
                vcs = df[vc_col].astype(str).str.strip()
                is_nova_vc = vcs.str.startswith('5468')
                is_harrier_ev_vc = vcs.str.startswith('5473')
                is_tayrona_vc = vcs.str.startswith('54831927A')
            else:
                is_nova_vc = pd.Series(False, index=df.index)
                is_harrier_ev_vc = pd.Series(False, index=df.index)
                is_tayrona_vc = pd.Series(False, index=df.index)
                
            # Override for Tayrona (Safari EV)
            if 'PRODUCT' in df.columns:
                is_tayrona = df['PRODUCT'].astype(str).str.strip().str.upper().str.contains('TAYRONA') | is_tayrona_vc
            else:
                is_tayrona = is_tayrona_vc
                
            res = np.where(is_tayrona, 'SAFARI EV',
                  np.where(is_nova_vc & is_missing, 'Punch EV (Nova)',
                  np.where(is_harrier_ev_vc & is_missing, 'Harrier EV', models_s)))
            return pd.Series(res, index=df.index).fillna('—')
            
        if not tcf1_alloc_df.empty:
            tcf1_alloc_df['Model'] = map_row_model(tcf1_alloc_df)
            tcf1_alloc_df['Cab location'] = tcf1_alloc_df.apply(ae.get_detailed_paint_summary_stage, axis=1)
        else:
            tcf1_alloc_df['Model'] = pd.Series(dtype='object')
            tcf1_alloc_df['Cab location'] = pd.Series(dtype='object')
            
        if not tcf2_alloc_df.empty:
            tcf2_alloc_df['Model'] = map_row_model(tcf2_alloc_df)
            tcf2_alloc_df['Cab location'] = tcf2_alloc_df.apply(ae.get_detailed_paint_summary_stage, axis=1)
        else:
            tcf2_alloc_df['Model'] = pd.Series(dtype='object')
            tcf2_alloc_df['Cab location'] = pd.Series(dtype='object')
            
        # Map Engine Part to Model in the raw drop data (VIN Generation)
        if tcf1_drops is not None and not tcf1_drops.empty:
            tcf1_drops['Model'] = map_row_model(tcf1_drops)
            
        if tcf2_drops is not None and not tcf2_drops.empty:
            tcf2_drops['Model'] = map_row_model(tcf2_drops)
        
        # ----------------- STAGEWISE MATERIAL SUMMARY -----------------
        # Get stages for all float report cabs
        if float_df is not None and not float_df.empty:
            float_stages_df = ae.get_paint_float_stages(float_df)
            
            # Combined stock registry for shortage calculation
            combined_true_stocks = {
                'engine': {**true_engine_tcf1, **true_engine_tcf2},
                'cockpit': {**true_cockpit_tcf1, **true_cockpit_tcf2},
                'wiring': {**true_wiring_tcf1, **true_wiring_tcf2}
            }
            
            shortage_report_df = ae.calculate_stagewise_shortage(float_stages_df, bom_df, combined_true_stocks)
        else:
            float_stages_df = pd.DataFrame()
            shortage_report_df = pd.DataFrame()

        # Scan every cab in the float for a missing/incomplete BOM entry, so cabs
        # that can't be matched to a BOM row never silently pass through as
        # "Ready for TCF" -- surfaced as an alert + quick-entry form on the homepage.
        missing_bom_df = ae.find_missing_bom_vcs(float_df, bom_df)

        # Build temp_float_df with stage, model, and engine mapping for all downstream views
        def get_summary_product_to_model(prod_name):
            prod = str(prod_name).strip().upper()
            if 'HORNBILL' in prod:
                return 'PUNCH'
            elif 'NOVA' in prod:
                return 'PUNCH.EV'
            elif 'ETURNA' in prod:
                return 'HARRIER.EV'
            elif 'GRAVITAS' in prod:
                return 'SAFARI'
            elif 'Q5' in prod:
                return 'HARRIER'
            elif 'TAYRONA' in prod:
                return 'SAFARI.EV'
            return 'UNKNOWN'
            
        def get_row_paint_stage(row):
            return ae.get_detailed_paint_summary_stage(row)

        if float_df is not None and not float_df.empty:
            temp_float_df = float_df.copy()
            temp_float_df['Model_Mapped'] = temp_float_df['PRODUCT'].apply(get_summary_product_to_model)
            temp_float_df['Stage'] = temp_float_df.apply(get_row_paint_stage, axis=1)
            if bom_df is not None and not bom_df.empty and 'VEHICLE CODE' in temp_float_df.columns:
                vc_to_engine = dict(zip(bom_df['Short Vehicle Code'].astype(str).str.strip(), bom_df['Engine'].astype(str).str.strip()))
                temp_float_df['Engine_Part'] = temp_float_df['VEHICLE CODE'].astype(str).str.strip().str[:9].map(vc_to_engine)
            
            # Map total float allocation status & blocking reason to temp_float_df
            alloc_status_map = {}
            alloc_reason_map = {}
            if 'tcf1_total_alloc_df' in locals() and not tcf1_total_alloc_df.empty and 'BIW NUMBER' in tcf1_total_alloc_df.columns:
                for idx, r in tcf1_total_alloc_df.iterrows():
                    b_key = str(r['BIW NUMBER']).strip()
                    alloc_status_map[b_key] = r.get('STATUS', '—')
                    alloc_reason_map[b_key] = r.get('BLOCKING_REASON', None)
            if 'tcf2_total_alloc_df' in locals() and not tcf2_total_alloc_df.empty and 'BIW NUMBER' in tcf2_total_alloc_df.columns:
                for idx, r in tcf2_total_alloc_df.iterrows():
                    b_key = str(r['BIW NUMBER']).strip()
                    alloc_status_map[b_key] = r.get('STATUS', '—')
                    alloc_reason_map[b_key] = r.get('BLOCKING_REASON', None)

            def get_cab_status(row):
                biw_str = str(row.get('BIW NUMBER', '')).strip()
                if biw_str in alloc_status_map:
                    return alloc_status_map[biw_str]
                hold_by = row.get('HOLD BY')
                if pd.notna(hold_by) and str(hold_by).strip() not in ['', 'None', 'nan']:
                    return '⚠️ Quality Hold'
                return '—'

            def get_cab_blocking_reason(row):
                biw_str = str(row.get('BIW NUMBER', '')).strip()
                reason = alloc_reason_map.get(biw_str)
                if pd.notna(reason) and str(reason).strip() not in ['', 'None', 'nan']:
                    return str(reason).strip()
                hold_by = row.get('HOLD BY')
                reasons_s = row.get('REASONS S')
                if pd.notna(hold_by) and str(hold_by).strip() not in ['', 'None', 'nan']:
                    r_str = f"Hold by {hold_by}"
                    if pd.notna(reasons_s) and str(reasons_s).strip() not in ['', 'None', 'nan']:
                        r_str += f": {reasons_s}"
                    return r_str
                return 'None (Clear)'

            temp_float_df['Status'] = temp_float_df.apply(get_cab_status, axis=1)
            temp_float_df['Blocking Reason'] = temp_float_df.apply(get_cab_blocking_reason, axis=1)
        else:
            temp_float_df = pd.DataFrame()

        ready_count = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf1_alloc_df.empty else 0
        blocked_count = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf1_alloc_df.empty else 0
        ready_count_tcf2 = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf2_alloc_df.empty else 0
        blocked_count_tcf2 = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf2_alloc_df.empty else 0

    st.session_state['last_generated_at'] = datetime.datetime.now()

except Exception as e:
    st.error(f"❌ Error while running calculations: {e}")
    st.info("Please verify that the uploaded files match the required structure and columns.")
    st.stop()

# ----------------- MAIN PANEL LAYOUT -----------------
# Helper function to render Total Float Details & Cab Search view
def render_total_float_details_view(float_df, default_line="All"):
    st.markdown("### 🔍 Total Float Details & Cab Search")
    st.markdown("""
        Search and filter cabs across the entire Paint Shop float using **BIW Number**, **Vehicle Code (VC)**, and **Paint Stage dropdown options**.
    """)
    
    if float_df is None or float_df.empty:
        st.info("Please load Paint Float report data in the Control Panel to view float details.")
        return
        
    df_search = float_df.copy()
    if 'Stage' not in df_search.columns:
        df_search['Stage'] = df_search.apply(ae.get_detailed_paint_summary_stage, axis=1)
    df_search['Cab location'] = df_search['Stage']
        
    # Search & Filter Controls
    c1, c2, c3, c4, c5 = st.columns([2, 2, 2.5, 2, 2])
    
    with c1:
        biw_query = st.text_input("🔍 BIW Number:", placeholder="e.g. 5012617", key=f"biw_search_{default_line}")
    with c2:
        vc_query = st.text_input("🚙 Vehicle Code (VC):", placeholder="e.g. 54972824A", key=f"vc_search_{default_line}")
    with c3:
        stages_available = ['All Stages'] + sorted(list(df_search['Stage'].dropna().unique()))
        selected_stage = st.selectbox("🎨 Paint Stage Filter:", options=stages_available, key=f"stage_filter_{default_line}")
    with c4:
        shop_options = ['All Shops', 'TCF1', 'TCF2']
        default_idx = shop_options.index(default_line) if default_line in shop_options else 0
        selected_shop = st.selectbox("🏭 Shop / Line Filter:", options=shop_options, index=default_idx, key=f"shop_filter_{default_line}")
    with c5:
        hold_filter = st.selectbox("🛑 Quality Hold Filter:", options=['All Cabs', 'Quality Hold Only', 'Clear Cabs Only'], key=f"hold_filter_{default_line}")
        
    # Apply filtering logic
    if biw_query.strip():
        df_search = df_search[df_search['BIW NUMBER'].astype(str).str.contains(biw_query.strip(), case=False, na=False)]
    if vc_query.strip():
        df_search = df_search[df_search['VEHICLE CODE'].astype(str).str.contains(vc_query.strip(), case=False, na=False)]
    if selected_stage != 'All Stages':
        df_search = df_search[df_search['Stage'] == selected_stage]
    if selected_shop != 'All Shops':
        df_search = df_search[df_search['SHOP'].astype(str).str.upper() == selected_shop.upper()]
    if hold_filter == 'Quality Hold Only':
        df_search = df_search[df_search['HOLD BY'].notna() & (df_search['HOLD BY'].astype(str).str.strip() != '') & (df_search['HOLD BY'].astype(str).str.upper() != 'NONE')]
    elif hold_filter == 'Clear Cabs Only':
        df_search = df_search[df_search['HOLD BY'].isna() | (df_search['HOLD BY'].astype(str).str.strip() == '') | (df_search['HOLD BY'].astype(str).str.upper() == 'NONE')]

    # KPI Summary Cards
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Matching Cabs", f"{len(df_search)} cabs")
    pbs_count = len(df_search[df_search['Stage'] == 'PBS FLOAT'])
    m2.metric("PBS Buffer Cabs", f"{pbs_count} cabs")
    hold_count = len(df_search[df_search['HOLD BY'].notna() & (df_search['HOLD BY'].astype(str).str.strip() != '') & (df_search['HOLD BY'].astype(str).str.upper() != 'NONE')])
    m3.metric("Quality Hold Cabs", f"{hold_count} cabs", delta_color="inverse")
    m4.metric("Total Plant Float", f"{len(float_df)} cabs")

    # Detailed Cab Inspector Card if BIW number searched
    if biw_query.strip() and len(df_search) == 1:
        cab = df_search.iloc[0]
        st.markdown(f"#### 🎴 Inspector Timeline for BIW #{cab.get('BIW NUMBER')}")
        is_dark = st.session_state.get('theme', '☀️ White Theme') == '🌙 Dark Theme'
        card_bg = "#1E293B" if is_dark else "#F8FAFC"
        border_c = "#334155" if is_dark else "#E2E8F0"
        
        st.markdown(f"""
        <div style="background-color: {card_bg}; border: 1px solid {border_c}; border-radius: 12px; padding: 1.25rem; margin-bottom: 1.5rem;">
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 1rem;">
                <div><strong>BIW Number:</strong> <br><span style="font-size: 1.1em; font-weight: bold; color: #3B82F6;">{cab.get('BIW NUMBER', '—')}</span></div>
                <div><strong>VIN:</strong> <br><span style="font-weight: bold;">{cab.get('VIN', '—')}</span></div>
                <div><strong>Vehicle Code:</strong> <br><code>{cab.get('VEHICLE CODE', '—')}</code></div>
                <div><strong>Product / Model:</strong> <br>{cab.get('PRODUCT', '—')} - {cab.get('MODEL', '—')}</div>
                <div><strong>Colour:</strong> <br>{cab.get('COLOUR', '—')}</div>
                <div><strong>Shop / Line:</strong> <br><strong>{cab.get('SHOP', '—')}</strong></div>
                <div><strong>Allocation Status:</strong> <br><span style="font-weight: bold;">{cab.get('Status', '—')}</span></div>
                <div><strong>Cab Location / Stage:</strong> <br><span style="background-color: #3B82F6; color: white; padding: 3px 8px; border-radius: 6px; font-weight: bold;">{cab.get('Cab location', cab.get('Stage', '—'))}</span></div>
            </div>
            <hr style="margin: 1rem 0; border: none; border-top: 1px solid {border_c};">
            <div style="font-size: 13px;">
                <strong>🛑 Quality Hold Status:</strong> {f"<span style='color: #EF4444; font-weight: bold;'>HOLD BY: {cab.get('HOLD BY')} | Reason: {cab.get('REASONS S')}</span>" if pd.notna(cab.get('HOLD BY')) and str(cab.get('HOLD BY')).strip() not in ['', 'None'] else "<span style='color: #10B981; font-weight: bold;'>✅ CLEAR (No Quality Hold)</span>"}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Interactive Data Table
    st.markdown("#### 📋 Float Cab Details")
    display_cols = ['BIW NUMBER', 'VIN', 'VEHICLE CODE', 'PRODUCT', 'MODEL', 'COLOUR', 'SHOP', 'Status', 'Blocking Reason', 'Cab location', 'Stage', 'HOLD BY', 'REASONS S', 'BIW LIFTING', 'PTCED', 'SEALANT', 'TOPCOAT']
    available_cols = [c for c in display_cols if c in df_search.columns]

    def _style_status_cell(val):
        val_str = str(val)
        if 'Ready' in val_str or '✅' in val_str:
            return 'background-color: rgba(16, 185, 129, 0.18); color: #059669; font-weight: 600;'
        elif 'Blocked' in val_str or '🚫' in val_str:
            return 'background-color: rgba(239, 68, 68, 0.18); color: #DC2626; font-weight: 600;'
        elif 'Hold' in val_str or '⚠️' in val_str:
            return 'background-color: rgba(245, 158, 11, 0.18); color: #B45309; font-weight: 600;'
        return ''

    display_df = df_search[available_cols]
    if 'Status' in available_cols:
        # Styler.applymap was removed in pandas 3.0 -- use .map instead.
        styled_display_df = display_df.style.map(_style_status_cell, subset=['Status'])
        st.dataframe(styled_display_df, use_container_width=True, hide_index=True)
    else:
        st.dataframe(display_df, use_container_width=True, hide_index=True)
    
    # Export options with colorful OpenPyXL styling and Ready to TCF Short VC Pivot + Blocked Reason Pivot
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    excel_buffer = io.BytesIO()
    df_export = df_search[available_cols].copy()
    
    # Build Pivot sheet 2 for ONLY Ready to TCF cabs
    df_search_pivot = df_search.copy()
    if 'Status' in df_search_pivot.columns:
        ready_mask = df_search_pivot['Status'].astype(str).str.contains('Ready for TCF', case=False, na=False)
        df_search_pivot = df_search_pivot[ready_mask].copy()
        
    if 'VEHICLE CODE' in df_search_pivot.columns:
        df_search_pivot['Short VC'] = df_search_pivot['VEHICLE CODE'].astype(str).str.strip().str[:9]
    elif 'VC' in df_search_pivot.columns:
        df_search_pivot['Short VC'] = df_search_pivot['VC'].astype(str).str.strip().str[:9]
    else:
        df_search_pivot['Short VC'] = ''
        
    if 'SHOP' not in df_search_pivot.columns:
        df_search_pivot['SHOP'] = 'All'
    if 'MODEL' not in df_search_pivot.columns:
        df_search_pivot['MODEL'] = df_search_pivot.get('PRODUCT', 'Unknown')
        
    if not df_search_pivot.empty:
        pivot_df = df_search_pivot.groupby(['SHOP', 'MODEL', 'Short VC']).size().reset_index(name='Ready to TCF Qty')
        tot_row = pd.DataFrame([{'SHOP': 'Total', 'MODEL': '', 'Short VC': '', 'Ready to TCF Qty': pivot_df['Ready to TCF Qty'].sum()}])
        pivot_df_with_tot = pd.concat([pivot_df, tot_row], ignore_index=True)
    else:
        pivot_df_with_tot = pd.DataFrame(columns=['SHOP', 'MODEL', 'Short VC', 'Ready to TCF Qty'])
        
    # Build Sheet 3 for Blocked Reason & Short VC Pivot
    df_blocked_cabs = df_search[df_search['Status'].astype(str).str.contains('Blocked|Hold', case=False, na=False)].copy()
    if 'VEHICLE CODE' in df_blocked_cabs.columns:
        df_blocked_cabs['Short VC'] = df_blocked_cabs['VEHICLE CODE'].astype(str).str.strip().str[:9]
    elif 'VC' in df_blocked_cabs.columns:
        df_blocked_cabs['Short VC'] = df_blocked_cabs['VC'].astype(str).str.strip().str[:9]
    else:
        df_blocked_cabs['Short VC'] = ''

    if 'SHOP' not in df_blocked_cabs.columns:
        df_blocked_cabs['SHOP'] = 'All'
    if 'MODEL' not in df_blocked_cabs.columns:
        df_blocked_cabs['MODEL'] = df_blocked_cabs.get('PRODUCT', 'Unknown')
    if 'Blocking Reason' not in df_blocked_cabs.columns:
        df_blocked_cabs['Blocking Reason'] = 'Unknown'

    if not df_blocked_cabs.empty:
        pivot_blocked = df_blocked_cabs.groupby(['SHOP', 'MODEL', 'Short VC', 'Blocking Reason']).size().reset_index(name='Blocked Cab Qty')
        tot_blocked_row = pd.DataFrame([{'SHOP': 'Total', 'MODEL': '', 'Short VC': '', 'Blocking Reason': '', 'Blocked Cab Qty': pivot_blocked['Blocked Cab Qty'].sum()}])
        pivot_blocked_with_tot = pd.concat([pivot_blocked, tot_blocked_row], ignore_index=True)
    else:
        pivot_blocked_with_tot = pd.DataFrame(columns=['SHOP', 'MODEL', 'Short VC', 'Blocking Reason', 'Blocked Cab Qty'])

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Float Details')
        pivot_df_with_tot.to_excel(writer, index=False, sheet_name='Ready to Upload Plan')
        pivot_blocked_with_tot.to_excel(writer, index=False, sheet_name='Blocked Reason Pivot')
        
        # Apply colorful & professional styling
        wb = writer.book
        
        header_fill_default = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_fill_blocked = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        alt_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        alt_fill_odd = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        alt_fill_blocked = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
        
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill_blocked = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        total_font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
        total_font_blocked = Font(name="Segoe UI", size=11, bold=True, color="C65911")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        total_border = Border(
            top=Side(style='thin', color='1F4E78'),
            bottom=Side(style='double', color='1F4E78'),
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9')
        )

        total_border_blocked = Border(
            top=Side(style='thin', color='C65911'),
            bottom=Side(style='double', color='C65911'),
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9')
        )
        
        ready_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ready_font = Font(name="Segoe UI", size=10, color="274E13", bold=True)
        
        blocked_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        blocked_font = Font(name="Segoe UI", size=10, color="C65911", bold=True)

        hold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        hold_font = Font(name="Segoe UI", size=10, color="806000", bold=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            is_blocked_sheet = (sheet_name == 'Blocked Reason Pivot')
            
            # Format Header Row
            ws.row_dimensions[1].height = 26
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill_blocked if is_blocked_sheet else header_fill_default
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            # Format Data Rows
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                is_last_row = (row_idx == ws.max_row) and (sheet_name in ['Ready to Upload Plan', 'Blocked Reason Pivot'])
                
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    val_str = str(cell.value or '').strip()
                    
                    if is_last_row:
                        cell.fill = total_fill_blocked if is_blocked_sheet else total_fill
                        cell.font = total_font_blocked if is_blocked_sheet else total_font
                        cell.border = total_border_blocked if is_blocked_sheet else total_border
                        if col_idx == ws.max_column or isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        # Zebra striping
                        if is_blocked_sheet:
                            cell.fill = alt_fill_blocked if row_idx % 2 == 0 else alt_fill_even
                        else:
                            cell.fill = alt_fill_odd if row_idx % 2 == 0 else alt_fill_even
                        cell.font = Font(name="Segoe UI", size=10)
                        
                        # Highlighting Status values in Float Details sheet
                        if 'Ready for TCF' in val_str:
                            cell.fill = ready_fill
                            cell.font = ready_font
                        elif 'Blocked' in val_str or 'Shortage' in val_str:
                            cell.fill = blocked_fill
                            cell.font = blocked_font
                        elif 'Hold' in val_str:
                            cell.fill = hold_fill
                            cell.font = hold_font
                            
                        # Align numbers to right, text to left
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Adjust column widths dynamically
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    st.download_button(
        label="📥 Export Float Details & Short VC / Blocked Pivots to Excel",
        data=excel_buffer.getvalue(),
        file_name=f"total_float_details_{default_line}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"export_float_details_{default_line}"
    )

active_clearance_shortage_alerts = evaluate_all_clearance_shortage_alerts(
    st.session_state.get('model_shortages_df'),
    st.session_state.get('nova_materials_df'),
    st.session_state.get('engine_df'),
    tcf1_drops,
    tcf2_drops
)

if active_clearance_shortage_alerts:
    alert_box_html = f"""
    <div style="background-color: #FEF2F2; border: 1.5px solid #EF4444; border-radius: 10px; padding: 14px 18px; margin-bottom: 16px;">
        <div style="font-weight: 700; font-size: 15px; color: #991B1B; display: flex; align-items: center; gap: 8px;">
            <span style="font-size: 18px;">🚨</span>
            <span>MATERIAL SHORTAGE ALERTS TRIGGERED ({len(active_clearance_shortage_alerts)} Critical Shortage Item{'s' if len(active_clearance_shortage_alerts)>1 else ''})</span>
        </div>
        <div style="font-size: 13px; color: #7F1D1D; margin-top: 8px; line-height: 1.6;">
    """
    for al in active_clearance_shortage_alerts:
        alert_box_html += f"• <b>{al['Model']} [{al['Trims']}] - {al['Part Name']}</b>: Clearance Stock is <b>{al['Clearance Qty']}</b> vs Demand <b>{al['Demand Qty']}</b> (<span style='color:#DC2626; font-weight:700;'>Shortage: -{al['Shortage Qty']} units</span>)<br>"
    alert_box_html += "</div></div>"
    st.markdown(alert_box_html, unsafe_allow_html=True)

# Toggle between TCF1, TCF2, Total Float Details, Combined Summary & Reports (Opening tab: Summary Report & Excel Download)
tcf_tabs = st.tabs([
    "📈 Summary Report & Excel Download",
    "🏭 TCF 1 Line (Altroz/Punch/Nova)", 
    "🏭 TCF 2 Line (Harrier/Safari)", 
    "🔍 Total Float Details & Search",
    "📋 Quality Hold Registry",
    "📱 Telegram Dispatcher"
])

# Helper function to style allocation dataframe rows
def style_alloc_table(df):
    if df.empty:
        return df
    
    is_dark = st.session_state.get('theme', '☀️ White Theme') == '🌙 Dark Theme'
    
    def get_row_style(row):
        status = row.get('STATUS')
        if status == '✅ Ready for TCF':
            if is_dark:
                return ['background-color: #064E3B; color: #D1FAE5'] * len(row)
            else:
                return ['background-color: #F0FAF4; color: #166534'] * len(row)
        elif status == '🚫 Blocked':
            if is_dark:
                return ['background-color: #7F1D1D; color: #FEE2E2'] * len(row)
            else:
                return ['background-color: #FFF5F5; color: #B91C1C'] * len(row)
        else:
            if is_dark:
                return ['background-color: #78350F; color: #FEF3C7'] * len(row)
            else:
                return ['background-color: #FFFBEB; color: #92400E'] * len(row)
            
    return df.style.apply(get_row_style, axis=1)

# Helper function to render horizontal bar charts of stocks
def plot_stock_chart(start, true, final, title):
    if not start:
        st.info(f"No stock data loaded for: {title}")
        return
    parts = list(start.keys())
    
    # Limit to top 15 parts to prevent visual clutter
    if len(parts) > 15:
        parts = parts[:15]
        
    start_vals = [start.get(p, 0) for p in parts]
    true_vals = [true.get(p, 0) for p in parts]
    final_vals = [final.get(p, 0) for p in parts]
    
    y = np.arange(len(parts))
    width = 0.25
    
    is_dark = st.session_state.get('theme', '☀️ White Theme') == '🌙 Dark Theme'
    
    # Theme configuration
    bg_color = '#0F172A' if is_dark else '#F8FAFC'
    face_color = '#1E293B' if is_dark else '#FFFFFF'
    label_color = '#94A3B8' if is_dark else '#374A6B'
    title_color = '#F8FAFC' if is_dark else '#1B2A4A'
    border_color = '#334155' if is_dark else '#E8ECF1'
    grid_color = '#334155' if is_dark else '#F1F5F9'
    
    fig, ax = plt.subplots(figsize=(10, min(6, len(parts)*0.4 + 1.5)))
    fig.patch.set_facecolor(bg_color)
    ax.set_facecolor(face_color)
    
    # Colors: Slate, Royal Blue / Light Indigo, Emerald Green
    start_color = '#64748B' if is_dark else '#94A3B8'
    true_color = '#6366F1' if is_dark else '#0D9488'
    final_color = '#10B981' if is_dark else '#22C55E'
    
    ax.barh(y - width, start_vals, width, label='Shift Start Stock', color=start_color, edgecolor=face_color, linewidth=0.5)
    ax.barh(y, true_vals, width, label='True Current Stock', color=true_color, edgecolor=face_color, linewidth=0.5)
    ax.barh(y + width, final_vals, width, label='Post-Alloc Virtual Stock', color=final_color, edgecolor=face_color, linewidth=0.5)
    
    ax.set_ylabel('Part Numbers', fontsize=10, fontweight='semibold', color=label_color)
    ax.set_xlabel('Quantity', fontsize=10, fontweight='semibold', color=label_color)
    ax.set_title(title, fontsize=12, fontweight='bold', pad=15, color=title_color)
    ax.set_yticks(y)
    ax.set_yticklabels(parts, fontsize=8, color=label_color)
    ax.tick_params(axis='x', colors=label_color)
    ax.legend(frameon=True, facecolor=face_color, edgecolor=border_color, labelcolor=label_color, fontsize=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color(border_color)
    ax.spines['bottom'].set_color(border_color)
    ax.grid(axis='x', color=grid_color, linewidth=0.5)
    plt.tight_layout()
    st.pyplot(fig)

# ----------------- TAB 2: TCF 1 LINE -----------------
with tcf_tabs[1]:
    # KPIs
    ready_count = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf1_alloc_df.empty else 0
    blocked_count = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf1_alloc_df.empty else 0
    issue_count = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'].str.startswith('⚠️', na=False)]) if not tcf1_alloc_df.empty else 0
    total_drops = int(tcf1_drops['VIN_Count'].sum()) if (tcf1_drops is not None and not tcf1_drops.empty and 'VIN_Count' in tcf1_drops.columns) else (len(tcf1_drops) if tcf1_drops is not None else 0)
    
    tcf1_ok = len(tcf1_queue)
    tcf1_hold = len(pbs_on_hold[pbs_on_hold['SHOP'] == 'TCF1']) if pbs_on_hold is not None else 0
    tcf1_total = tcf1_ok + tcf1_hold
    
    kpi_cols = st.columns(4)
    kpi_cols[0].metric("VIN Generation", f"{total_drops} cabs", help="Cabs built in TCF1 since shift start")
    kpi_cols[1].metric("PBS Current Stock", f"{tcf1_total} cabs (OK: {tcf1_ok} | Hold: {tcf1_hold})", help="Total cabs in TCF1 PBS buffer (Active unblocked + Quality holds)")
    kpi_cols[2].metric("✅ Ready for TCF", f"{ready_count} cabs", delta=f"+{ready_count} alloc")
    kpi_cols[3].metric("🚫 Blocked (Stock Out)", f"{blocked_count} cabs", delta=f"-{blocked_count} wait", delta_color="inverse")
    
    line_subtabs = st.tabs(["📋 FIFO Allocation Queue", "🔍 Total Float Details & Search"])
    
    # Subtab 1: Queue
    with line_subtabs[0]:
        st.markdown("### TCF1 FIFO Buffer Queue Status")
        if tcf1_alloc_df.empty:
            st.info("No active cabs in TCF1 PBS queue.")
        else:
            # Filter to show only Ready for TCF and Blocked statuses
            filtered_df = tcf1_alloc_df[tcf1_alloc_df['STATUS'].isin(['✅ Ready for TCF', '🚫 Blocked'])].copy()
            
            # Apply any manual planner overrides from session_state
            if 'tcf1_manual_overrides' not in st.session_state:
                st.session_state.tcf1_manual_overrides = {}
            for biw_key, override in st.session_state.tcf1_manual_overrides.items():
                mask = filtered_df['BIW NUMBER'].astype(str) == str(biw_key)
                if mask.any():
                    filtered_df.loc[mask, 'STATUS'] = override['status']
                    filtered_df.loc[mask, 'BLOCKING_REASON'] = override['reason']
            
            c_srch1, c_srch2 = st.columns([3, 3])
            with c_srch1:
                search_biw = st.text_input("🔍 Quick Search by BIW Number:", key="tcf1_biw_search")
            with c_srch2:
                loc_options = ['All Locations'] + sorted(list(filtered_df['Cab location'].dropna().unique())) if 'Cab location' in filtered_df.columns else ['All Locations']
                selected_loc = st.selectbox("📍 Cab Location Filter:", options=loc_options, key="tcf1_loc_search")
                
            if search_biw:
                filtered_df = filtered_df[filtered_df['BIW NUMBER'].astype(str).str.contains(search_biw.strip())]
            if selected_loc != 'All Locations':
                filtered_df = filtered_df[filtered_df['Cab location'] == selected_loc]
                
            display_cols = ['BIW NUMBER', 'Model', 'Trim', 'VEHICLE CODE', 'STATUS', 'BLOCKING_REASON', 'Cab location', 'Engine_Part', 'Cockpit_Part', 'Wiring_Part', 'Engine_Stock_After', 'Cockpit_Stock_After', 'Wiring_Stock_After']
            
            st.caption("✏️ **Planner Edit Mode** — Click any cell in the **Status** or **Blocking Reason** column to change it. Changes are saved automatically.")
            
            status_options = ['✅ Ready for TCF', '🚫 Blocked', '⚠️ PBS Hold']
            
            edited_tcf1 = st.data_editor(
                filtered_df[display_cols],
                use_container_width=True,
                hide_index=True,
                key="tcf1_queue_editor",
                column_config={
                    "BIW NUMBER": st.column_config.TextColumn("BIW NUMBER", disabled=True),
                    "Model": st.column_config.TextColumn("Model", disabled=True),
                    "Trim": st.column_config.TextColumn("Trim", disabled=True),
                    "VEHICLE CODE": st.column_config.TextColumn("VEHICLE CODE", disabled=True),
                    "STATUS": st.column_config.SelectboxColumn(
                        "STATUS",
                        options=status_options,
                        required=True,
                        help="Select cab status: Ready, Blocked, or PBS Hold"
                    ),
                    "BLOCKING_REASON": st.column_config.TextColumn(
                        "BLOCKING REASON",
                        help="Enter blocking/hold reason (e.g., quality issue, part shortage, PBS hold)"
                    ),
                    "Cab location": st.column_config.TextColumn("Cab Location", disabled=True),
                    "Engine_Part": st.column_config.TextColumn("Engine Part", disabled=True),
                    "Cockpit_Part": st.column_config.TextColumn("Cockpit Part", disabled=True),
                    "Wiring_Part": st.column_config.TextColumn("Wiring Part", disabled=True),
                    "Engine_Stock_After": st.column_config.NumberColumn("Eng Stock After", disabled=True),
                    "Cockpit_Stock_After": st.column_config.NumberColumn("CK Stock After", disabled=True),
                    "Wiring_Stock_After": st.column_config.NumberColumn("WH Stock After", disabled=True),
                },
            )
            
            # Detect and persist planner edits
            original_display = filtered_df[display_cols].reset_index(drop=True)
            edited_display = edited_tcf1.reset_index(drop=True)
            if not original_display.equals(edited_display):
                for i in range(len(edited_display)):
                    orig_status = str(original_display.at[i, 'STATUS']) if i < len(original_display) else ''
                    new_status = str(edited_display.at[i, 'STATUS'])
                    orig_reason = str(original_display.at[i, 'BLOCKING_REASON']) if i < len(original_display) else ''
                    new_reason = str(edited_display.at[i, 'BLOCKING_REASON'])
                    if orig_status != new_status or orig_reason != new_reason:
                        biw_key = str(edited_display.at[i, 'BIW NUMBER'])
                        st.session_state.tcf1_manual_overrides[biw_key] = {
                            'status': new_status,
                            'reason': new_reason
                        }
                st.toast("✅ Planner override saved!", icon="✏️")
                # Update filtered_df to reflect edits for Excel downloads
                filtered_df.update(edited_tcf1)
            
            # Excel download buttons for Ready to TCF and Blocked with Reason
            import io
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            import openpyxl.utils
            
            dl_cols = st.columns(2)
            
            # Ready to TCF Excel
            ready_df_tcf1 = filtered_df[filtered_df['STATUS'] == '✅ Ready for TCF'][display_cols].copy()
            if not ready_df_tcf1.empty:
                buf_ready = io.BytesIO()
                with pd.ExcelWriter(buf_ready, engine='openpyxl') as writer:
                    ready_df_tcf1.to_excel(writer, index=False, sheet_name='Ready to TCF1')
                    ws = writer.sheets['Ready to TCF1']
                    hdr_fill = PatternFill(start_color='D8F3E5', end_color='D8F3E5', fill_type='solid')
                    hdr_font = Font(name='Calibri', size=11, bold=True, color='1B4D32')
                    thin_b = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
                    for c in range(1, len(ready_df_tcf1.columns) + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(ready_df_tcf1) + 2):
                        for c in range(1, len(ready_df_tcf1.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(name='Calibri', size=10)
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
                        
                    # Sheet 2: ready to upload (Pivot of Short VC and count)
                    ready_full_1 = filtered_df[filtered_df['STATUS'] == '✅ Ready for TCF'].copy()
                    vc_s1 = ready_full_1['VEHICLE CODE'] if 'VEHICLE CODE' in ready_full_1.columns else (ready_full_1['VC'] if 'VC' in ready_full_1.columns else pd.Series('', index=ready_full_1.index))
                    ready_full_1['Short VC'] = vc_s1.astype(str).str.strip().str[:9]
                    pivot_df_1 = ready_full_1.groupby('Short VC').size().reset_index(name='Count')
                    tot_row_1 = pd.DataFrame([{'Short VC': 'Total', 'Count': pivot_df_1['Count'].sum()}])
                    pivot_full_1 = pd.concat([pivot_df_1, tot_row_1], ignore_index=True)
                    
                    pivot_full_1.to_excel(writer, index=False, sheet_name='ready to upload')
                    ws2 = writer.sheets['ready to upload']
                    for c in range(1, len(pivot_full_1.columns) + 1):
                        cell = ws2.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(pivot_full_1) + 2):
                        is_tot = (r == len(pivot_full_1) + 1)
                        for c in range(1, len(pivot_full_1.columns) + 1):
                            cell = ws2.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if is_tot:
                                cell.font = Font(name='Calibri', size=11, bold=True)
                                cell.fill = hdr_fill
                            else:
                                cell.font = Font(name='Calibri', size=10)
                    for col in ws2.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws2.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 15)
                with dl_cols[0]:
                    st.download_button(
                        label=f"📥 Ready to TCF1 ({len(ready_df_tcf1)} cabs)",
                        data=buf_ready.getvalue(),
                        file_name="TCF1_Ready_to_Build.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_ready_tcf1"
                    )
            
            # Blocked with Reason Excel (includes both 🚫 Blocked and ⚠️ PBS Hold)
            blocked_df_tcf1 = filtered_df[filtered_df['STATUS'].isin(['🚫 Blocked', '⚠️ PBS Hold'])][display_cols].copy()
            if not blocked_df_tcf1.empty:
                buf_blocked = io.BytesIO()
                with pd.ExcelWriter(buf_blocked, engine='openpyxl') as writer:
                    blocked_df_tcf1.to_excel(writer, index=False, sheet_name='Blocked TCF1')
                    ws = writer.sheets['Blocked TCF1']
                    hdr_fill = PatternFill(start_color='FFD1D1', end_color='FFD1D1', fill_type='solid')
                    hdr_font = Font(name='Calibri', size=11, bold=True, color='5C1D1B')
                    thin_b = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
                    for c in range(1, len(blocked_df_tcf1.columns) + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(blocked_df_tcf1) + 2):
                        for c in range(1, len(blocked_df_tcf1.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(name='Calibri', size=10)
                            # Highlight blocking reason column in light red
                            if c == 5:  # BLOCKING_REASON column
                                cell.fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
                with dl_cols[1]:
                    st.download_button(
                        label=f"📥 Blocked with Reason ({len(blocked_df_tcf1)} cabs)",
                        data=buf_blocked.getvalue(),
                        file_name="TCF1_Blocked_with_Reason.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_blocked_tcf1"
                    )
            
    # Subtab 2: Total Float Details
    with line_subtabs[1]:
        render_total_float_details_view(temp_float_df, default_line="TCF1")

# ----------------- TAB 3: TCF 2 LINE -----------------
with tcf_tabs[2]:
    # KPIs
    ready_count_tcf2 = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf2_alloc_df.empty else 0
    blocked_count_tcf2 = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf2_alloc_df.empty else 0
    issue_count_tcf2 = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'].str.startswith('⚠️', na=False)]) if not tcf2_alloc_df.empty else 0
    total_drops_tcf2 = int(tcf2_drops['VIN_Count'].sum()) if (tcf2_drops is not None and not tcf2_drops.empty and 'VIN_Count' in tcf2_drops.columns) else (len(tcf2_drops) if tcf2_drops is not None else 0)
    
    tcf2_ok = len(tcf2_queue)
    tcf2_hold = len(pbs_on_hold[pbs_on_hold['SHOP'] == 'TCF2']) if pbs_on_hold is not None else 0
    tcf2_total = tcf2_ok + tcf2_hold
    
    kpi_cols_tcf2 = st.columns(4)
    kpi_cols_tcf2[0].metric("VIN Generation", f"{total_drops_tcf2} cabs", help="Cabs built in TCF2 since shift start")
    kpi_cols_tcf2[1].metric("PBS Current Stock", f"{tcf2_total} cabs (OK: {tcf2_ok} | Hold: {tcf2_hold})", help="Total cabs in TCF2 PBS buffer (Active unblocked + Quality holds)")
    kpi_cols_tcf2[2].metric("✅ Ready for TCF", f"{ready_count_tcf2} cabs", delta=f"+{ready_count_tcf2} alloc")
    kpi_cols_tcf2[3].metric("🚫 Blocked (Stock Out)", f"{blocked_count_tcf2} cabs", delta=f"-{blocked_count_tcf2} wait", delta_color="inverse")
    
    line_subtabs_tcf2 = st.tabs(["📋 FIFO Allocation Queue", "🔍 Total Float Details & Search"])
    
    # Subtab 1: Queue
    with line_subtabs_tcf2[0]:
        st.markdown("### TCF2 FIFO Buffer Queue Status")
        if tcf2_alloc_df.empty:
            st.info("No active cabs in TCF2 PBS queue.")
        else:
            # Filter to show only Ready for TCF and Blocked statuses
            filtered_df_tcf2 = tcf2_alloc_df[tcf2_alloc_df['STATUS'].isin(['✅ Ready for TCF', '🚫 Blocked'])].copy()
            
            # Apply any manual planner overrides from session_state
            if 'tcf2_manual_overrides' not in st.session_state:
                st.session_state.tcf2_manual_overrides = {}
            for biw_key, override in st.session_state.tcf2_manual_overrides.items():
                mask = filtered_df_tcf2['BIW NUMBER'].astype(str) == str(biw_key)
                if mask.any():
                    filtered_df_tcf2.loc[mask, 'STATUS'] = override['status']
                    filtered_df_tcf2.loc[mask, 'BLOCKING_REASON'] = override['reason']
            
            c_srch1_tcf2, c_srch2_tcf2 = st.columns([3, 3])
            with c_srch1_tcf2:
                search_biw_tcf2 = st.text_input("🔍 Quick Search by BIW Number:", key="tcf2_biw_search")
            with c_srch2_tcf2:
                loc_options_tcf2 = ['All Locations'] + sorted(list(filtered_df_tcf2['Cab location'].dropna().unique())) if 'Cab location' in filtered_df_tcf2.columns else ['All Locations']
                selected_loc_tcf2 = st.selectbox("📍 Cab Location Filter:", options=loc_options_tcf2, key="tcf2_loc_search")
                
            if search_biw_tcf2:
                filtered_df_tcf2 = filtered_df_tcf2[filtered_df_tcf2['BIW NUMBER'].astype(str).str.contains(search_biw_tcf2.strip())]
            if selected_loc_tcf2 != 'All Locations':
                filtered_df_tcf2 = filtered_df_tcf2[filtered_df_tcf2['Cab location'] == selected_loc_tcf2]
                
            display_cols = ['BIW NUMBER', 'Model', 'Trim', 'VEHICLE CODE', 'STATUS', 'BLOCKING_REASON', 'Cab location', 'Engine_Part', 'Cockpit_Part', 'Wiring_Part', 'Engine_Stock_After', 'Cockpit_Stock_After', 'Wiring_Stock_After']
            
            st.caption("✏️ **Planner Edit Mode** — Click any cell in the **Status** or **Blocking Reason** column to change it. Changes are saved automatically.")
            
            status_options = ['✅ Ready for TCF', '🚫 Blocked', '⚠️ PBS Hold']
            
            edited_tcf2 = st.data_editor(
                filtered_df_tcf2[display_cols],
                use_container_width=True,
                hide_index=True,
                key="tcf2_queue_editor",
                column_config={
                    "BIW NUMBER": st.column_config.TextColumn("BIW NUMBER", disabled=True),
                    "Model": st.column_config.TextColumn("Model", disabled=True),
                    "Trim": st.column_config.TextColumn("Trim", disabled=True),
                    "VEHICLE CODE": st.column_config.TextColumn("VEHICLE CODE", disabled=True),
                    "STATUS": st.column_config.SelectboxColumn(
                        "STATUS",
                        options=status_options,
                        required=True,
                        help="Select cab status: Ready, Blocked, or PBS Hold"
                    ),
                    "BLOCKING_REASON": st.column_config.TextColumn(
                        "BLOCKING REASON",
                        help="Enter blocking/hold reason (e.g., quality issue, part shortage, PBS hold)"
                    ),
                    "Cab location": st.column_config.TextColumn("Cab Location", disabled=True),
                    "Engine_Part": st.column_config.TextColumn("Engine Part", disabled=True),
                    "Cockpit_Part": st.column_config.TextColumn("Cockpit Part", disabled=True),
                    "Wiring_Part": st.column_config.TextColumn("Wiring Part", disabled=True),
                    "Engine_Stock_After": st.column_config.NumberColumn("Eng Stock After", disabled=True),
                    "Cockpit_Stock_After": st.column_config.NumberColumn("CK Stock After", disabled=True),
                    "Wiring_Stock_After": st.column_config.NumberColumn("WH Stock After", disabled=True),
                },
            )
            
            # Detect and persist planner edits
            original_display_tcf2 = filtered_df_tcf2[display_cols].reset_index(drop=True)
            edited_display_tcf2 = edited_tcf2.reset_index(drop=True)
            if not original_display_tcf2.equals(edited_display_tcf2):
                for i in range(len(edited_display_tcf2)):
                    orig_status = str(original_display_tcf2.at[i, 'STATUS']) if i < len(original_display_tcf2) else ''
                    new_status = str(edited_display_tcf2.at[i, 'STATUS'])
                    orig_reason = str(original_display_tcf2.at[i, 'BLOCKING_REASON']) if i < len(original_display_tcf2) else ''
                    new_reason = str(edited_display_tcf2.at[i, 'BLOCKING_REASON'])
                    if orig_status != new_status or orig_reason != new_reason:
                        biw_key = str(edited_display_tcf2.at[i, 'BIW NUMBER'])
                        st.session_state.tcf2_manual_overrides[biw_key] = {
                            'status': new_status,
                            'reason': new_reason
                        }
                st.toast("✅ Planner override saved!", icon="✏️")
                # Update filtered_df_tcf2 to reflect edits for Excel downloads
                filtered_df_tcf2.update(edited_tcf2)
            
            # Excel download buttons for Ready to TCF and Blocked with Reason
            import io
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            import openpyxl.utils
            
            dl_cols_tcf2 = st.columns(2)
            
            # Ready to TCF Excel
            ready_df_tcf2 = filtered_df_tcf2[filtered_df_tcf2['STATUS'] == '✅ Ready for TCF'][display_cols].copy()
            if not ready_df_tcf2.empty:
                buf_ready2 = io.BytesIO()
                with pd.ExcelWriter(buf_ready2, engine='openpyxl') as writer:
                    ready_df_tcf2.to_excel(writer, index=False, sheet_name='Ready to TCF2')
                    ws = writer.sheets['Ready to TCF2']
                    hdr_fill = PatternFill(start_color='D8F3E5', end_color='D8F3E5', fill_type='solid')
                    hdr_font = Font(name='Calibri', size=11, bold=True, color='1B4D32')
                    thin_b = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
                    for c in range(1, len(ready_df_tcf2.columns) + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(ready_df_tcf2) + 2):
                        for c in range(1, len(ready_df_tcf2.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(name='Calibri', size=10)
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
                        
                    # Sheet 2: ready to upload (Pivot of Short VC and count)
                    ready_full_2 = filtered_df_tcf2[filtered_df_tcf2['STATUS'] == '✅ Ready for TCF'].copy()
                    vc_s2 = ready_full_2['VEHICLE CODE'] if 'VEHICLE CODE' in ready_full_2.columns else (ready_full_2['VC'] if 'VC' in ready_full_2.columns else pd.Series('', index=ready_full_2.index))
                    ready_full_2['Short VC'] = vc_s2.astype(str).str.strip().str[:9]
                    pivot_df_2 = ready_full_2.groupby('Short VC').size().reset_index(name='Count')
                    tot_row_2 = pd.DataFrame([{'Short VC': 'Total', 'Count': pivot_df_2['Count'].sum()}])
                    pivot_full_2 = pd.concat([pivot_df_2, tot_row_2], ignore_index=True)
                    
                    pivot_full_2.to_excel(writer, index=False, sheet_name='ready to upload')
                    ws2_2 = writer.sheets['ready to upload']
                    for c in range(1, len(pivot_full_2.columns) + 1):
                        cell = ws2_2.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(pivot_full_2) + 2):
                        is_tot = (r == len(pivot_full_2) + 1)
                        for c in range(1, len(pivot_full_2.columns) + 1):
                            cell = ws2_2.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if is_tot:
                                cell.font = Font(name='Calibri', size=11, bold=True)
                                cell.fill = hdr_fill
                            else:
                                cell.font = Font(name='Calibri', size=10)
                    for col in ws2_2.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws2_2.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 15)
                with dl_cols_tcf2[0]:
                    st.download_button(
                        label=f"📥 Ready to TCF2 ({len(ready_df_tcf2)} cabs)",
                        data=buf_ready2.getvalue(),
                        file_name="TCF2_Ready_to_Build.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_ready_tcf2"
                    )
            
            # Blocked with Reason Excel (includes both 🚫 Blocked and ⚠️ PBS Hold)
            blocked_df_tcf2 = filtered_df_tcf2[filtered_df_tcf2['STATUS'].isin(['🚫 Blocked', '⚠️ PBS Hold'])][display_cols].copy()
            if not blocked_df_tcf2.empty:
                buf_blocked2 = io.BytesIO()
                with pd.ExcelWriter(buf_blocked2, engine='openpyxl') as writer:
                    blocked_df_tcf2.to_excel(writer, index=False, sheet_name='Blocked TCF2')
                    ws = writer.sheets['Blocked TCF2']
                    hdr_fill = PatternFill(start_color='FFD1D1', end_color='FFD1D1', fill_type='solid')
                    hdr_font = Font(name='Calibri', size=11, bold=True, color='5C1D1B')
                    thin_b = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
                    for c in range(1, len(blocked_df_tcf2.columns) + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_b
                    for r in range(2, len(blocked_df_tcf2) + 2):
                        for c in range(1, len(blocked_df_tcf2.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_b
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(name='Calibri', size=10)
                            # Highlight blocking reason column in light red
                            if c == 5:  # BLOCKING_REASON column
                                cell.fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
                with dl_cols_tcf2[1]:
                    st.download_button(
                        label=f"📥 Blocked with Reason ({len(blocked_df_tcf2)} cabs)",
                        data=buf_blocked2.getvalue(),
                        file_name="TCF2_Blocked_with_Reason.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_blocked_tcf2"
                    )
            
    # Subtab 2: Total Float Details
    with line_subtabs_tcf2[1]:
        render_total_float_details_view(temp_float_df, default_line="TCF2")

# ----------------- TAB 4: TOTAL FLOAT DETAILS & SEARCH -----------------
with tcf_tabs[3]:
    render_total_float_details_view(temp_float_df, default_line="All")

# ----------------- TAB 5: QUALITY HOLD REGISTRY -----------------
with tcf_tabs[4]:
    st.markdown("### 📋 Quality Hold Registry")
    st.markdown("Overview of all vehicles currently placed on quality hold in the Paint Shop and PBS buffer.")
    
    # 1. Compute PBS Quality Holds
    if not pbs_on_hold.empty:
        pbs_on_hold_cleaned = pbs_on_hold.drop_duplicates(subset=['BIW NUMBER']).sort_values(by=['SHOP', 'PBS LIFT'], ascending=[True, True]).copy()
        if bom_df is not None and not bom_df.empty:
            vc_to_engine = dict(zip(bom_df['Short Vehicle Code'].astype(str).str.strip(), bom_df['Engine'].astype(str).str.strip()))
            short_vcs = pbs_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str[:9]
            mapped_engines = short_vcs.map(vc_to_engine)
            pbs_on_hold_cleaned['Model'] = mapped_engines.map(engine_to_model)
        else:
            pbs_on_hold_cleaned['Model'] = pd.Series(dtype='object', index=pbs_on_hold_cleaned.index)
            
        if 'PRODUCT' in pbs_on_hold_cleaned.columns:
            is_tayrona = pbs_on_hold_cleaned['PRODUCT'].astype(str).str.strip().str.upper().str.contains('TAYRONA') | \
                         pbs_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str.startswith('54831927A')
        else:
            is_tayrona = pbs_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str.startswith('54831927A')
            
        pbs_on_hold_cleaned['Model'] = np.where(is_tayrona, 'SAFARI EV', pbs_on_hold_cleaned['Model'])
        pbs_on_hold_cleaned['Model'] = pbs_on_hold_cleaned['Model'].fillna('—')
            
        colour_col = None
        for col in pbs_on_hold_cleaned.columns:
            if str(col).strip().upper() in ['COLOUR', 'COLOR']:
                colour_col = col
                break
        if colour_col:
            pbs_on_hold_cleaned['Colour'] = pbs_on_hold_cleaned[colour_col].fillna('—')
        else:
            pbs_on_hold_cleaned['Colour'] = '—'
    else:
        pbs_on_hold_cleaned = pd.DataFrame()

    # 2. Compute Paint Shop Quality Holds (from PTCED to before PBS Lift)
    if float_df is not None and not float_df.empty:
        ps_hold_mask = float_df['PBS LIFT'].isna() & float_df['PTCED'].notna() & float_df['HOLD BY'].notna() & (float_df['HOLD BY'].astype(str).str.strip() != '') & (float_df['HOLD BY'].astype(str).str.strip() != 'nan')
        paintshop_on_hold = float_df[ps_hold_mask].copy()
    else:
        paintshop_on_hold = pd.DataFrame()

    if not paintshop_on_hold.empty:
        paintshop_on_hold_cleaned = paintshop_on_hold.drop_duplicates(subset=['BIW NUMBER']).sort_values(by=['SHOP', 'PTCED'], ascending=[True, True]).copy()
        if bom_df is not None and not bom_df.empty:
            vc_to_engine = dict(zip(bom_df['Short Vehicle Code'].astype(str).str.strip(), bom_df['Engine'].astype(str).str.strip()))
            short_vcs = paintshop_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str[:9]
            mapped_engines = short_vcs.map(vc_to_engine)
            paintshop_on_hold_cleaned['Model'] = mapped_engines.map(engine_to_model)
        else:
            paintshop_on_hold_cleaned['Model'] = pd.Series(dtype='object', index=paintshop_on_hold_cleaned.index)
            
        if 'PRODUCT' in paintshop_on_hold_cleaned.columns:
            is_tayrona = paintshop_on_hold_cleaned['PRODUCT'].astype(str).str.strip().str.upper().str.contains('TAYRONA') | \
                         paintshop_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str.startswith('54831927A')
        else:
            is_tayrona = paintshop_on_hold_cleaned['VEHICLE CODE'].astype(str).str.strip().str.startswith('54831927A')
            
        paintshop_on_hold_cleaned['Model'] = np.where(is_tayrona, 'SAFARI EV', paintshop_on_hold_cleaned['Model'])
        paintshop_on_hold_cleaned['Model'] = paintshop_on_hold_cleaned['Model'].fillna('—')
            
        colour_col = None
        for col in paintshop_on_hold_cleaned.columns:
            if str(col).strip().upper() in ['COLOUR', 'COLOR']:
                colour_col = col
                break
        if colour_col:
            paintshop_on_hold_cleaned['Colour'] = paintshop_on_hold_cleaned[colour_col].fillna('—')
        else:
            paintshop_on_hold_cleaned['Colour'] = '—'
    else:
        paintshop_on_hold_cleaned = pd.DataFrame()

    tot_pbs_h = len(pbs_on_hold_cleaned)
    tot_ps_h = len(paintshop_on_hold_cleaned)

    metric_cols = st.columns(3)
    metric_cols[0].metric("PBS Buffer Holds", f"{tot_pbs_h} cabs", help="Cabs in PBS Buffer on Quality hold")
    metric_cols[1].metric("Paint Shop Holds", f"{tot_ps_h} cabs", help="Cabs in Paint Shop (PTCED to before PBS Lift) on Quality hold")
    metric_cols[2].metric("Total Quality Holds", f"{tot_pbs_h + tot_ps_h} cabs")
    
    st.markdown("---")
    
    # --- SECTION 1: PBS QUALITY HOLDS ---
    st.markdown("### 🛑 PBS Quality Holds Registry (PBS Buffer)")
    if pbs_on_hold_cleaned.empty:
        st.success("🎉 Excellent! No cabs currently on quality hold in the PBS buffer.")
    else:
        st.warning(f"⚠️ {tot_pbs_h} unique cabs are currently held in PBS and skipped from Clear-to-Build checks.")
        
        display_hold_cols = [col for col in ['BIW NUMBER', 'Model', 'Colour', 'VIN', 'VEHICLE CODE', 'SHOP', 'HOLD BY', 'REASONS S', 'PBS LIFT'] if col in pbs_on_hold_cleaned.columns]
        st.dataframe(
            pbs_on_hold_cleaned[display_hold_cols],
            use_container_width=True,
            hide_index=True
        )
        
        import io
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            pbs_on_hold_cleaned[display_hold_cols].to_excel(writer, index=False, sheet_name='PBS Quality Holds')
        st.download_button(
            label="📥 Export PBS Quality Holds to Excel",
            data=excel_buffer.getvalue(),
            file_name="pbs_quality_holds.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="export_pbs_quality_holds"
        )

    st.markdown("---")

    # --- SECTION 2: PAINT SHOP QUALITY HOLDS ---
    st.markdown("### 🎨 Paint Shop Quality Holds Registry (PTCED to Before PBS Lift)")
    if paintshop_on_hold_cleaned.empty:
        st.success("🎉 Excellent! No cabs currently on quality hold in the Paint Shop stage.")
    else:
        st.warning(f"⚠️ {tot_ps_h} unique cabs are currently held in Paint Shop (between PTCED and before PBS Lift).")
        
        display_ps_cols = [col for col in ['BIW NUMBER', 'Model', 'Colour', 'VIN', 'VEHICLE CODE', 'SHOP', 'HOLD BY', 'REASONS S', 'PTCED', 'SEALANT', 'TOPCOAT'] if col in paintshop_on_hold_cleaned.columns]
        st.dataframe(
            paintshop_on_hold_cleaned[display_ps_cols],
            use_container_width=True,
            hide_index=True
        )
        
        import io
        excel_buffer_ps = io.BytesIO()
        with pd.ExcelWriter(excel_buffer_ps, engine='openpyxl') as writer:
            paintshop_on_hold_cleaned[display_ps_cols].to_excel(writer, index=False, sheet_name='Paint Shop Quality Holds')
        st.download_button(
            label="📥 Export Paint Shop Quality Holds to Excel",
            data=excel_buffer_ps.getvalue(),
            file_name="paintshop_quality_holds.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="export_paintshop_quality_holds"
        )
        


# ----------------- TAB 6: TELEGRAM DISPATCHER -----------------
with tcf_tabs[5]:
    st.markdown("### 📱 Telegram Report Dispatcher")
    st.markdown("Send live shift production summaries and material shortage alerts directly to Telegram channels, groups, or planners.")

    # Collapsed credentials expander (hidden by default)
    with st.expander("⚙️ Telegram Bot Settings (Click to Edit Token / Chat ID)", expanded=False):
        st.markdown("<small style='color:#8896AB'>Configure Telegram Bot API Key & Target Chat ID</small>", unsafe_allow_html=True)
        
        input_token = st.text_input("Telegram Bot API Token:", value=st.session_state.telegram_token, type="password", key="input_telegram_token")
        input_chat_id = st.text_input("Telegram Chat ID / Group ID:", value=st.session_state.telegram_chat_id, key="input_telegram_chat_id")
        
        st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
        btn_t1, btn_t2 = st.columns(2)
        with btn_t1:
            if st.button("💾 Save Credentials", type="primary", use_container_width=True, key="save_tg_creds_btn"):
                st.session_state.telegram_token = input_token.strip()
                st.session_state.telegram_chat_id = input_chat_id.strip()
                try:
                    dl.save_metadata('telegram_token', input_token.strip())
                    dl.save_metadata('telegram_chat_id', input_chat_id.strip())
                except Exception:
                    pass
                st.toast("💾 Telegram credentials saved to database!", icon="💾")

        with btn_t2:
            if st.button("🧪 Send Test Msg", type="secondary", use_container_width=True, key="test_tg_creds_btn"):
                test_msg = "<b>🤖 TML Planner Dashboard Connected!</b>\n\nTelegram Bot integration successfully verified."
                success, status_lbl = dl.send_telegram_message(input_token.strip(), input_chat_id.strip(), test_msg)
                if success:
                    st.success(status_lbl)
                else:
                    st.error(status_lbl)

    st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
    st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("#### 📊 Report Message Preview & Dispatch")
        st.markdown("<small style='color:#8896AB'>Select report type to preview and dispatch via Telegram</small>", unsafe_allow_html=True)

        # Helper function for detailed blocked reasons summary
        def format_blocked_summary(alloc_df):
            if alloc_df is None or alloc_df.empty:
                return "0"
            
            blocked_df = alloc_df[alloc_df['STATUS'].astype(str).str.contains('Blocked|Hold')].copy()
            tot_blocked = len(blocked_df)
            if tot_blocked == 0:
                return "0"
                
            reason_counts = {}
            for idx, r in blocked_df.iterrows():
                reason = str(r.get('BLOCKING_REASON', 'Unspecified')).strip()
                if 'Shortage:' in reason:
                    clean_part = reason.replace('Shortage:', '').split('(')[0].strip()
                    tokens = clean_part.split()
                    if len(tokens) >= 2 and tokens[-1].isdigit():
                        clean_r = " ".join(tokens[:-1])
                    else:
                        clean_r = clean_part
                elif 'Quality' in reason or 'Hold' in reason or 'QA' in reason:
                    clean_r = "QA Hold"
                elif 'BOM' in reason:
                    clean_r = "BOM Incomplete"
                else:
                    clean_r = reason[:20]
                    
                reason_counts[clean_r] = reason_counts.get(clean_r, 0) + 1
                
            breakdown_items = [f"{cnt} {r_lbl}" for r_lbl, cnt in reason_counts.items()]
            breakdown_str = ", ".join(breakdown_items)
            return f"{tot_blocked} ({breakdown_str})"

        # ----------------- REPORT 1: TCF1 & TCF2 PPC PLANNER DASHBOARD REPORT -----------------
        t1_vin_gen = int(tcf1_drops['VIN_Count'].sum()) if (tcf1_drops is not None and not tcf1_drops.empty and 'VIN_Count' in tcf1_drops.columns) else (len(tcf1_drops) if tcf1_drops is not None else 0)
        t2_vin_gen = int(tcf2_drops['VIN_Count'].sum()) if (tcf2_drops is not None and not tcf2_drops.empty and 'VIN_Count' in tcf2_drops.columns) else (len(tcf2_drops) if tcf2_drops is not None else 0)

        t1_ready = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf1_alloc_df.empty else 0
        t1_shortages_cnt = len(tcf1_alloc_df[tcf1_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf1_alloc_df.empty else 0
        t1_blocked_summary = format_blocked_summary(tcf1_alloc_df)

        t2_ready = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '✅ Ready for TCF']) if not tcf2_alloc_df.empty else 0
        t2_shortages_cnt = len(tcf2_alloc_df[tcf2_alloc_df['STATUS'] == '🚫 Blocked']) if not tcf2_alloc_df.empty else 0
        t2_blocked_summary = format_blocked_summary(tcf2_alloc_df)

        t1_pbs_total = len(pbs_all[pbs_all['SHOP'] == 'TCF1']) if (float_df is not None and not float_df.empty and 'pbs_all' in locals() and not pbs_all.empty) else len(tcf1_alloc_df)
        t1_qa_hold = len(pbs_on_hold[pbs_on_hold['SHOP'] == 'TCF1']) if ('pbs_on_hold' in locals() and not pbs_on_hold.empty) else 0

        t2_pbs_total = len(pbs_all[pbs_all['SHOP'] == 'TCF2']) if (float_df is not None and not float_df.empty and 'pbs_all' in locals() and not pbs_all.empty) else len(tcf2_alloc_df)
        t2_qa_hold = len(pbs_on_hold[pbs_on_hold['SHOP'] == 'TCF2']) if ('pbs_on_hold' in locals() and not pbs_on_hold.empty) else 0

        # Nova Total Float Qty (124) and Nova Today VIN Qty (30)
        nova_total_float_qty = 0
        if 'paint_summary_dict' in locals() and paint_summary_dict and 'PUNCH.EV' in paint_summary_dict:
            nova_total_float_qty = paint_summary_dict['PUNCH.EV'].get('TOTAL FLOAT', 0)
        elif float_df is not None and not float_df.empty:
            is_nova_mask = (
                float_df['PRODUCT'].astype(str).str.upper().str.contains('NOVA') |
                float_df['VEHICLE CODE'].astype(str).str.startswith('5468')
            )
            nova_total_float_qty = len(float_df[is_nova_mask])

        nova_vin_qty = 0
        if tcf1_drops is not None and not tcf1_drops.empty:
            vin_match_col = 'Model_Family' if 'Model_Family' in tcf1_drops.columns else ('Model' if 'Model' in tcf1_drops.columns else None)
            if vin_match_col:
                nova_drops = tcf1_drops[tcf1_drops[vin_match_col] == 'PUNCH.EV']
            else:
                nova_drops = tcf1_drops[tcf1_drops['VEHICLE CODE'].astype(str).str.startswith('5468')]
            nova_vin_qty = int(nova_drops['VIN_Count'].sum()) if 'VIN_Count' in nova_drops.columns else len(nova_drops)
        elif not tcf1_alloc_df.empty:
            nova_cabs = tcf1_alloc_df[
                tcf1_alloc_df['Model'].astype(str).str.contains('Nova|Punch EV|PUNCH.EV', case=False, na=False, regex=True) |
                tcf1_alloc_df['VEHICLE CODE'].astype(str).str.startswith('5468')
            ]
            nova_vin_qty = len(nova_cabs)
        tcf1_drop_val = 0
        tcf1_paint_val = 0
        tcf2_drop_val = 0
        tcf2_paint_val = 0
        t60_val = 0
        t40_val = 0
        
        if shop_totals:
            tcf1_drop_val = int(shop_totals.get('TCF DROP', 0))
            tcf2_drop_val = int(shop_totals.get('TCF2 DROP', 0))
            t60_val = int(shop_totals.get('T60', 0))
            t40_val = int(shop_totals.get('T40', 0))
            
        if shop_vehicles_df is not None and not shop_vehicles_df.empty:
            tcf1_m = shop_vehicles_df[shop_vehicles_df['Model'].isin(['PUNCH', 'PUNCH Exports', 'PUNCH EV'])]
            tcf2_m = shop_vehicles_df[shop_vehicles_df['Model'].isin(['HARRIER EV', 'SAFARI', 'HARRIER'])]
            
            if not tcf1_m.empty:
                tcf1_paint_val = int(tcf1_m['Paint Lifting'].sum())
                sum_t60 = int(tcf1_m['T60'].sum())
                if sum_t60 > 0 or t60_val == 0:
                    t60_val = sum_t60
            if not tcf2_m.empty:
                tcf2_paint_val = int(tcf2_m['Paint Lifting'].sum())
                sum_t40 = int(tcf2_m['T40'].sum())
                if sum_t40 > 0 or t40_val == 0:
                    t40_val = sum_t40

        now_str_r1 = format_ist_now("%d-%m-%Y %I:%M %p")
        tg_report_1_text = f"📊 TCF1 & TCF2 PPC REPORT\n"
        tg_report_1_text += f"⏰ Report Time: {now_str_r1}\n\n"
        tg_report_1_text += f"🏭 TCF1 LINE (Punch / Punch EV):\n"
        tg_report_1_text += f" • 🚜 Dropping: {tcf1_drop_val}\n"
        tg_report_1_text += f" • 🎨 Paint Lifting: {tcf1_paint_val}\n"
        tg_report_1_text += f" • ⏱️ T60: {t60_val}\n"
        tg_report_1_text += f" • ✅ Ready for TCF: {t1_ready}\n"
        tg_report_1_text += f" • 🚫 Shortages: {t1_blocked_summary}\n\n"
        tg_report_1_text += f"🏭 TCF2 LINE (Harrier / Safari):\n"
        tg_report_1_text += f" • 🚜 Dropping: {tcf2_drop_val}\n"
        tg_report_1_text += f" • 🎨 Paint Lifting: {tcf2_paint_val}\n"
        tg_report_1_text += f" • ⏱️ T40: {t40_val}\n"
        tg_report_1_text += f" • ✅ Ready for TCF: {t2_ready}\n"
        tg_report_1_text += f" • 🚫 Shortages: {t2_blocked_summary}\n\n"
        tg_report_1_text += f"📦 PBS Cab details:\n\n"
        tg_report_1_text += f" • 🚜 TCF1: {t1_pbs_total} ({t1_qa_hold} QA hold, {t1_shortages_cnt} Material Shortage)\n"
        tg_report_1_text += f" • 🚜 TCF2: {t2_pbs_total} ({t2_qa_hold} QA hold, {t2_shortages_cnt} Material Shortage)\n\n"
        tg_report_1_text += f"⚡ Punch EV (Nova) VIN Qty: {nova_vin_qty}\n"
        tg_report_1_text += f"⏲️ Current Material clearance after 06:30 AM:\n"

        if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None:
            for idx, r_n in st.session_state.nova_materials_df.iterrows():
                m_name = str(r_n['Material']).strip()
                m_name_clean = m_name.replace('Craddle', 'Cradle')
                if 'Tube Frame' in m_name_clean and 'Tube Frame (' not in m_name_clean:
                    m_name_clean = m_name_clean.replace('Tube Frame(', 'Tube Frame (')
                if 'new_nova_input_vals' in locals() and m_name in new_nova_input_vals:
                    open_qty = int(new_nova_input_vals[m_name])
                else:
                    open_qty = int(r_n['Clearance Qty'])
                
                if open_qty < nova_vin_qty:
                    defic = nova_vin_qty - open_qty
                    icon = "🔴"
                    tg_report_1_text += f" • {icon} SHORTAGE: {m_name_clean}: {open_qty} (VIN Demand: {nova_vin_qty}, Deficit: -{defic})\n"
                elif open_qty == 0:
                    icon = "🔴"
                    tg_report_1_text += f" • {icon} {m_name_clean}: {open_qty}\n"
                else:
                    icon = "🟢"
                    tg_report_1_text += f" • {icon} {m_name_clean}: {open_qty}\n"

        if 'model_shortages_df' in st.session_state and st.session_state.model_shortages_df is not None and not st.session_state.model_shortages_df.empty:
            tg_report_1_text += f"\n📦 Model-Wise Material Shortage Alerts:\n"
            for idx_ms, r_ms in st.session_state.model_shortages_df.iterrows():
                ms_mod = str(r_ms['Model']).strip()
                ms_trm = str(r_ms.get('Trims', 'All Trims')).strip()
                ms_part = str(r_ms['Part Name']).strip()
                ms_c_qty = int(r_ms['Clearance Qty'])
                ms_d_qty = get_demand_qty_for_model_trims(ms_mod, ms_trm, tcf1_drops, tcf2_drops)
                if ms_c_qty < ms_d_qty:
                    ms_def = ms_d_qty - ms_c_qty
                    tg_report_1_text += f" • 🔴 SHORTAGE: {ms_mod} [{ms_trm}] - {ms_part}: {ms_c_qty} (Demand: {ms_d_qty}, Deficit: -{ms_def})\n"
                else:
                    tg_report_1_text += f" • 🟢 {ms_mod} [{ms_trm}] - {ms_part}: {ms_c_qty} (Demand: {ms_d_qty})\n"

        # ----------------- REPORT 2: PUNCH EV (NOVA) EXECUTIVE STATUS REPORT -----------------
        now_time_r2 = format_ist_nearest_15min().replace(" ", "")  # e.g. "04.15PM"
        nova_paint_float_cnt = 0
        nova_pbs_cnt = 0

        if 'paint_summary_dict' in locals() and paint_summary_dict and 'PUNCH.EV' in paint_summary_dict:
            m_nova = paint_summary_dict['PUNCH.EV']
            nova_paint_float_cnt = m_nova.get('TOTAL FLOAT', 0)
            nova_pbs_cnt = m_nova.get('PBS FLOAT', 0)
        elif float_df is not None and not float_df.empty:
            is_nova_mask = (
                float_df['PRODUCT'].astype(str).str.upper().str.contains('NOVA') |
                float_df['VEHICLE CODE'].astype(str).str.startswith('5468')
            )
            nova_float_cabs = float_df[is_nova_mask]
            nova_paint_float_cnt = len(nova_float_cabs)
            nova_pbs_cnt = len(nova_float_cabs[nova_float_cabs['PBS LIFT'].notna()])

        tg_report_2_text = f"Dear sir,\n\n"
        tg_report_2_text += f"Nova Status as on {now_time_r2}\n\n"
        tg_report_2_text += f"VIN: {nova_vin_qty}\n\n"
        tg_report_2_text += f"Current Paint Float: {nova_paint_float_cnt}\n"
        tg_report_2_text += f"PBS: {nova_pbs_cnt}\n\n"
        tg_report_2_text += f"Today's Material Clearance (after 06:30 AM):\n\n"

        if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None:
            for idx, r_n in st.session_state.nova_materials_df.iterrows():
                m_name = str(r_n['Material']).strip()
                m_name_clean = m_name.replace('Craddle', 'Cradle')
                if 'Tube Frame' in m_name_clean and 'Tube Frame (' not in m_name_clean:
                    m_name_clean = m_name_clean.replace('Tube Frame(', 'Tube Frame (')
                if 'new_nova_input_vals' in locals() and m_name in new_nova_input_vals:
                    open_qty = int(new_nova_input_vals[m_name])
                else:
                    open_qty = int(r_n['Clearance Qty'])
                
                # Add * to lower stock qty only (stock < VIN Qty)
                if open_qty < nova_vin_qty:
                    defic = nova_vin_qty - open_qty
                    tg_report_2_text += f"🚨 *SHORTAGE: {m_name_clean}: {open_qty} (Demand: {nova_vin_qty}, Deficit: -{defic})*\n"
                else:
                    tg_report_2_text += f"{m_name_clean}: {open_qty}\n"

        if 'model_shortages_df' in st.session_state and st.session_state.model_shortages_df is not None and not st.session_state.model_shortages_df.empty:
            tg_report_2_text += f"\nModel Shortages:\n"
            for idx_ms, r_ms in st.session_state.model_shortages_df.iterrows():
                ms_mod = str(r_ms['Model']).strip()
                ms_trm = str(r_ms.get('Trims', 'All Trims')).strip()
                ms_part = str(r_ms['Part Name']).strip()
                ms_c_qty = int(r_ms['Clearance Qty'])
                ms_d_qty = get_demand_qty_for_model_trims(ms_mod, ms_trm, tcf1_drops, tcf2_drops)
                if ms_c_qty < ms_d_qty:
                    ms_def = ms_d_qty - ms_c_qty
                    tg_report_2_text += f"🚨 *SHORTAGE: {ms_mod} [{ms_trm}] - {ms_part}: {ms_c_qty} (Demand: {ms_d_qty}, Deficit: -{ms_def})*\n"
                else:
                    tg_report_2_text += f"{ms_mod} [{ms_trm}] - {ms_part}: {ms_c_qty}\n"
        # Build Report 3: TCF Dropping vs. Paint Lifting Status
        tcf1_gap_str = f"\n*Gap:{tcf1_drop_val - tcf1_paint_val:02d}*" if tcf1_drop_val >= tcf1_paint_val else ""
        tcf2_gap_str = f"\n *Gap: {tcf2_drop_val - tcf2_paint_val:02d}* " if tcf2_drop_val >= tcf2_paint_val else ""
        
        tcf1_pbs_detail_str = f"{t1_ready} cabs ({t1_qa_hold} QA hold, {t1_shortages_cnt} Material Shortage)"
        tcf2_pbs_detail_str = f"{t2_ready} cabs ({t2_qa_hold} QA hold, {t2_shortages_cnt} Material Shortage)"

        tg_report_3_text = f"""Dear Sir

TCF Dropping vs. Paint Lifting Status:

TCF1:
Dropping: {tcf1_drop_val}
Paint Lifting: {tcf1_paint_val}{tcf1_gap_str}

TCF2:
Dropping : {tcf2_drop_val}
Paint Lifting: {tcf2_paint_val}{tcf2_gap_str}

Dropping Float:
*T60: {t60_val}*
*T40: {t40_val}*

Available Cabs for VIN Generation:

TCF1: {tcf1_pbs_detail_str}

TCF2: {tcf2_pbs_detail_str}"""

        # ----------------- QUICK ACTIONS: SEND BOTH SCHEDULED REPORTS -----------------
        # Moved to the top so the most common one-click action doesn't require
        # scrolling past the full custom-dispatcher form to find it.
        with st.container(border=True):
            qa_col1, qa_col2 = st.columns([3, 1.3])
            with qa_col1:
                st.markdown("##### ⚡ Quick Action")
                st.caption("Send the TCF1 & TCF2 PPC Report and the Nova Status Report together, in one tap.")
            with qa_col2:
                if st.button("🚀 Send BOTH Reports Now", type="secondary", use_container_width=True, key="send_both_tg_reports_btn"):
                    ok1, res_msg1 = dl.send_telegram_message(st.session_state.telegram_token, st.session_state.telegram_chat_id, tg_report_1_text)
                    ok2, res_msg2 = dl.send_telegram_message(st.session_state.telegram_token, st.session_state.telegram_chat_id, tg_report_2_text)
                    if ok1 and ok2:
                        st.toast("🚀 Both reports successfully sent to Telegram!", icon="🚀")
                        st.success("✅ Both reports (TCF1 & TCF2 PPC Report & Nova Status Report) dispatched successfully!")
                    else:
                        if not ok1:
                            st.error(f"Report 1 Error: {res_msg1}")
                        if not ok2:
                            st.error(f"Report 2 Error: {res_msg2}")

        st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)
        st.caption("📅 **Scheduled Report Templates** (first 3 tabs) — preview and send one at a time · 💬 **Manual Dispatcher** (last tab) — free-text, screenshots, or file attachments")

        # Tabbed preview & dispatch options
        report_tab1, report_tab2, report_tab3, report_tab4 = st.tabs([
            "📅 Report 1: TCF1 & TCF2 PPC Report",
            "📅 Report 2: Punch EV (Nova) Status Report",
            "📅 Report 3: TCF Dropping vs. Paint Lifting Status",
            "💬 Custom Telegram Dispatcher"
        ])

        with report_tab1:
            st.markdown("##### 📊 TCF1 & TCF2 PPC Report Preview")
            st.markdown(f"<div style='background: rgba(15, 23, 42, 0.05); border-radius: 8px; padding: 14px; font-family: monospace; font-size: 13px; white-space: pre-wrap; word-break: break-all;'>{tg_report_1_text}</div>", unsafe_allow_html=True)
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            if st.button("🚀 Send TCF1 & TCF2 Report to Telegram", type="primary", use_container_width=True, key="send_tg_report_1_btn"):
                ok, res_msg = dl.send_telegram_message(st.session_state.telegram_token, st.session_state.telegram_chat_id, tg_report_1_text)
                if ok:
                    st.toast("🚀 TCF1 & TCF2 Report successfully sent to Telegram!", icon="🚀")
                    st.success("✅ TCF1 & TCF2 Report dispatched successfully!")
                else:
                    st.error(res_msg)

        with report_tab2:
            st.markdown("##### ⚡ Punch EV (Nova) Status Report Preview")
            st.markdown(f"<div style='background: rgba(15, 23, 42, 0.05); border-radius: 8px; padding: 14px; font-family: monospace; font-size: 13px; white-space: pre-wrap; word-break: break-all;'>{tg_report_2_text}</div>", unsafe_allow_html=True)
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            if st.button("🚀 Send Nova Status Report to Telegram", type="primary", use_container_width=True, key="send_tg_report_2_btn"):
                ok, res_msg = dl.send_telegram_message(st.session_state.telegram_token, st.session_state.telegram_chat_id, tg_report_2_text)
                if ok:
                    st.toast("🚀 Nova Status Report successfully sent to Telegram!", icon="🚀")
                    st.success("✅ Nova Status Report dispatched successfully!")
                else:
                    st.error(res_msg)

        with report_tab3:
            st.markdown("##### 🏭 TCF Dropping vs. Paint Lifting Status Report Preview")
            st.markdown(f"<div style='background: rgba(15, 23, 42, 0.05); border-radius: 8px; padding: 14px; font-family: monospace; font-size: 13px; white-space: pre-wrap; word-break: break-all;'>{tg_report_3_text}</div>", unsafe_allow_html=True)
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            if st.button("🚀 Send Dropping vs. Paint Lifting Report to Telegram", type="primary", use_container_width=True, key="send_tg_report_3_btn"):
                ok, res_msg = dl.send_telegram_message(st.session_state.telegram_token, st.session_state.telegram_chat_id, tg_report_3_text)
                if ok:
                    st.toast("🚀 Dropping vs. Paint Lifting Report successfully sent to Telegram!", icon="🚀")
                    st.success("✅ Dropping vs. Paint Lifting Report dispatched successfully!")
                else:
                    st.error(res_msg)

        with report_tab4:
            st.markdown("##### 💬 Custom / Manual Telegram Dispatcher")
            st.markdown("<small style='color:#8896AB'>Paste any custom text message, paste screenshots directly from clipboard, or attach Excel / document files to send to your mobile via Telegram Bot.</small>", unsafe_allow_html=True)
            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

            custom_text = st.text_area("✍️ Custom Text Message (Optional if uploading file/image):", height=130, placeholder="Type or paste text message here...", key="custom_tg_text_input")
            
            c_col1, c_col2 = st.columns(2)
            img_bytes_to_send = None
            img_filename_to_send = "screenshot.png"

            with c_col1:
                st.markdown("##### 🖼️ Image / Screenshot Attachment")
                st.markdown("<small style='color:#8896AB'>Capture with Snipping Tool (Win + Shift + S), then click below:</small>", unsafe_allow_html=True)
                
                pasted_img_res = paste_image_button(
                    label="📋 Paste Screenshot from Clipboard",
                    background_color="#2563EB",
                    text_color="#FFFFFF",
                    hover_background_color="#1D4ED8",
                    key="custom_tg_paste_btn"
                )
                
                custom_img = st.file_uploader("Or Upload Image File", type=["png", "jpg", "jpeg", "webp"], key="custom_tg_img_uploader")

                if pasted_img_res and pasted_img_res.image_data is not None:
                    st.success("✅ Screenshot captured from Clipboard!")
                    st.image(pasted_img_res.image_data, caption="Clipboard Screenshot Preview", use_column_width=True)
                    img_buf = io.BytesIO()
                    pasted_img_res.image_data.save(img_buf, format="PNG")
                    img_bytes_to_send = img_buf.getvalue()
                elif custom_img is not None:
                    img_bytes_to_send = custom_img.getvalue()
                    img_filename_to_send = custom_img.name

            with c_col2:
                st.markdown("##### 📁 Excel / Document / Image File Attachment (Option 2)")
                st.markdown("<small style='color:#8896AB'>Upload ANY files (Excel, PDF, Images PNG/JPG, Word, CSV, ZIP, etc.):</small>", unsafe_allow_html=True)
                custom_docs = st.file_uploader("Upload Files (Excel, PDF, Images, Documents, etc.)", accept_multiple_files=True, key="custom_tg_doc_uploader")

            st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
            if st.button("🚀 Send Custom Message / Files via Telegram Bot", type="primary", use_container_width=True, key="send_custom_tg_msg_btn"):
                has_text = bool(custom_text and custom_text.strip())
                has_img = img_bytes_to_send is not None
                has_docs = bool(custom_docs)

                if not has_text and not has_img and not has_docs:
                    st.warning("⚠️ Please enter a text message, paste a screenshot, or select file(s) to send.")
                else:
                    bot_token = st.session_state.get('telegram_token', '')
                    chat_id = st.session_state.get('telegram_chat_id', '')

                    successes = []
                    errors = []
                    caption_used = False

                    # 1. Send Text if provided (and no file attachments)
                    if has_text and not has_img and not has_docs:
                        ok, msg = dl.send_telegram_message(bot_token, chat_id, custom_text.strip())
                        if ok:
                            successes.append("Custom text message sent successfully!")
                        else:
                            errors.append(f"Text Message Error: {msg}")

                    # 2. Send Col 1 / Clipboard Image if provided
                    if has_img:
                        caption_str = custom_text.strip() if (has_text and not caption_used) else ""
                        ok, msg = dl.send_telegram_photo(bot_token, chat_id, img_bytes_to_send, caption=caption_str, filename=img_filename_to_send)
                        if ok:
                            successes.append(f"Image ({img_filename_to_send}) sent successfully!")
                            caption_used = True
                        else:
                            errors.append(f"Image Error: {msg}")

                    # 3. Send Files from Col 2 (Images, Excel, PDF, Word, etc.)
                    if has_docs:
                        for doc_file in custom_docs:
                            f_bytes = doc_file.getvalue()
                            f_name = doc_file.name
                            caption_str = custom_text.strip() if (has_text and not caption_used) else ""
                            
                            # Auto-detect if file is an image format
                            ext = os.path.splitext(f_name)[1].lower()
                            if ext in ['.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp']:
                                ok, msg = dl.send_telegram_photo(bot_token, chat_id, f_bytes, caption=caption_str, filename=f_name)
                                if ok:
                                    successes.append(f"Image ({f_name}) sent successfully!")
                                    caption_used = True
                                else:
                                    errors.append(f"File Error ({f_name}): {msg}")
                            else:
                                ok, msg = dl.send_telegram_document(bot_token, chat_id, f_bytes, caption=caption_str, filename=f_name)
                                if ok:
                                    successes.append(f"File ({f_name}) sent successfully!")
                                    caption_used = True
                                else:
                                    errors.append(f"File Error ({f_name}): {msg}")

                    if successes:
                        for s_msg in successes:
                            st.toast(f"🚀 {s_msg}", icon="🚀")
                            st.success(f"✅ {s_msg}")
                    if errors:
                        for e_msg in errors:
                            st.error(f"❌ {e_msg}")

        st.markdown("<div style='height: 12px;'></div>", unsafe_allow_html=True)

# ----------------- TAB 1: SUMMARY REPORT & EXCEL DOWNLOAD -----------------
with tcf_tabs[0]:
    # --- SECTION -1: BOM COMPLETENESS ALERT & QUICK-ENTRY ---
    if 'missing_bom_df' in locals() and not missing_bom_df.empty:
        affected_cabs = int(missing_bom_df['Cab Count'].sum())
        st.error(
            f"🚨 **BOM not available for {len(missing_bom_df)} Short Vehicle Code(s)** "
            f"({affected_cabs} cab(s) affected). These cabs may be miscounted as Ready-to-TCF "
            "with blank/NaN parts until BOM is entered below."
        )
        with st.expander(f"⚠️ Fix Missing/Incomplete BOM — {len(missing_bom_df)} Short VC(s)", expanded=True):
            st.dataframe(missing_bom_df, use_container_width=True, hide_index=True)
            st.markdown("###### ➕ Enter BOM for a Short VC")
            with st.form("missing_bom_entry_form", clear_on_submit=True):
                sel_vc = st.selectbox("Short Vehicle Code", options=missing_bom_df['Short VC'].tolist())
                fb1, fb2, fb3 = st.columns(3)
                with fb1:
                    in_engine = st.text_input("Engine / Battery Part No.")
                with fb2:
                    in_cockpit = st.text_input("Cockpit Part No.")
                with fb3:
                    in_wiring = st.text_input("Front Wiring Part No.")
                submitted = st.form_submit_button("💾 Save BOM Entry")
                if submitted:
                    if not (in_engine.strip() or in_cockpit.strip() or in_wiring.strip()):
                        st.warning("Enter at least one part number before saving.")
                    else:
                        try:
                            dl.save_single_bom_entry(sel_vc, in_wiring, in_cockpit, in_engine)
                            st.success(f"Saved BOM for {sel_vc} to the database. Refreshing report...")
                            st.session_state.run_report = True
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not save BOM entry: {e}")

    # --- SECTION 0: SHOP-WISE PLANT PRODUCTION SUMMARY ---
    if shop_totals is not None or shop_vehicles_df is not None:
        st.markdown("### 🏭 Shop-Wise Plant Production Summary (Daily Report)")
        if shop_totals:
            rep_date = shop_totals.get('Date', shop_totals.get('REPORT DATE', '03/08/2026'))
            cap_col1, cap_col2 = st.columns([2, 2])
            with cap_col1:
                st.caption(f"📅 **Report Date**: {rep_date}")
            with cap_col2:
                last_gen = st.session_state.get('last_generated_at')
                if last_gen:
                    st.caption(f"🕒 **Last Generated**: {last_gen.strftime('%d-%b-%Y %I:%M %p')}")

            # Sticky KPI bar: stays pinned to the top while scrolling through the
            # rest of the (often long) summary report below.
            st.markdown("""
                <style>
                .st-key-sticky_kpi_bar {
                    position: sticky;
                    top: 2.75rem;
                    z-index: 998;
                    background-color: var(--background-color, #0e1117);
                    padding: 0.5rem 0 0.75rem 0;
                    border-bottom: 1px solid rgba(128,128,128,0.25);
                }
                </style>
            """, unsafe_allow_html=True)

            with st.container(key="sticky_kpi_bar"):
                s_kpi1, s_kpi2, s_kpi3, s_kpi4, s_kpi5, s_kpi6 = st.columns(6)
                s_kpi1.metric("TCF1 VIN Count", f"{shop_totals.get('TCF VIN', 0)} cabs")
                s_kpi2.metric("TCF2 VIN Count", f"{shop_totals.get('TCF2 VIN', 0)} cabs")
                s_kpi3.metric("Total TCF Dropping", f"{int(shop_totals.get('TCF DROP', 0)) + int(shop_totals.get('TCF2 DROP', 0))} cabs")
                s_kpi4.metric("Paint Lifting", f"{shop_totals.get('PAINT', 0)} cabs")
                s_kpi5.metric("T60 Count", f"{shop_totals.get('T60', 0)} cabs")
                s_kpi6.metric("T40 Count", f"{shop_totals.get('T40', 0)} cabs")
            
        if shop_vehicles_df is not None and not shop_vehicles_df.empty:
            st.markdown("#### 🚗 Model-Wise Production Matrix (TCF1 & TCF2 Breakdown)")
            
            tcf1_models = ['PUNCH', 'PUNCH Exports', 'PUNCH EV']
            tcf2_models = ['HARRIER EV', 'SAFARI', 'HARRIER']
            
            df1 = shop_vehicles_df[shop_vehicles_df['Model'].isin(tcf1_models)].copy()
            df2 = shop_vehicles_df[shop_vehicles_df['Model'].isin(tcf2_models)].copy()
            
            t1_vin = int(df1['VIN'].sum()) if not df1.empty else 0
            t1_drop = int(df1['Drop'].sum()) if not df1.empty else 0
            t1_paint = int(df1['Paint Lifting'].sum()) if not df1.empty else 0
            t1_t60 = int(df1['T60'].sum()) if not df1.empty else 0
            t1_t40 = int(df1['T40'].sum()) if not df1.empty else 0

            t2_vin = int(df2['VIN'].sum()) if not df2.empty else 0
            t2_drop = int(df2['Drop'].sum()) if not df2.empty else 0
            t2_paint = int(df2['Paint Lifting'].sum()) if not df2.empty else 0
            t2_t60 = int(df2['T60'].sum()) if not df2.empty else 0
            t2_t40 = int(df2['T40'].sum()) if not df2.empty else 0

            g_vin = t1_vin + t2_vin
            g_drop = t1_drop + t2_drop
            g_paint = t1_paint + t2_paint
            g_t60 = t1_t60 + t2_t60
            g_t40 = t1_t40 + t2_t40

            html_table = f"""<style>
.matrix-card {{
    background: rgba(15, 23, 42, 0.02);
    border: 1px solid rgba(226, 232, 240, 0.9);
    border-radius: 14px;
    padding: 12px;
    margin-top: 8px;
    margin-bottom: 16px;
    box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05);
}}
.matrix-table {{
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    border-radius: 10px;
    overflow: hidden;
    font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
}}
.matrix-table th {{
    background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
    color: #38bdf8;
    padding: 12px 16px;
    font-size: 13px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    text-align: center;
    border-bottom: 2px solid #0284c7;
}}
.matrix-table th:first-child {{ text-align: left; padding-left: 20px; }}
.matrix-table td {{
    padding: 11px 16px;
    font-size: 14px;
    color: #1e293b;
    text-align: center;
    border-bottom: 1px solid #e2e8f0;
    font-weight: 600;
}}
.matrix-table td:first-child {{ text-align: left; padding-left: 20px; }}
.tr-tcf1 {{ background-color: #ffffff; }}
.tr-tcf1:hover {{ background-color: #f0f9ff; }}
.tr-tcf2 {{ background-color: #ffffff; }}
.tr-tcf2:hover {{ background-color: #f0fdf4; }}
.tr-tcf1-tot td {{
    background: linear-gradient(90deg, #1e40af 0%, #3b82f6 100%) !important;
    color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 14px !important;
    border-top: 2px solid #1d4ed8 !important;
    border-bottom: 2px solid #1d4ed8 !important;
}}
.tr-tcf2-tot td {{
    background: linear-gradient(90deg, #0f766e 0%, #14b8a6 100%) !important;
    color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 14px !important;
    border-top: 2px solid #0d9488 !important;
    border-bottom: 2px solid #0d9488 !important;
}}
.tr-grand-tot td {{
    background: linear-gradient(90deg, #312e81 0%, #4f46e5 100%) !important;
    color: #fbbf24 !important;
    font-weight: 900 !important;
    font-size: 15px !important;
    letter-spacing: 0.5px;
    border-top: 3px solid #6366f1 !important;
}}
.badge-tcf1 {{
    background: #dbeafe;
    color: #1e40af;
    padding: 3px 8px;
    border-radius: 6px;
    font-size: 11px;
    font-weight: 700;
    display: inline-block;
}}
.badge-tcf2 {{
    background: #ccfbf1;
    color: #0f766e;
    padding: 3px 8px;
    border-radius: 6px;
    font-size: 11px;
    font-weight: 700;
    display: inline-block;
}}
</style>
<div class="matrix-card">
<table class="matrix-table">
<thead>
<tr>
<th>Model</th>
<th>VIN</th>
<th>Drop</th>
<th>Paint Lifting</th>
<th>T60</th>
<th>T40</th>
</tr>
</thead>
<tbody>"""
            
            # TCF1 Rows
            if not df1.empty:
                for _, r in df1.iterrows():
                    html_table += f"""<tr class="tr-tcf1"><td><span class="badge-tcf1">TCF1</span> &nbsp; <b>{r['Model']}</b></td><td>{r['VIN']}</td><td>{r['Drop']}</td><td>{r['Paint Lifting']}</td><td>{r['T60']}</td><td>{r['T40']}</td></tr>"""
            
            # TCF1 Total
            html_table += f"""<tr class="tr-tcf1-tot"><td>🔹 TOTAL TCF1</td><td>{t1_vin}</td><td>{t1_drop}</td><td>{t1_paint}</td><td>{t1_t60}</td><td>{t1_t40}</td></tr>"""

            # TCF2 Rows
            if not df2.empty:
                for _, r in df2.iterrows():
                    html_table += f"""<tr class="tr-tcf2"><td><span class="badge-tcf2">TCF2</span> &nbsp; <b>{r['Model']}</b></td><td>{r['VIN']}</td><td>{r['Drop']}</td><td>{r['Paint Lifting']}</td><td>{r['T60']}</td><td>{r['T40']}</td></tr>"""
            
            # TCF2 Total
            html_table += f"""<tr class="tr-tcf2-tot"><td>🔸 TOTAL TCF2</td><td>{t2_vin}</td><td>{t2_drop}</td><td>{t2_paint}</td><td>{t2_t60}</td><td>{t2_t40}</td></tr>"""

            # Grand Total
            html_table += f"""<tr class="tr-grand-tot"><td>🏆 GRAND TOTAL PLANT</td><td>{g_vin}</td><td>{g_drop}</td><td>{g_paint}</td><td>{g_t60}</td><td>{g_t40}</td></tr></tbody></table></div>"""
            
            st.markdown(html_table, unsafe_allow_html=True)
            
            # Excel export button for Model Wise Matrix
            df1_sub = df1.copy()
            t1_row = pd.DataFrame([{'Model': 'TOTAL TCF1', 'VIN': t1_vin, 'Drop': t1_drop, 'Paint Lifting': t1_paint, 'T60': t1_t60, 'T40': t1_t40}])
            df2_sub = df2.copy()
            t2_row = pd.DataFrame([{'Model': 'TOTAL TCF2', 'VIN': t2_vin, 'Drop': t2_drop, 'Paint Lifting': t2_paint, 'T60': t2_t60, 'T40': t2_t40}])
            gt_row = pd.DataFrame([{'Model': 'GRAND TOTAL PLANT', 'VIN': g_vin, 'Drop': g_drop, 'Paint Lifting': g_paint, 'T60': g_t60, 'T40': g_t40}])
            
            export_matrix_df = pd.concat([df1_sub, t1_row, df2_sub, t2_row, gt_row], ignore_index=True)
            
            buf_matrix = io.BytesIO()
            with pd.ExcelWriter(buf_matrix, engine='openpyxl') as writer:
                export_matrix_df.to_excel(writer, index=False, sheet_name='Production Matrix')
            st.download_button(
                label="📥 Export Model-Wise Production Matrix to Excel",
                data=buf_matrix.getvalue(),
                file_name="Model_Wise_Production_Matrix.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_prod_matrix_btn"
            )
            
        st.markdown("---")

    st.markdown("### 📊 Paint Shop Float Summary")
    st.markdown("""
        This report displays the paint shop buffer status by stage and model, matching the exact layout of the paint shop tracker sheet.
    """)
    
    if paint_summary_dict or (float_df is not None and not float_df.empty):
        # Map product names to internal TCF models
        def get_summary_product_to_model(prod_name):
            prod = str(prod_name).strip().upper()
            if 'HORNBILL' in prod:
                return 'PUNCH'
            elif 'NOVA' in prod:
                return 'PUNCH.EV'
            elif 'ETURNA' in prod:
                return 'HARRIER.EV'
            elif 'GRAVITAS' in prod:
                return 'SAFARI'
            elif 'Q5' in prod:
                return 'HARRIER'
            elif 'TAYRONA' in prod:
                return 'SAFARI.EV'
            return 'UNKNOWN'
            
        def get_row_paint_stage(row):
            return ae.get_detailed_paint_summary_stage(row)
                
        stages_list = [
            'PBS FLOAT', 
            'PBS TO POLISHING', 
            'POLISHING TO TOPCOAT', 
            'TOPCOAT TO WETSANDING G ROOFBLACK', 
            'TOPCOAT TO WETSANDING G FRESH', 
            'WETSANDING G TO SEALANT', 
            'PT ENTRY TO SEALANT', 
            'BIW LIFTING G TO PT', 
            'PT BYPASS'
        ]
        
        tcf1_models = ['PUNCH', 'PUNCH.EV']
        tcf2_models = ['HARRIER.EV', 'SAFARI', 'HARRIER', 'SAFARI.EV']
        
        rows = []
        
        def get_today_vin_count(vgl_df, model_name):
            if vgl_df is None or vgl_df.empty:
                return 0
            col = 'Model_Family' if 'Model_Family' in vgl_df.columns else ('Model' if 'Model' in vgl_df.columns else None)
            if not col:
                return 0
            sub = vgl_df[vgl_df[col] == model_name]
            if sub.empty and 'Model' in vgl_df.columns and col != 'Model':
                sub = vgl_df[vgl_df['Model'] == model_name]
            if 'VIN_Count' in sub.columns:
                return int(sub['VIN_Count'].sum())
            return len(sub)
            
        if paint_summary_dict:
            # Extract numbers directly from PPC_Float_Report_Paint_...
            # TCF1 Line
            for model in tcf1_models:
                m_dict = paint_summary_dict.get(model, {})
                today_vin = get_today_vin_count(tcf1_drops, model)
                row_data = {
                    'Paint Float': 'TCF1',
                    'MODEL': model,
                    'Today VIN': today_vin
                }
                for stage in stages_list:
                    row_data[stage] = m_dict.get(stage, 0)
                row_data['TOTAL UPTO SEALANT'] = m_dict.get('TOTAL UPTO SEALANT', (
                    row_data['PBS FLOAT'] + 
                    row_data['PBS TO POLISHING'] + 
                    row_data['POLISHING TO TOPCOAT'] + 
                    row_data['TOPCOAT TO WETSANDING G ROOFBLACK'] + 
                    row_data['TOPCOAT TO WETSANDING G FRESH'] + 
                    row_data['WETSANDING G TO SEALANT']
                ))
                row_data['TOTAL FLOAT'] = m_dict.get('TOTAL FLOAT', row_data['TOTAL UPTO SEALANT'] + row_data.get('PT ENTRY TO SEALANT', 0) + row_data.get('BIW LIFTING G TO PT', 0) + row_data.get('PT BYPASS', 0))
                rows.append(row_data)

            # TCF1 TOTAL
            tcf1_subtotal = {
                'Paint Float': 'TCF1',
                'MODEL': 'TCF1 TOTAL',
                'Today VIN': sum(r['Today VIN'] for r in rows if r['Paint Float'] == 'TCF1')
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                tcf1_subtotal[col] = sum(r[col] for r in rows if r['Paint Float'] == 'TCF1')
            rows.append(tcf1_subtotal)

            # TCF2 Line
            tcf2_rows_start_idx = len(rows)
            for model in tcf2_models:
                m_dict = paint_summary_dict.get(model, {})
                today_vin = get_today_vin_count(tcf2_drops, model)
                row_data = {
                    'Paint Float': 'TCF2',
                    'MODEL': model,
                    'Today VIN': today_vin
                }
                for stage in stages_list:
                    row_data[stage] = m_dict.get(stage, 0)
                row_data['TOTAL UPTO SEALANT'] = m_dict.get('TOTAL UPTO SEALANT', (
                    row_data['PBS FLOAT'] + 
                    row_data['PBS TO POLISHING'] + 
                    row_data['POLISHING TO TOPCOAT'] + 
                    row_data['TOPCOAT TO WETSANDING G ROOFBLACK'] + 
                    row_data['TOPCOAT TO WETSANDING G FRESH'] + 
                    row_data['WETSANDING G TO SEALANT']
                ))
                row_data['TOTAL FLOAT'] = m_dict.get('TOTAL FLOAT', row_data['TOTAL UPTO SEALANT'] + row_data.get('PT ENTRY TO SEALANT', 0) + row_data.get('BIW LIFTING G TO PT', 0) + row_data.get('PT BYPASS', 0))
                rows.append(row_data)

            # TCF2 TOTAL
            tcf2_subtotal = {
                'Paint Float': 'TCF2',
                'MODEL': 'TCF2 TOTAL',
                'Today VIN': sum(r['Today VIN'] for r in rows[tcf2_rows_start_idx:] if r['Paint Float'] == 'TCF2')
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                tcf2_subtotal[col] = sum(r[col] for r in rows[tcf2_rows_start_idx:] if r['Paint Float'] == 'TCF2')
            rows.append(tcf2_subtotal)

            # GRAND TOTAL
            grand_total = {
                'Paint Float': '',
                'MODEL': 'GRAND TOTAL',
                'Today VIN': tcf1_subtotal['Today VIN'] + tcf2_subtotal['Today VIN']
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                grand_total[col] = tcf1_subtotal[col] + tcf2_subtotal[col]
            rows.append(grand_total)

            summary_df = pd.DataFrame(rows)
        elif float_df is not None and not float_df.empty:
            temp_float_df = float_df.copy()
            temp_float_df['Model_Mapped'] = temp_float_df['PRODUCT'].apply(get_summary_product_to_model)
            temp_float_df['Stage'] = temp_float_df.apply(get_row_paint_stage, axis=1)
            
            # TCF1 Line
            tcf1_sub_df = temp_float_df[temp_float_df['SHOP'] == 'TCF1']
            for model in tcf1_models:
                model_df = tcf1_sub_df[tcf1_sub_df['Model_Mapped'] == model]
                today_vin = get_today_vin_count(tcf1_drops, model)
                
                row_data = {
                    'Paint Float': 'TCF1',
                    'MODEL': model,
                    'Today VIN': today_vin
                }
                
                total_float = 0
                for stage in stages_list:
                    cnt = len(model_df[model_df['Stage'] == stage])
                    row_data[stage] = cnt
                    total_float += cnt
                    
                row_data['TOTAL FLOAT'] = total_float
                row_data['TOTAL UPTO SEALANT'] = (
                    row_data['PBS FLOAT'] + 
                    row_data['PBS TO POLISHING'] + 
                    row_data['POLISHING TO TOPCOAT'] + 
                    row_data['TOPCOAT TO WETSANDING G ROOFBLACK'] + 
                    row_data['TOPCOAT TO WETSANDING G FRESH'] + 
                    row_data['WETSANDING G TO SEALANT']
                )
                rows.append(row_data)
                
            # TCF1 TOTAL
            tcf1_subtotal = {
                'Paint Float': 'TCF1',
                'MODEL': 'TCF1 TOTAL',
                'Today VIN': sum(r['Today VIN'] for r in rows if r['Paint Float'] == 'TCF1')
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                tcf1_subtotal[col] = sum(r[col] for r in rows if r['Paint Float'] == 'TCF1')
            rows.append(tcf1_subtotal)
            
            # TCF2 Line
            tcf2_sub_df = temp_float_df[temp_float_df['SHOP'] == 'TCF2']
            tcf2_rows_start_idx = len(rows)
            for model in tcf2_models:
                model_df = tcf2_sub_df[tcf2_sub_df['Model_Mapped'] == model]
                today_vin = get_today_vin_count(tcf2_drops, model)
                
                row_data = {
                    'Paint Float': 'TCF2',
                    'MODEL': model,
                    'Today VIN': today_vin
                }
                
                total_float = 0
                for stage in stages_list:
                    cnt = len(model_df[model_df['Stage'] == stage])
                    row_data[stage] = cnt
                    total_float += cnt
                    
                row_data['TOTAL FLOAT'] = total_float
                row_data['TOTAL UPTO SEALANT'] = (
                    row_data['PBS FLOAT'] + 
                    row_data['PBS TO POLISHING'] + 
                    row_data['POLISHING TO TOPCOAT'] + 
                    row_data['TOPCOAT TO WETSANDING G ROOFBLACK'] + 
                    row_data['TOPCOAT TO WETSANDING G FRESH'] + 
                    row_data['WETSANDING G TO SEALANT']
                )
                rows.append(row_data)
                
            # TCF2 TOTAL
            tcf2_subtotal = {
                'Paint Float': 'TCF2',
                'MODEL': 'TCF2 TOTAL',
                'Today VIN': sum(r['Today VIN'] for r in rows[tcf2_rows_start_idx:] if r['Paint Float'] == 'TCF2')
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                tcf2_subtotal[col] = sum(r[col] for r in rows[tcf2_rows_start_idx:] if r['Paint Float'] == 'TCF2')
            rows.append(tcf2_subtotal)
            
            # GRAND TOTAL
            grand_total = {
                'Paint Float': '',
                'MODEL': 'GRAND TOTAL',
                'Today VIN': tcf1_subtotal['Today VIN'] + tcf2_subtotal['Today VIN']
            }
            for col in ['TOTAL FLOAT'] + stages_list + ['TOTAL UPTO SEALANT']:
                grand_total[col] = tcf1_subtotal[col] + tcf2_subtotal[col]
            rows.append(grand_total)
            
            summary_df = pd.DataFrame(rows)
        else:
            summary_df = pd.DataFrame()
        
        # Rename columns to match user copy perfectly
        display_col_mapping = {
            'Paint Float': 'Paint Float',
            'MODEL': 'MODEL',
            'TOTAL FLOAT': 'TOTAL FLOAT',
            'PBS FLOAT': 'PBS FLOAT',
            'PBS TO POLISHING': 'PBS TO POLISHING',
            'POLISHING TO TOPCOAT': 'POLISHING TO TOPCOAT',
            'TOPCOAT TO WETSANDING G ROOFBLACK': 'TOPCOAT TO WETSANDING G ROOFBLACK',
            'TOPCOAT TO WETSANDING G FRESH': 'TOPCOAT TO WETSANDING G FRESH',
            'WETSANDING G TO SEALANT': 'WETSANDING G TO SEALANT',
            'TOTAL UPTO SEALANT': 'TOTAL UPTO SEALANT',
            'PT ENTRY TO SEALANT': 'PT ENTRY TO SEALANT',
            'BIW LIFTING G TO PT': 'BIW LIFTING G TO PT',
            'PT BYPASS': 'PT BYPASS',
            'Today VIN': 'Today VIN'
        }
        
        summary_df = summary_df[[
            'Paint Float', 'MODEL', 'TOTAL FLOAT', 'PBS FLOAT', 'PBS TO POLISHING',
            'POLISHING TO TOPCOAT', 'TOPCOAT TO WETSANDING G ROOFBLACK',
            'TOPCOAT TO WETSANDING G FRESH', 'WETSANDING G TO SEALANT',
            'TOTAL UPTO SEALANT', 'PT ENTRY TO SEALANT', 'BIW LIFTING G TO PT',
            'PT BYPASS', 'Today VIN'
        ]].rename(columns=display_col_mapping)
        
        # Helper to generate beautiful wrapped HTML table for summary float report
        is_dark_theme = st.session_state.get('theme', '☀️ White Theme') == '🌙 Dark Theme'
        
        def render_html_float_summary(df, is_dark):
            th_bg = "#1F2937" if is_dark else "#F3F4F6"
            th_text = "#FAFAFA" if is_dark else "#374151"
            td_border = "#30363D" if is_dark else "#E5E7EB"
            text_color = "#FAFAFA" if is_dark else "#111827"
            
            html = f"""
            <div style="overflow-x: auto; border: 1px solid {td_border}; border-radius: 12px; margin-bottom: 2rem; background-color: {'#161B22' if is_dark else '#FFFFFF'}; box-shadow: 0 4px 12px rgba(0,0,0,0.03);">
            <table style="width: 100%; border-collapse: collapse; font-family: 'Inter', sans-serif; font-size: 12px; color: {text_color};">
                <thead>
                    <tr style="background-color: {th_bg}; border-bottom: 2px solid {td_border};">
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: left; color: {th_text}; font-weight: 600;">Paint Float</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: left; color: {th_text}; font-weight: 600; width: 110px;">MODEL</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 60px; word-wrap: break-word; white-space: normal;">TOTAL FLOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 60px; word-wrap: break-word; white-space: normal;">PBS FLOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; word-wrap: break-word; white-space: normal;">PBS TO POLISHING</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; word-wrap: break-word; white-space: normal;">POLISHING TO TOPCOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 100px; max-width: 120px; word-wrap: break-word; white-space: normal;">TOPCOAT TO WETSANDING G ROOFBLACK</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 100px; max-width: 120px; word-wrap: break-word; white-space: normal;">TOPCOAT TO WETSANDING G FRESH</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 100px; max-width: 120px; word-wrap: break-word; white-space: normal;">WETSANDING G TO SEALANT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; max-width: 100px; word-wrap: break-word; white-space: normal;">TOTAL UPTO SEALANT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; max-width: 100px; word-wrap: break-word; white-space: normal;">PT ENTRY TO SEALANT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; max-width: 100px; word-wrap: break-word; white-space: normal;">BIW LIFTING G TO PT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 60px; word-wrap: break-word; white-space: normal;">PT BYPASS</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 60px; word-wrap: break-word; white-space: normal;">Today VIN</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for idx_r, row_r in df.iterrows():
                model_val = str(row_r.get('MODEL', '')).strip()
                
                row_bg = "transparent"
                row_text = text_color
                font_weight = "normal"
                
                if 'TOTAL' in model_val and 'GRAND' not in model_val:
                    row_bg = "#3b1f3c" if is_dark else "#f2dcdb"
                    row_text = "#f2dcdb" if is_dark else "#5c1d1b"
                    font_weight = "bold"
                elif 'GRAND TOTAL' in model_val:
                    row_bg = "#4a3f00" if is_dark else "#ffffc5"
                    row_text = "#ffff00" if is_dark else "#806000"
                    font_weight = "bold"
                    
                html += f'<tr style="background-color: {row_bg}; color: {row_text}; font-weight: {font_weight}; border-bottom: 1px solid {td_border};">'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: left;">{row_r.get("Paint Float", "")}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: left;">{row_r.get("MODEL", "")}</td>'
                
                for col in ['TOTAL FLOAT', 'PBS FLOAT', 'PBS TO POLISHING', 'POLISHING TO TOPCOAT', 
                            'TOPCOAT TO WETSANDING G ROOFBLACK', 'TOPCOAT TO WETSANDING G FRESH', 
                            'WETSANDING G TO SEALANT', 'TOTAL UPTO SEALANT', 'PT ENTRY TO SEALANT', 
                            'BIW LIFTING G TO PT', 'PT BYPASS', 'Today VIN']:
                    val = row_r.get(col, 0)
                    val_str = str(val) if pd.notna(val) else "0"
                    html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_str}</td>'
                html += '</tr>'
                
            html += """
                </tbody>
            </table>
            </div>
            """
            return html
            
        # Display the first table (wrapped nicely in HTML)
        st.markdown(render_html_float_summary(summary_df, is_dark_theme), unsafe_allow_html=True)
        
        # ----------------- ENGINE & BATTERY REQUIREMENT SUMMARY REPORT -----------------
        st.markdown("<div style='height: 1.5rem;'></div>", unsafe_allow_html=True)
        st.markdown("### 📊 Engine & Battery Requirement Summary")
        st.markdown("""
            This report summarizes raw engine inventory status against paint shop float and computes clear-to-build requirements.
        """)
        
        # Compute dictionary values for engines
        engine_stocks_dict = {}
        engine_ta_dict = {}
        if 'engine_df' in st.session_state and st.session_state.engine_df is not None:
            for idx, r_eng in st.session_state.engine_df.iterrows():
                p_no = str(r_eng['Engine Part No']).strip()
                engine_stocks_dict[p_no] = int(r_eng.get('Clearance After 6:30AM', 0))
                engine_ta_dict[p_no] = str(r_eng.get('TA Code', '—')).strip()
                
        # Create mapping dictionary from Short Vehicle Code -> Engine
        vc_to_engine = {}
        if bom_df is not None and not bom_df.empty:
            vc_to_engine = dict(zip(bom_df['Short Vehicle Code'].astype(str).str.strip(), bom_df['Engine'].astype(str).str.strip()))
            
        # Today VIN per engine part
        today_vin_dict = {}
        for vgl_df in [tcf1_drops, tcf2_drops]:
            if vgl_df is not None and not vgl_df.empty:
                vc_col = 'VEHICLE CODE' if 'VEHICLE CODE' in vgl_df.columns else ('VC' if 'VC' in vgl_df.columns else None)
                if vc_col:
                    vgl_df['Engine_Part'] = vgl_df[vc_col].astype(str).str.strip().str[:9].map(vc_to_engine)
                    for part in vgl_df['Engine_Part'].dropna().unique():
                        p_str = str(part).strip()
                        if p_str in ['None', 'nan', '0', '']:
                            continue
                        sub = vgl_df[vgl_df['Engine_Part'] == part]
                        cnt = int(sub['VIN_Count'].sum()) if 'VIN_Count' in sub.columns else len(sub)
                        today_vin_dict[p_str] = today_vin_dict.get(p_str, 0) + cnt
                
        # Float demands per engine part
        pbs_float_dict = {}
        upto_sealant_dict = {}
        total_float_dict = {}
        
        stages_upto_sealant = [
            'PBS FLOAT', 
            'PBS TO POLISHING', 
            'POLISHING TO TOPCOAT', 
            'TOPCOAT TO WETSANDING G ROOFBLACK', 
            'TOPCOAT TO WETSANDING G FRESH', 
            'WETSANDING G TO SEALANT'
        ]
        
        if paint_summary_vc_dict:
            for svc, counts in paint_summary_vc_dict.items():
                p_str = vc_to_engine.get(svc)
                if p_str and str(p_str).strip() not in ['None', 'nan', '0']:
                    p_clean = str(p_str).strip()
                    total_float_dict[p_clean] = total_float_dict.get(p_clean, 0) + counts['TOTAL FLOAT']
                    pbs_float_dict[p_clean] = pbs_float_dict.get(p_clean, 0) + counts['PBS FLOAT']
                    upto_sealant_dict[p_clean] = upto_sealant_dict.get(p_clean, 0) + counts['TOTAL UPTO SEALANT']
        elif temp_float_df is not None and not temp_float_df.empty:
            vc_col = 'VEHICLE CODE' if 'VEHICLE CODE' in temp_float_df.columns else ('VC' if 'VC' in temp_float_df.columns else None)
            if vc_col:
                temp_float_df['Engine_Part'] = temp_float_df[vc_col].astype(str).str.strip().str[:9].map(vc_to_engine)
                for idx, row_f in temp_float_df.iterrows():
                    part = row_f.get('Engine_Part')
                    if pd.isna(part):
                        continue
                    p_str = str(part).strip()
                    stage = row_f.get('Stage', '')
                    
                    total_float_dict[p_str] = total_float_dict.get(p_str, 0) + 1
                    if stage == 'PBS FLOAT':
                        pbs_float_dict[p_str] = pbs_float_dict.get(p_str, 0) + 1
                    if stage in stages_upto_sealant:
                        upto_sealant_dict[p_str] = upto_sealant_dict.get(p_str, 0) + 1
                
        # Build TCF1 rows
        punch_parts = [
            ("54850000PTP001", "Punch MT SA"),
            ("54850000PTP002", "Punch AMT SA"),
            ("54970000PTP002", "Punch TC MCE"),
            ("54970000PTP003", "Punch MCE MT"),
            ("54970000PTP004", "Punch MCE AMT"),
            ("54970000PTP005", "Punch MCE CNG MT"),
            ("54970000PTP031", "Punch MCE CNG AMT")
        ]
        
        table2_rows = []
        for part, model in punch_parts:
            clearance = engine_stocks_dict.get(part, 0)
            today_vin = today_vin_dict.get(part, 0)
            bal = clearance - today_vin
            pbs = pbs_float_dict.get(part, 0)
            sealant = upto_sealant_dict.get(part, 0)
            total = total_float_dict.get(part, 0)
            table2_rows.append({
                'Engine Part No': part,
                'Model': model,
                'TA Code': engine_ta_dict.get(part, '—'),
                'Clearance After 6:30AM': clearance,
                'Today VIN': today_vin,
                'Bal': bal,
                'PBS FLOAT': pbs,
                'Float UPTO SEALANT': sealant,
                'TOTAL FLOAT': total,
                'With respect to PBS FLOAT': bal - pbs,
                'With respect to Sealant FLOAT': bal - sealant,
                'With respect to Total FLOAT': bal - total,
                'Type': 'row'
            })
            
        subtotal_1_2 = {
            'Engine Part No': '',
            'Model': '1.2 Lit Total',
            'TA Code': '',
            'Clearance After 6:30AM': '',
            'Today VIN': sum(r['Today VIN'] for r in table2_rows),
            'Bal': '',
            'PBS FLOAT': sum(r['PBS FLOAT'] for r in table2_rows),
            'Float UPTO SEALANT': sum(r['Float UPTO SEALANT'] for r in table2_rows),
            'TOTAL FLOAT': sum(r['TOTAL FLOAT'] for r in table2_rows),
            'With respect to PBS FLOAT': '',
            'With respect to Sealant FLOAT': '',
            'With respect to Total FLOAT': '',
            'Type': 'subtotal'
        }
        table2_rows.append(subtotal_1_2)
        
        # Nova
        part_nova = "546816111212"
        model_nova = "Nova"
        if 'nova_materials_df' in st.session_state and st.session_state.nova_materials_df is not None and not st.session_state.nova_materials_df.empty:
            clearance_nova = int(st.session_state.nova_materials_df['Clearance Qty'].min())
        else:
            clearance_nova = 182
        today_vin_nova = today_vin_dict.get(part_nova, 0)
        bal_nova = clearance_nova - today_vin_nova
        pbs_nova = pbs_float_dict.get(part_nova, 0)
        sealant_nova = upto_sealant_dict.get(part_nova, 0)
        total_nova = total_float_dict.get(part_nova, 0)
        row_nova = {
            'Engine Part No': part_nova,
            'Model': model_nova,
            'TA Code': engine_ta_dict.get(part_nova, '5468'),
            'Clearance After 6:30AM': clearance_nova,
            'Today VIN': today_vin_nova,
            'Bal': bal_nova,
            'PBS FLOAT': pbs_nova,
            'Float UPTO SEALANT': sealant_nova,
            'TOTAL FLOAT': total_nova,
            'With respect to PBS FLOAT': bal_nova - pbs_nova,
            'With respect to Sealant FLOAT': bal_nova - sealant_nova,
            'With respect to Total FLOAT': bal_nova - total_nova,
            'Type': 'row'
        }
        table2_rows.append(row_nova)

        # TCF1 Grand Total
        tcf1_grand = {
            'Engine Part No': '',
            'Model': 'TCF1',
            'TA Code': '',
            'Clearance After 6:30AM': '',
            'Today VIN': subtotal_1_2['Today VIN'] + row_nova['Today VIN'],
            'Bal': '',
            'PBS FLOAT': subtotal_1_2['PBS FLOAT'] + row_nova['PBS FLOAT'],
            'Float UPTO SEALANT': subtotal_1_2['Float UPTO SEALANT'] + row_nova['Float UPTO SEALANT'],
            'TOTAL FLOAT': subtotal_1_2['TOTAL FLOAT'] + row_nova['TOTAL FLOAT'],
            'With respect to PBS FLOAT': '',
            'With respect to Sealant FLOAT': '',
            'With respect to Total FLOAT': '',
            'Type': 'total'
        }
        table2_rows.append(tcf1_grand)
        
        # Build TCF2 rows
        tcf2_parts = [
            ("572900000118", "Harrier / Safari Diesel AT"),
            ("572900000120", "Harrier / Safari Diesel MT"),
            ("54780000PTP001", "Harrier / Safari Petrol TGDI MT"),
            ("54780000PTP002", "Harrier / Safari Petrol TGDI AT")
        ]
        
        tcf2_start_idx = len(table2_rows)
        for part, model in tcf2_parts:
            clearance = engine_stocks_dict.get(part, 0)
            today_vin = today_vin_dict.get(part, 0)
            bal = clearance - today_vin
            pbs = pbs_float_dict.get(part, 0)
            sealant = upto_sealant_dict.get(part, 0)
            total = total_float_dict.get(part, 0)
            table2_rows.append({
                'Engine Part No': part,
                'Model': model,
                'TA Code': engine_ta_dict.get(part, '—'),
                'Clearance After 6:30AM': clearance,
                'Today VIN': today_vin,
                'Bal': bal,
                'PBS FLOAT': pbs,
                'Float UPTO SEALANT': sealant,
                'TOTAL FLOAT': total,
                'With respect to PBS FLOAT': bal - pbs,
                'With respect to Sealant FLOAT': bal - sealant,
                'With respect to Total FLOAT': bal - total,
                'Type': 'row'
            })
            
        subtotal_2_0 = {
            'Engine Part No': '',
            'Model': '2 Lit Total',
            'TA Code': '',
            'Clearance After 6:30AM': '',
            'Today VIN': sum(r['Today VIN'] for r in table2_rows[tcf2_start_idx:]),
            'Bal': '',
            'PBS FLOAT': sum(r['PBS FLOAT'] for r in table2_rows[tcf2_start_idx:]),
            'Float UPTO SEALANT': sum(r['Float UPTO SEALANT'] for r in table2_rows[tcf2_start_idx:]),
            'TOTAL FLOAT': sum(r['TOTAL FLOAT'] for r in table2_rows[tcf2_start_idx:]),
            'With respect to PBS FLOAT': '',
            'With respect to Sealant FLOAT': '',
            'With respect to Total FLOAT': '',
            'Type': 'subtotal'
        }
        table2_rows.append(subtotal_2_0)
        
        # Harrier EV
        part_hev = "547380400103"
        model_hev = "Harrier EV"
        clearance_hev = 160
        today_vin_hev = today_vin_dict.get(part_hev, 0)
        bal_hev = clearance_hev - today_vin_hev
        pbs_hev = pbs_float_dict.get(part_hev, 0)
        sealant_hev = upto_sealant_dict.get(part_hev, 0)
        total_hev = total_float_dict.get(part_hev, 0)
        row_hev = {
            'Engine Part No': part_hev,
            'Model': model_hev,
            'TA Code': engine_ta_dict.get(part_hev, '5473'),
            'Clearance After 6:30AM': clearance_hev,
            'Today VIN': today_vin_hev,
            'Bal': bal_hev,
            'PBS FLOAT': pbs_hev,
            'Float UPTO SEALANT': sealant_hev,
            'TOTAL FLOAT': total_hev,
            'With respect to PBS FLOAT': bal_hev - pbs_hev,
            'With respect to Sealant FLOAT': bal_hev - sealant_hev,
            'With respect to Total FLOAT': bal_hev - total_hev,
            'Type': 'row'
        }
        table2_rows.append(row_hev)
        
        # TCF2 Grand Total
        tcf2_grand = {
            'Engine Part No': '',
            'Model': 'TCF2',
            'TA Code': '',
            'Clearance After 6:30AM': '',
            'Today VIN': subtotal_2_0['Today VIN'] + row_hev['Today VIN'],
            'Bal': '',
            'PBS FLOAT': subtotal_2_0['PBS FLOAT'] + row_hev['PBS FLOAT'],
            'Float UPTO SEALANT': subtotal_2_0['Float UPTO SEALANT'] + row_hev['Float UPTO SEALANT'],
            'TOTAL FLOAT': subtotal_2_0['TOTAL FLOAT'] + row_hev['TOTAL FLOAT'],
            'With respect to PBS FLOAT': '',
            'With respect to Sealant FLOAT': '',
            'With respect to Total FLOAT': '',
            'Type': 'total'
        }
        table2_rows.append(tcf2_grand)
        
        # Render Table 2 in beautiful HTML with rowspan/colspan
        def render_html_table_2(rows, is_dark):
            th_bg = "#1F2937" if is_dark else "#F3F4F6"
            th_text = "#FAFAFA" if is_dark else "#374151"
            td_border = "#30363D" if is_dark else "#E5E7EB"
            text_color = "#FAFAFA" if is_dark else "#111827"
            
            clearance_bg = "#1b4d32" if is_dark else "#d8f3e5"
            clearance_text = "#FAFAFA" if is_dark else "#1b4d32"
            
            bal_bg = "#4a274c" if is_dark else "#f2dcdb"
            bal_text = "#FAFAFA" if is_dark else "#5c1d1b"
            
            alert_bg = "#5c1d1d" if is_dark else "#ffd1d1"
            alert_text = "#FAFAFA" if is_dark else "#5c1d1d"
            
            html = f"""
            <div style="overflow-x: auto; border: 1px solid {td_border}; border-radius: 12px; margin-bottom: 2rem; background-color: {'#161B22' if is_dark else '#FFFFFF'}; box-shadow: 0 4px 12px rgba(0,0,0,0.03);">
            <table style="width: 100%; border-collapse: collapse; font-family: 'Inter', sans-serif; font-size: 12px; color: {text_color};">
                <thead>
                    <tr style="background-color: {th_bg}; border-bottom: 1px solid {td_border};">
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; vertical-align: middle;">Engine / Battery Part No</th>
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: left; color: {th_text}; font-weight: 600; width: 180px; vertical-align: middle;">Model</th>
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; vertical-align: middle;">TA Code</th>
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 80px; white-space: normal; vertical-align: middle;">Clearance After 6:30AM</th>
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; vertical-align: middle;">Today VIN</th>
                        <th rowspan="2" style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; background-color: {bal_bg}; color: {bal_text}; vertical-align: middle;">Bal</th>
                        <th colspan="3" style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600;">Paint Float</th>
                        <th colspan="3" style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600;">Engine & Battery requirement</th>
                    </tr>
                    <tr style="background-color: {th_bg}; border-bottom: 2px solid {td_border};">
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600;">PBS FLOAT</th>
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600;">Float UPTO SEALANT</th>
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600;">TOTAL FLOAT</th>
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 90px; white-space: normal;">With respect to PBS FLOAT</th>
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 90px; white-space: normal;">With respect to Sealant FLOAT</th>
                        <th style="padding: 6px; border: 1px solid {td_border}; text-align: center; color: {th_text}; font-weight: 600; min-width: 90px; white-space: normal;">With respect to Total FLOAT</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for r_data in rows:
                r_type = r_data['Type']
                
                row_bg = "transparent"
                row_text = text_color
                font_weight = "normal"
                
                if r_type == 'subtotal':
                    row_bg = "#005b8a" if is_dark else "#00B0F0"
                    row_text = "#FFFFFF"
                    font_weight = "bold"
                elif r_type == 'total':
                    row_bg = "#7f7f00" if is_dark else "#ffff00"
                    row_text = "#FAFAFA" if is_dark else "#000000"
                    font_weight = "bold"
                    
                html += f'<tr style="background-color: {row_bg}; color: {row_text}; font-weight: {font_weight}; border-bottom: 1px solid {td_border};">'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["Engine Part No"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: left;">{r_data["Model"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["TA Code"]}</td>'
                
                val_clearance = r_data["Clearance After 6:30AM"]
                if val_clearance != "" and r_type == 'row':
                    html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center; background-color: {clearance_bg}; color: {clearance_text}; font-weight: bold;">{val_clearance}</td>'
                else:
                    html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_clearance}</td>'
                    
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["Today VIN"]}</td>'
                
                val_bal = r_data["Bal"]
                if val_bal != "" and r_type == 'row':
                    html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center; background-color: {bal_bg}; color: {bal_text}; font-weight: bold;">{val_bal}</td>'
                else:
                    html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_bal}</td>'
                    
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["PBS FLOAT"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["Float UPTO SEALANT"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r_data["TOTAL FLOAT"]}</td>'
                
                for col_k in ["With respect to PBS FLOAT", "With respect to Sealant FLOAT", "With respect to Total FLOAT"]:
                    val_req = r_data[col_k]
                    if val_req != "" and r_type == 'row':
                        if isinstance(val_req, (int, float)) and val_req < 0:
                            html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center; background-color: {alert_bg}; color: {alert_text}; font-weight: bold;">{val_req}</td>'
                        else:
                            html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_req}</td>'
                    else:
                        html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_req}</td>'
                        
                html += '</tr>'
            html += """
                </tbody>
            </table>
            </div>
            """
            return html
            
        st.markdown(render_html_table_2(table2_rows, is_dark_theme), unsafe_allow_html=True)
        
        # ----------------- COCKPIT & WIRING SHORTAGE REPORTS (EXACT USER FORMAT) -----------------
        st.markdown("---")
        st.markdown("### 🧩 Cockpit & Wiring Shortage Reports")
        st.markdown("""
            These reports track Cockpit and Wiring Harness shortage requirements in the exact layout matching engine summary sheet models.
        """)
        
        # Helper to build shortage table for Cockpit or Front Wiring
        def build_formatted_shortage_table(part_col_name, stock_tcf1, stock_tcf2, bom_df, float_df, paint_summary_vc_dict, tcf1_drops, tcf2_drops, only_shortage=True):
            if bom_df is None or bom_df.empty:
                return pd.DataFrame()
                
            local_engine_to_model = {}
            local_engine_to_line = {}
            if 'engine_df' in st.session_state and not st.session_state.engine_df.empty:
                local_engine_to_model = dict(zip(st.session_state.engine_df['Engine Part No'].astype(str).str.strip(), st.session_state.engine_df['Model']))
                local_engine_to_line = dict(zip(st.session_state.engine_df['Engine Part No'].astype(str).str.strip(), st.session_state.engine_df['TCF Line']))
            else:
                local_engine_to_model = {item['Engine Part No']: item['Model'] for item in engine_default_data}
                local_engine_to_line = {item['Engine Part No']: item['TCF Line'] for item in engine_default_data}

            local_engine_to_model['546816111212'] = 'Punch EV (Nova)'
            local_engine_to_line['546816111212'] = 'TCF1'
            local_engine_to_model['547380400103'] = 'Harrier EV'
            local_engine_to_line['547380400103'] = 'TCF2'

            vc_to_part = dict(zip(bom_df['Short Vehicle Code'].astype(str).str.strip(), bom_df[part_col_name].astype(str).str.strip()))
            
            part_to_models = {}
            part_to_line = {}
            
            for idx, row in bom_df.iterrows():
                eng = str(row.get('Engine', '')).strip()
                part = str(row.get(part_col_name, '')).strip()
                mdl = local_engine_to_model.get(eng, '')
                line = local_engine_to_line.get(eng, '')
                if mdl and part and part not in ['None', 'nan', '0']:
                    part_to_models.setdefault(part, set()).add(mdl)
                    if line:
                        part_to_line[part] = line
                        
            pbs_dict = {}
            sealant_dict = {}
            total_dict = {}
            stages_upto_sealant = ['PBS FLOAT', 'PBS TO POLISHING', 'POLISHING TO TOPCOAT', 'TOPCOAT TO WETSANDING G ROOFBLACK', 'TOPCOAT TO WETSANDING G FRESH', 'WETSANDING G TO SEALANT']
            
            if paint_summary_vc_dict:
                for svc, counts in paint_summary_vc_dict.items():
                    p = vc_to_part.get(svc)
                    if p and str(p).strip() not in ['None', 'nan', '0']:
                        p_clean = str(p).strip()
                        total_dict[p_clean] = total_dict.get(p_clean, 0) + counts['TOTAL FLOAT']
                        pbs_dict[p_clean] = pbs_dict.get(p_clean, 0) + counts['PBS FLOAT']
                        sealant_dict[p_clean] = sealant_dict.get(p_clean, 0) + counts['TOTAL UPTO SEALANT']
            elif float_df is not None and not float_df.empty:
                vc_col = 'VEHICLE CODE' if 'VEHICLE CODE' in float_df.columns else ('VC' if 'VC' in float_df.columns else None)
                if vc_col:
                    for idx, row in float_df.iterrows():
                        vc = str(row.get(vc_col, '')).strip()[:9]
                        p = vc_to_part.get(vc)
                        if p and str(p).strip() not in ['None', 'nan', '0']:
                            p_clean = str(p).strip()
                            total_dict[p_clean] = total_dict.get(p_clean, 0) + 1
                            stg = row.get('Stage', '')
                            if stg == 'PBS FLOAT':
                                pbs_dict[p_clean] = pbs_dict.get(p_clean, 0) + 1
                            if stg in stages_upto_sealant:
                                sealant_dict[p_clean] = sealant_dict.get(p_clean, 0) + 1
                            
            today_vin_dict = {}
            for vgl_df in [tcf1_drops, tcf2_drops]:
                if vgl_df is not None and not vgl_df.empty:
                    vc_col = 'VEHICLE CODE' if 'VEHICLE CODE' in vgl_df.columns else ('VC' if 'VC' in vgl_df.columns else None)
                    if vc_col:
                        mapped_parts = vgl_df[vc_col].astype(str).str.strip().str[:9].map(vc_to_part)
                        for idx, part in mapped_parts.items():
                            if pd.notna(part) and str(part).strip() not in ['None', 'nan', '0', '']:
                                p_str = str(part).strip()
                                cnt = int(vgl_df.loc[idx, 'VIN_Count']) if 'VIN_Count' in vgl_df.columns and pd.notna(vgl_df.loc[idx, 'VIN_Count']) else 1
                                today_vin_dict[p_str] = today_vin_dict.get(p_str, 0) + cnt

            table_rows = []
            header_part_name = 'Cockpit Part Number' if 'Cockpit' in part_col_name else 'Wiring Part Number'
            
            stk_1 = stock_tcf1 if stock_tcf1 is not None else {}
            stk_2 = stock_tcf2 if stock_tcf2 is not None else {}

            for part, mdls in part_to_models.items():
                line = part_to_line.get(part, 'TCF1')
                stock_dict = stk_1 if line == 'TCF1' else stk_2
                if stock_dict and part not in stock_dict:
                    continue

                stock = stock_dict.get(part, 0)
                today_vin = today_vin_dict.get(part, 0)
                pbs = pbs_dict.get(part, 0)
                sealant = sealant_dict.get(part, 0)
                total = total_dict.get(part, 0)
                
                sh_pbs = stock - today_vin - pbs
                sh_sealant = stock - today_vin - sealant
                sh_total = stock - today_vin - total
                
                # Filter based on only_shortage parameter
                if not only_shortage or (sh_pbs < 0 or sh_sealant < 0 or sh_total < 0):
                    table_rows.append({
                        header_part_name: part,
                        'Model': ', '.join(sorted(mdls)),
                        'LINE': line,
                        'Clearance After 6:30AM': stock,
                        'Today VIN': today_vin,
                        'Paint TOTAL FLOAT': total,
                        'PBS FLOAT': pbs,
                        'Cabs Float UPTO SEALANT': sealant,
                        'Shortage PBS FLOAT': sh_pbs,
                        'Shortage Upto Sealant': sh_sealant,
                        'Shortage TOTAL FLOAT': sh_total
                    })
                
            return pd.DataFrame(table_rows)

        df_cpt_shortage = build_formatted_shortage_table('Cockpit', tcf1_cockpit_start, tcf2_cockpit_start, bom_df, temp_float_df, paint_summary_vc_dict, tcf1_drops, tcf2_drops, only_shortage=True)
        df_wir_shortage = build_formatted_shortage_table('Front Wiring', tcf1_wiring_start, tcf2_wiring_start, bom_df, temp_float_df, paint_summary_vc_dict, tcf1_drops, tcf2_drops, only_shortage=True)
        df_cpt_all = build_formatted_shortage_table('Cockpit', tcf1_cockpit_start, tcf2_cockpit_start, bom_df, temp_float_df, paint_summary_vc_dict, tcf1_drops, tcf2_drops, only_shortage=False)
        df_wir_all = build_formatted_shortage_table('Front Wiring', tcf1_wiring_start, tcf2_wiring_start, bom_df, temp_float_df, paint_summary_vc_dict, tcf1_drops, tcf2_drops, only_shortage=False)

        # Evaluate Today VIN excess alerts across Engine, Nova Aggregates, Cockpit, and Wiring
        excess_alerts = []
        
        # 1. Engine
        for r_eng in table2_rows:
            if r_eng.get('Type') == 'row':
                part_no = r_eng.get('Engine Part No', '')
                model_name = r_eng.get('Model', '')
                cl_val = r_eng.get('Clearance After 6:30AM')
                vin_val = r_eng.get('Today VIN', 0)
                if isinstance(cl_val, (int, float)) and vin_val > cl_val:
                    excess_alerts.append({
                        'Category': 'Engine',
                        'Model / Part': f"{model_name} ({part_no})" if part_no else model_name,
                        'Clearance 6:30 AM': cl_val,
                        'Today VIN': vin_val,
                        'Excess Qty': vin_val - cl_val
                    })
                    
        # 2. Nova Aggregates
        nova_df_check = st.session_state.get('nova_materials_df')
        if nova_df_check is not None and not nova_df_check.empty:
            vin_nova = today_vin_dict.get("546816111212", 0)
            for idx, r_n in nova_df_check.iterrows():
                m_name = r_n.get('Material', 'Aggregate')
                c_q = int(r_n.get('Clearance Qty', 0))
                if vin_nova > c_q:
                    excess_alerts.append({
                        'Category': 'Nova Aggregate',
                        'Model / Part': f"Punch EV - {m_name}",
                        'Clearance 6:30 AM': c_q,
                        'Today VIN': vin_nova,
                        'Excess Qty': vin_nova - c_q
                    })
                    
        # 3. Cockpit
        if df_cpt_all is not None and not df_cpt_all.empty:
            for idx, r_c in df_cpt_all.iterrows():
                p_hdr = 'Cockpit Part Number'
                c_no = r_c.get(p_hdr, '')
                m_descr = r_c.get('Model', '')
                cl_c = r_c.get('Clearance After 6:30AM', 0)
                vin_c = r_c.get('Today VIN', 0)
                if isinstance(cl_c, (int, float)) and vin_c > cl_c:
                    excess_alerts.append({
                        'Category': 'Cockpit',
                        'Model / Part': f"{c_no} ({m_descr})",
                        'Clearance 6:30 AM': cl_c,
                        'Today VIN': vin_c,
                        'Excess Qty': vin_c - cl_c
                    })
                    
        # 4. Wiring
        if df_wir_all is not None and not df_wir_all.empty:
            for idx, r_w in df_wir_all.iterrows():
                p_hdr = 'Wiring Part Number'
                w_no = r_w.get(p_hdr, '')
                m_descr = r_w.get('Model', '')
                cl_w = r_w.get('Clearance After 6:30AM', 0)
                vin_w = r_w.get('Today VIN', 0)
                if isinstance(cl_w, (int, float)) and vin_w > cl_w:
                    excess_alerts.append({
                        'Category': 'Wiring',
                        'Model / Part': f"{w_no} ({m_descr})",
                        'Clearance 6:30 AM': cl_w,
                        'Today VIN': vin_w,
                        'Excess Qty': vin_w - cl_w
                    })

        # Render alert banner if any excess exists
        if excess_alerts:
            header_title = f"🚨 ALERT: TODAY VIN GENERATION EXCEEDS CLEARANCE AFTER 6:30 AM ({len(excess_alerts)} Item{'s' if len(excess_alerts) > 1 else ''})"
            bg_card = '#2C121A' if is_dark_theme else '#FFF1F2'
            title_color = '#FDA4AF' if is_dark_theme else '#9F1239'
            sub_color = '#94A3B8' if is_dark_theme else '#64748B'
            text_col = '#FAFAFA' if is_dark_theme else '#111827'
            th_bg = '#4C1D24' if is_dark_theme else '#FFE4E6'
            border_hdr = '#881337' if is_dark_theme else '#FECDD3'
            border_td = '#5F1D28' if is_dark_theme else '#FFE4E6'
            cat_col = '#FB7185' if is_dark_theme else '#E11D48'
            vin_col = '#F43F5E' if is_dark_theme else '#BE123C'
            exc_bg = '#881337' if is_dark_theme else '#FECDD3'
            exc_col = '#FFF' if is_dark_theme else '#9F1239'
            
            rows_html = ""
            for a in excess_alerts:
                rows_html += (
                    f"<tr>"
                    f"<td style='padding: 7px 8px; border: 1px solid {border_td}; font-weight: 600; color: {cat_col};'>{a['Category']}</td>"
                    f"<td style='padding: 7px 8px; border: 1px solid {border_td};'>{a['Model / Part']}</td>"
                    f"<td style='padding: 7px 8px; border: 1px solid {border_td}; text-align: center;'>{a['Clearance 6:30 AM']}</td>"
                    f"<td style='padding: 7px 8px; border: 1px solid {border_td}; text-align: center; font-weight: bold; color: {vin_col};'>{a['Today VIN']}</td>"
                    f"<td style='padding: 7px 8px; border: 1px solid {border_td}; text-align: center; font-weight: bold; background-color: {exc_bg}; color: {exc_col};'>+{a['Excess Qty']}</td>"
                    f"</tr>"
                )
                
            alert_box_html = (
                f"<div style='border: 2px solid #E11D48; border-radius: 10px; background-color: {bg_card}; padding: 14px; margin-top: 15px; margin-bottom: 20px;'>"
                f"<div style='font-weight: 700; font-size: 15px; color: {title_color}; margin-bottom: 6px;'>{header_title}</div>"
                f"<div style='font-size: 12px; color: {sub_color}; margin-bottom: 10px;'>The following parts / aggregates have Today VIN quantity exceeding the 6:30 AM Clearance quantity:</div>"
                f"<table style='width: 100%; border-collapse: collapse; font-family: sans-serif; font-size: 12px; color: {text_col};'>"
                f"<thead>"
                f"<tr style='background-color: {th_bg}; text-align: left;'>"
                f"<th style='padding: 8px; border: 1px solid {border_hdr};'>Category</th>"
                f"<th style='padding: 8px; border: 1px solid {border_hdr};'>Model / Part Description</th>"
                f"<th style='padding: 8px; border: 1px solid {border_hdr}; text-align: center;'>Clearance After 6:30AM</th>"
                f"<th style='padding: 8px; border: 1px solid {border_hdr}; text-align: center;'>Today VIN</th>"
                f"<th style='padding: 8px; border: 1px solid {border_hdr}; text-align: center;'>Excess VIN Qty</th>"
                f"</tr>"
                f"</thead>"
                f"<tbody>{rows_html}</tbody>"
                f"</table>"
                f"</div>"
            )
            st.markdown(alert_box_html, unsafe_allow_html=True)


        def render_html_formatted_shortage(df, part_header_name, is_dark):
            if df.empty:
                return "<p style='color: #6B7280; font-style: italic;'>No data available.</p>"
                
            th_bg_orange = "#382315" if is_dark else "#FCE4D6"
            th_text_orange = "#FAFAFA" if is_dark else "#73330D"
            
            th_bg_blue = "#1A2B4C" if is_dark else "#BDD7EE"
            th_text_blue = "#FAFAFA" if is_dark else "#1A2B4C"

            th_bg_blue2 = "#1E3A5F" if is_dark else "#9BC2E6"
            
            td_border = "#30363D" if is_dark else "#E5E7EB"
            text_color = "#FAFAFA" if is_dark else "#111827"
            
            alert_bg = "#5c1d1d" if is_dark else "#FFD1D1"
            alert_text = "#FAFAFA" if is_dark else "#5C1D1B"
            
            html = f"""
            <div style="overflow-x: auto; border: 1px solid {td_border}; border-radius: 12px; margin-bottom: 2rem; background-color: {'#161B22' if is_dark else '#FFFFFF'}; box-shadow: 0 4px 12px rgba(0,0,0,0.03);">
            <table style="width: 100%; border-collapse: collapse; font-family: 'Inter', sans-serif; font-size: 12px; color: {text_color};">
                <thead>
                    <tr style="border-bottom: 2px solid {td_border};">
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_orange}; color: {th_text_orange}; font-weight: bold;">{part_header_name}</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: left; background-color: {th_bg_orange}; color: {th_text_orange}; font-weight: bold; width: 220px;">Model</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_orange}; color: {th_text_orange}; font-weight: bold;">LINE</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue}; color: {th_text_blue}; font-weight: bold;">Clearance After 6:30AM</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue}; color: {th_text_blue}; font-weight: bold;">Today VIN</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue}; color: {th_text_blue}; font-weight: bold;">Paint TOTAL FLOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue}; color: {th_text_blue}; font-weight: bold;">PBS FLOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue}; color: {th_text_blue}; font-weight: bold;">Cabs Float UPTO SEALANT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue2}; color: {th_text_blue}; font-weight: bold;">Shortage PBS FLOAT</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue2}; color: {th_text_blue}; font-weight: bold;">Shortage Upto Sealant</th>
                        <th style="padding: 10px 8px; border: 1px solid {td_border}; text-align: center; background-color: {th_bg_blue2}; color: {th_text_blue}; font-weight: bold;">Shortage TOTAL FLOAT</th>
                    </tr>
                </thead>
                <tbody>
            """
            for idx, r in df.iterrows():
                html += f'<tr style="border-bottom: 1px solid {td_border};">'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center; font-weight: 600;">{r[part_header_name]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: left;">{r["Model"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["LINE"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["Clearance After 6:30AM"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["Today VIN"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["Paint TOTAL FLOAT"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["PBS FLOAT"]}</td>'
                html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{r["Cabs Float UPTO SEALANT"]}</td>'
                
                for col_sh in ["Shortage PBS FLOAT", "Shortage Upto Sealant", "Shortage TOTAL FLOAT"]:
                    val_sh = r[col_sh]
                    if isinstance(val_sh, (int, float)) and val_sh < 0:
                        html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center; background-color: {alert_bg}; color: {alert_text}; font-weight: bold;">{val_sh}</td>'
                    else:
                        html += f'<td style="padding: 8px; border: 1px solid {td_border}; text-align: center;">{val_sh}</td>'
                        
                html += '</tr>'
            html += "</tbody></table></div>"
            return html

        st.markdown("#### 🚗 Cockpit Shortage Report")
        st.markdown(render_html_formatted_shortage(df_cpt_shortage, "Cockpit Part Number", is_dark_theme), unsafe_allow_html=True)

        st.markdown("#### ⚡ Wiring Harness Shortage Report")
        st.markdown(render_html_formatted_shortage(df_wir_shortage, "Wiring Part Number", is_dark_theme), unsafe_allow_html=True)
        
        # Excel generator with beautiful color schemes matching attached copy
        import io
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        import openpyxl.utils
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Sheet 1: Summary Report
            summary_df.to_excel(writer, index=False, sheet_name='Summary Report')
            workbook = writer.book
            worksheet = writer.sheets['Summary Report']
            
            font_header = Font(name='Calibri', size=11, bold=True, color='000000')
            fill_header = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Peach
            
            font_subtotal = Font(name='Calibri', size=11, bold=True, color='000000')
            fill_subtotal = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid') # Pink/Lavender
            
            font_grand_total = Font(name='Calibri', size=11, bold=True, color='000000')
            fill_grand_total = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Yellow
            
            font_normal = Font(name='Calibri', size=11, color='000000')
            
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF')
            )
            
            for col_idx in range(1, len(summary_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
            for row_idx in range(2, len(summary_df) + 2):
                model_val = str(worksheet.cell(row=row_idx, column=2).value).strip()
                is_subtotal = 'TOTAL' in model_val and 'GRAND' not in model_val
                is_grand = 'GRAND TOTAL' in model_val
                
                for col_idx in range(1, len(summary_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if is_subtotal:
                        cell.font = font_subtotal
                        cell.fill = fill_subtotal
                    elif is_grand:
                        cell.font = font_grand_total
                        cell.fill = fill_grand_total
                    else:
                        cell.font = font_normal
                        
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            worksheet.row_dimensions[1].height = 28
            for row_idx in range(2, len(summary_df) + 2):
                worksheet.row_dimensions[row_idx].height = 20
                
            # Sheet 2: Engine & Battery Requirement Summary
            worksheet2 = workbook.create_sheet('Engine & Battery Requirement')
            worksheet2.row_dimensions[1].height = 25
            worksheet2.row_dimensions[2].height = 25
            
            worksheet2.merge_cells('A1:A2')
            worksheet2.merge_cells('B1:B2')
            worksheet2.merge_cells('C1:C2')
            worksheet2.merge_cells('D1:D2')
            worksheet2.merge_cells('E1:E2')
            worksheet2.merge_cells('F1:F2')
            worksheet2.merge_cells('G1:I1')
            worksheet2.merge_cells('J1:L1')
            
            worksheet2['A1'] = "Engine / Battery Part No"
            worksheet2['B1'] = "Model"
            worksheet2['C1'] = "TA Code"
            worksheet2['D1'] = "Clearance After 6:30AM"
            worksheet2['E1'] = "Today VIN"
            worksheet2['F1'] = "Bal"
            worksheet2['G1'] = "Paint Float"
            worksheet2['J1'] = "Engine & Battery requirement"
            
            worksheet2['G2'] = "PBS FLOAT"
            worksheet2['H2'] = "Float UPTO SEALANT"
            worksheet2['I2'] = "TOTAL FLOAT"
            worksheet2['J2'] = "With respect to PBS FLOAT"
            worksheet2['K2'] = "With respect to Sealant FLOAT"
            worksheet2['L2'] = "With respect to Total FLOAT"
            
            for r in [1, 2]:
                for c in range(1, 13):
                    cell = worksheet2.cell(row=r, column=c)
                    cell.font = font_header
                    if r == 1 and c == 6:
                        cell.fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid') # Purple/Pink
                    else:
                        cell.fill = fill_header
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                    
            for r_idx, row_d in enumerate(table2_rows, start=3):
                worksheet2.row_dimensions[r_idx].height = 20
                row_type = row_d['Type']
                
                fill_row = None
                font_row = font_normal
                
                if row_type == 'subtotal':
                    fill_row = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') # Blue
                    font_row = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                elif row_type == 'total':
                    fill_row = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Yellow
                    font_row = Font(name='Calibri', size=11, bold=True, color='000000')
                    
                columns_list = [
                    'Engine Part No', 'Model', 'TA Code', 'Clearance After 6:30AM',
                    'Today VIN', 'Bal', 'PBS FLOAT', 'Float UPTO SEALANT', 'TOTAL FLOAT',
                    'With respect to PBS FLOAT', 'With respect to Sealant FLOAT', 'With respect to Total FLOAT'
                ]
                
                for c_idx, col_key in enumerate(columns_list, start=1):
                    cell = worksheet2.cell(row=r_idx, column=c_idx)
                    val = row_d[col_key]
                    cell.value = val
                    cell.border = thin_border
                    cell.font = font_row
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if row_type == 'row':
                        if col_key == 'Clearance After 6:30AM':
                            cell.fill = PatternFill(start_color='D8F3E5', end_color='D8F3E5', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, bold=True, color='1B4D32')
                        elif col_key == 'Bal':
                            cell.fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, bold=True, color='5C1D1B')
                        elif col_key in ['With respect to PBS FLOAT', 'With respect to Sealant FLOAT', 'With respect to Total FLOAT']:
                            if isinstance(val, (int, float)) and val < 0:
                                cell.fill = PatternFill(start_color='FFD1D1', end_color='FFD1D1', fill_type='solid')
                                cell.font = Font(name='Calibri', size=11, bold=True, color='5C1D1B')
                    elif fill_row:
                        cell.fill = fill_row
                        
            for col in worksheet2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                worksheet2.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # Function to format openpyxl sheet for Cockpit / Wiring Shortage or All Parts
            def format_openpyxl_shortage_sheet(sheet_name, df_data, part_col_hdr, target_writer=None):
                w_target = target_writer if target_writer is not None else writer
                if df_data.empty:
                    return
                df_data.to_excel(w_target, index=False, sheet_name=sheet_name)
                ws = w_target.sheets[sheet_name]
                ws.row_dimensions[1].height = 28
                
                fill_peach = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                fill_blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                fill_blue2 = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
                
                for c_i, col_name in enumerate(df_data.columns, start=1):
                    cell = ws.cell(row=1, column=c_i)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
                    if c_i <= 3:
                        cell.fill = fill_peach
                    elif c_i <= 8:
                        cell.fill = fill_blue
                    else:
                        cell.fill = fill_blue2
                        
                for r_i, r_val in enumerate(df_data.iterrows(), start=2):
                    ws.row_dimensions[r_i].height = 20
                    row_dict = r_val[1]
                    for c_i, col_name in enumerate(df_data.columns, start=1):
                        cell = ws.cell(row=r_i, column=c_i)
                        val = row_dict[col_name]
                        cell.value = val
                        cell.border = thin_border
                        cell.font = font_normal
                        cell.alignment = Alignment(horizontal='center' if col_name != 'Model' else 'left', vertical='center')
                        
                        if 'Shortage' in col_name and isinstance(val, (int, float)) and val < 0:
                            cell.fill = PatternFill(start_color='FFD1D1', end_color='FFD1D1', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, bold=True, color='5C1D1B')
                            
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            format_openpyxl_shortage_sheet('Cockpit Shortage', df_cpt_shortage, 'Cockpit Part Number')
            format_openpyxl_shortage_sheet('Wiring Shortage', df_wir_shortage, 'Wiring Part Number')
                
        excel_data = excel_buffer.getvalue()

        # Build 2-Sheet Excel workbook for Cockpit & Wiring Report (All Parts: Cockpit, Wiring)
        all_parts_excel_buffer = io.BytesIO()
        with pd.ExcelWriter(all_parts_excel_buffer, engine='openpyxl') as writer_all:
            format_openpyxl_shortage_sheet('Cockpit', df_cpt_all, 'Cockpit Part Number', target_writer=writer_all)
            format_openpyxl_shortage_sheet('Wiring', df_wir_all, 'Wiring Part Number', target_writer=writer_all)
        all_parts_excel_data = all_parts_excel_buffer.getvalue()
        
        st.markdown("---")
        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            st.download_button(
                label="📥 Export Summary Reports to Excel",
                data=excel_data,
                file_name="paint_shop_float_and_requirements_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_summary_report"
            )
        with col_exp2:
            st.download_button(
                label="📥 Download Cockpit & Wiring Report (All Parts - 2 Sheets)",
                data=all_parts_excel_data,
                file_name="Cockpit_and_Wiring_Report_All_Parts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_cockpit_wiring_all_parts"
            )

        # --- FINAL REMARK: BOM completeness check result ---
        st.markdown("---")
        if 'missing_bom_df' in locals() and missing_bom_df.empty and bom_df is not None and not bom_df.empty:
            st.success("✅ All BOM data checked — no error found. Every Short VC in the current float has a complete BOM match.")
    else:
        st.info("Please load Paint Float data in the Control Panel to view the summary report.")
